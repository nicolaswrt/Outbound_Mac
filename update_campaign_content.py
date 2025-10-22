# update_campaign_content.py
# -*- coding: utf-8 -*-
"""
Update Campaign Content:
Ordnet einer bestehenden Sonar-Campaign ein Layout-Template zu und setzt Variablen
über den Endpunkt /campaigns/{id}/updateContent.

Public:
- run_update_campaign_content(job, status_callback=None, progress_callback=None, headless=True) -> str|None
    Erwartet EINEN Job (ein Campaign-Link/ID), sendet den POST und speichert eine JSON-Logdatei.
    Rückgabe: Pfad zur Logdatei (oder None bei Fehler im Schreiben).

Optional (für Excel-Workflow):
- parse_update_excel(xlsx_path) -> list[{"campaign": str, "variables": dict}]
- build_jobs_from_excel(xlsx_path, template_path, supported_languages, use_json_variables=False) -> list[job-dicts]
"""

from __future__ import annotations

import os
import re
import json
import time
import csv
import sqlite3
import tempfile
import shutil
from typing import Callable, Dict, Any, List, Tuple, Optional

import requests

# Optionaler Selenium-Fallback zum Cookie-Refresh
from selenium import webdriver
from selenium.webdriver.firefox.service import Service as FxService
from selenium.webdriver.firefox.options import Options as FxOptions

# Firefox-Profil ermitteln (SSO-Cookies)
from utils import get_firefox_profile


# -------------------- Konfiguration --------------------

SONAR_WEB_DOMAIN = "sonar-eu.amazon.com"
SONAR_SERVICE_HOST = "sonar-service-eu-ca-dub.dub.proxy.amazon.com"
SERVICE_BASE = f"https://{SONAR_SERVICE_HOST}"

GECKODRIVER_PATH = "geckodriver.exe" if os.name == "nt" else "geckodriver"


def _copy_sqlite_readonly(src_path: str) -> Tuple[str, str]:
    """Gesperrte cookies.sqlite in ein Temp-Verzeichnis kopieren, um sie ro zu öffnen."""
    if not os.path.exists(src_path):
        raise FileNotFoundError(f"cookies.sqlite not found at: {src_path}")
    tmpdir = tempfile.mkdtemp(prefix="ff_cookies_")
    dst = os.path.join(tmpdir, "cookies.sqlite")
    shutil.copy2(src_path, dst)
    return dst, tmpdir


def _load_firefox_cookies_for_suffixes(profile_path: str, suffixes: List[str]) -> requests.cookies.RequestsCookieJar:
    """
    Lädt Cookies aus Firefox für alle Host-Suffixe in 'suffixes'.
    Nimmt zusätzlich generische .amazon.com-Cookies mit.
    """
    cookies_db = os.path.join(profile_path, "cookies.sqlite")
    cleanup_dir = None
    try:
        conn = sqlite3.connect(f"file:{cookies_db}?mode=ro", uri=True)
    except sqlite3.OperationalError:
        copied_path, cleanup_dir = _copy_sqlite_readonly(cookies_db)
        conn = sqlite3.connect(copied_path)

    jar = requests.cookies.RequestsCookieJar()
    try:
        cur = conn.cursor()
        wanted = list(suffixes) + [".amazon.com", "amazon.com"]
        seen = set()  # (name, domain, path)
        for suf in wanted:
            cur.execute(
                "SELECT name, value, host, path, isSecure FROM moz_cookies WHERE host LIKE ?",
                (f"%{suf}",)
            )
            for name, value, host, path, isSecure in cur.fetchall():
                key = (name, host, path)
                if key in seen:
                    continue
                seen.add(key)
                jar.set(name, value, domain=host, path=path, secure=bool(isSecure))
    finally:
        conn.close()
        if cleanup_dir and os.path.isdir(cleanup_dir):
            shutil.rmtree(cleanup_dir, ignore_errors=True)
    return jar


def _build_session_from_firefox(profile_path: str) -> requests.Session:
    jar = _load_firefox_cookies_for_suffixes(
        profile_path,
        [SONAR_WEB_DOMAIN, SONAR_SERVICE_HOST]
    )
    s = requests.Session()
    s.cookies = jar
    s.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:140.0) Gecko/20100101 Firefox/140.0",
        "Accept": "application/json, text/plain, */*",
        "Content-Type": "application/json;charset=utf-8",
        "Origin": f"https://{SONAR_WEB_DOMAIN}",
        "Referer": f"https://{SONAR_WEB_DOMAIN}/",
        "Connection": "keep-alive",
    })
    return s


def _cookiejar_from_selenium_cookies(cookies_list):
    jar = requests.cookies.RequestsCookieJar()
    for c in cookies_list:
        jar.set(
            c.get("name"),
            c.get("value"),
            domain=c.get("domain"),
            path=c.get("path", "/"),
            secure=bool(c.get("secure", False))
        )
    return jar


def _selenium_refresh_session_cookies(profile_path: str, headless: bool = True):
    """Öffnet Sonar-Web mit dem Profil und liefert frische Cookies zurück."""
    options = FxOptions()
    options.add_argument("-profile")
    options.add_argument(profile_path)
    if headless:
        options.add_argument("--headless")
        options.add_argument("--width=1920")
        options.add_argument("--height=1080")

    service = FxService(GECKODRIVER_PATH)
    driver = None
    try:
        driver = webdriver.Firefox(service=service, options=options)
        driver.get(f"https://{SONAR_WEB_DOMAIN}/")
        time.sleep(3)
        return _cookiejar_from_selenium_cookies(driver.get_cookies())
    finally:
        if driver is not None:
            try:
                driver.quit()
            except Exception:
                pass


# -------------------- HTTP Helpers --------------------

def _safe_parse_json(resp: requests.Response) -> dict:
    txt = resp.text or ""
    # <<< NEU: leere Antwort als Erfolg akzeptieren >>>
    if not txt.strip():
        return {}  # success without payload
    ct = (resp.headers.get("Content-Type") or "").lower()
    if "json" in ct or txt.strip().startswith(("{", "[")):
        try:
            return resp.json()
        except Exception:
            try:
                return json.loads(txt)
            except Exception:
                pass
    snippet = txt[:500].replace("\n", " ").replace("\r", " ")
    raise RuntimeError(f"HTTP {resp.status_code} non-JSON response. Snippet: {snippet!r}")



def _post_json(session: requests.Session, url: str, payload: Dict,
               timeout: Tuple[int, int] = (10, 60)) -> dict:
    resp = session.post(url, json=payload, timeout=timeout)
    # <<< NEU: 204 oder leerer Body → OK >>>
    if resp.status_code in (200, 204) and not (resp.text or "").strip():
        return {}
    if resp.status_code >= 400:
        try:
            _ = _safe_parse_json(resp)
        except Exception:
            pass
        resp.raise_for_status()
    return _safe_parse_json(resp)



def _post_json_with_refresh(session: requests.Session,
                            url: str,
                            payload: Dict,
                            profile_path: str,
                            headless: bool = True,
                            timeout: Tuple[int, int] = (10, 60)) -> dict:
    try:
        return _post_json(session, url, payload, timeout=timeout)
    except requests.HTTPError as e:
        status = getattr(e.response, "status_code", None)
        if status in (401, 403):
            fresh = _selenium_refresh_session_cookies(profile_path, headless=headless)
            if fresh:
                session.cookies = fresh
                return _post_json(session, url, payload, timeout=timeout)
        raise


# -------------------- Payload Builder --------------------

def _as_api_string(value) -> str:
    """Serialisiert API-Werte als String (bool → 'true'/'false', dict/list → JSON)."""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    if value is None:
        return ""
    return str(value)


def _norm_lang(code: str) -> str:
    """language_en_GB / en-GB / en_GB → en_GB"""
    c = str(code or "").strip().strip('"').strip("'")
    if c.startswith("language_"):
        c = c.split("language_", 1)[1]
    c = c.replace("-", "_")
    if "_" in c:
        a, b = c.split("_", 1)
        c = f"{a.lower()}_{b.upper()}"
    return c


def _build_variables(template_path: str,
                     supported_languages: List[str] | None,
                     use_json_vars: bool,
                     extra_vars: Dict[str, Any] | None) -> List[Dict[str, Any]]:
    langs = [_norm_lang(x) for x in (supported_languages or [])]
    vars_: List[Dict[str, Any]] = [
        {"name": "Correios.supported-languages",   "value": json.dumps(langs, ensure_ascii=False)},
        {"name": "Correios.managed-template-path", "value": _as_api_string(template_path)},
        {"name": "Correios.use-json-variables",    "value": "true" if use_json_vars else "false"},
    ]
    for k, v in (extra_vars or {}).items():
        vars_.append({"name": str(k), "value": _as_api_string(v)})
    return vars_


# -------------------- Excel helpers (Option A) --------------------

def _norm(s: str) -> str:
    import re as _re
    return _re.sub(r'[^a-z0-9]+', '', (s or '').strip().lower())

_CAMPAIGN_HEADERS = {
    "sonarlink", "sonar", "campaign", "campaignurl", "campaignlink", "campaignid"
}




def _read_xlsx_rows(path: str) -> List[Dict[str, Any]]:
    """Liest .xlsx in Liste von Dicts (Keys = originale Header)."""
    rows: List[Dict[str, Any]] = []
    # 1) pandas
    try:
        import pandas as pd
        df = pd.read_excel(path)
        for _, r in df.iterrows():
            d = {}
            for k, v in r.to_dict().items():
                key = str(k) if k is not None else ""
                if hasattr(pd, "isna") and pd.isna(v):
                    d[key] = ""
                else:
                    d[key] = v
            rows.append(d)
        return rows
    except Exception:
        pass
    # 2) openpyxl
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename=path, read_only=True, data_only=True)
        ws = wb.active
        headers: List[str] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(c) if c is not None else "" for c in row]
                continue
            d: Dict[str, Any] = {}
            for j, val in enumerate(row):
                key = headers[j] if j < len(headers) else f"col{j+1}"
                d[key] = "" if val is None else val
            rows.append(d)
        wb.close()
        return rows
    except Exception as e:
        raise RuntimeError(
            f"Could not read Excel. Please install 'pandas' or 'openpyxl'.\n\n{e}"
        )


def parse_update_excel(path: str) -> List[Dict[str, Any]]:
    """
    Parsed eine Update-Excel in Zeilen:
      Rückgabe: [{"campaign": "<URL/ID>", "variables": {name: value, ...}}, ...]
    Pflicht: Spalte 'Sonar Link' (exakt) ODER engeres Synonym (s. _CAMPAIGN_HEADERS).
    Alle Spalten, deren normalisierter Header in _ALLOWED_VARS liegt, werden als Variablen übernommen.
    """
    rows = _read_xlsx_rows(path)
    if not rows:
        raise ValueError("Excel is empty.")

    headers = list(rows[0].keys())
    norm_headers = [_norm(h) for h in headers]

    # Kampagnen-Spalte: exakter 'Sonar Link' bevorzugt
    campaign_idx: Optional[int] = None
    for i, h in enumerate(headers):
        if h.strip().lower() == "sonar link":
            campaign_idx = i
            break
    if campaign_idx is None:
        for i, nh in enumerate(norm_headers):
            if nh in _CAMPAIGN_HEADERS:
                campaign_idx = i
                break
    if campaign_idx is None:
        raise ValueError(
            "No campaign column found. Please use 'Sonar Link' (or Campaign/Campaign URL/Campaign ID/Sonar)."
        )

    out: List[Dict[str, Any]] = []
    for row in rows:
        values = [row.get(h, "") for h in headers]
        camp_val = values[campaign_idx]
        if camp_val is None or str(camp_val).strip() == "":
            raise ValueError("Row without campaign value (Sonar Link / Campaign / Campaign ID).")

        vars_map: Dict[str, Any] = {}
        for i, h in enumerate(headers):
            if i == campaign_idx:
                continue
            col_name = str(h).strip()
            val = values[i]
            if val is None:
                continue
            sval = str(val).strip()
            if not sval:
                continue
            # Variablenname = originaler Spaltenname (z. B. "url", "notificationTitle", ...)
            vars_map[col_name] = sval

        out.append({"campaign": str(camp_val).strip(), "variables": vars_map})


    return out


def build_jobs_from_excel(path: str,
                          template_path: str,
                          supported_languages: List[str] | None,
                          use_json_variables: bool = False) -> List[Dict[str, Any]]:
    """
    Baut komplette Job-Dicts aus der Excel (für direkten Batch-Run,
    falls du nicht über das UI pro Link einen Job baust).
    """
    rows = parse_update_excel(path)
    jobs: List[Dict[str, Any]] = []
    for r in rows:
        jobs.append({
            "campaigns": [r["campaign"]],
            "template_path": template_path,
            "supported_languages": supported_languages or [],
            "use_json_variables": use_json_variables,
            "extra_variables": r["variables"],
        })
    return jobs


# -------------------- Kernfunktion (ein Job) --------------------

_CAMPAIGN_IN_HASH = re.compile(r"#/\d+/(?:campaigns|programs)/(\d+)")
_DIGITS = re.compile(r"\b(\d{6,})\b")

def _to_campaign_id(s: str) -> int:
    """Extrahiert die Campaign-ID aus URL/#/MP/campaigns/<id> oder nimmt eine reine ID."""
    s = (s or "").strip()
    if not s:
        raise ValueError("empty campaign reference")
    if s.isdigit():
        return int(s)
    m = _CAMPAIGN_IN_HASH.search(s)
    if m:
        return int(m.group(1))
    m = _DIGITS.search(s)
    if m:
        return int(m.group(1))
    raise ValueError(f"cannot parse campaign id from: {s!r}")


def run_update_campaign_content(
    job: Dict[str, Any],
    status_callback: Callable[[str], None] | None = None,
    progress_callback: Callable[[int], None] | None = None,
    headless: bool = True,
) -> str | None:
    """
    Erwartet EINEN Job:

      job = {
        "campaigns": [ "<URL oder ID>" ],          # genau 1 Eintrag
        "template_path": "/LAYOUT-TEMPLATES/xxxx",
        "supported_languages": ["en_GB"],          # kurze Codes!
        "use_json_variables": False,
        "extra_variables": { "notificationTitle": "Hello", "url": "https://…", ... }
      }

    Rückgabe: Pfad zur JSON-Logdatei (oder None, falls Schreiben fehlschlug).
    """
    def tell(msg: str):
        if status_callback:
            status_callback(msg)

    # 1) Kampagne extrahieren
    campaigns = job.get("campaigns") or []
    if not campaigns:
        raise ValueError("job.campaigns must contain exactly one element")
    if len(campaigns) != 1:
        raise ValueError("this backend expects one campaign per job")
    cid = _to_campaign_id(str(campaigns[0]))

    template_path       = job.get("template_path") or ""
    supported_languages = job.get("supported_languages") or []
    use_json_vars       = bool(job.get("use_json_variables", False))
    extra_vars          = job.get("extra_variables") or {}

    if not template_path:
        raise ValueError("template_path is required")

    # 2) Session (Cookies aus Firefox)
    profile_path = get_firefox_profile()
    if not profile_path or not os.path.isdir(profile_path):
        raise RuntimeError("No Firefox profile found – open Firefox once and sign in.")

    sess = _build_session_from_firefox(profile_path)

    # 3) Payload bauen
    variables = _build_variables(template_path, supported_languages, use_json_vars, extra_vars)
    payload = {"campaignId": cid, "variables": variables}

    # 4) POST (als JSON)
    url = f"{SERVICE_BASE}/campaigns/{cid}/updateContent"
    tell(f"Updating content for campaign {cid} …")
    try:
        resp = _post_json_with_refresh(sess, url, payload, profile_path, headless=headless, timeout=(10, 60))

    except Exception as e:
        # vollständige Fehlermeldung nach oben, damit UI sie anzeigt
        raise

    if progress_callback:
        progress_callback(0)

    # 5) Logdatei schreiben (Request/Response)
    out_path: Optional[str] = None
    try:
        ts = time.strftime("%Y%m%d_%H%M%S")
        out_path = os.path.abspath(f"updateContent_{cid}_{ts}.json")
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump({"request": payload, "response": resp}, f, ensure_ascii=False, indent=2)

    except Exception:
        out_path = None

    tell(f"Campaign {cid} updated.")
    return out_path


# -------------------- Optionaler Batch-Runner (CLI/Tests) --------------------

def _export_results_csv(path: str, rows: List[Dict[str, Any]]) -> None:
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["campaignId", "templatePath", "variablesCount", "ok", "error"])
        for r in rows:
            w.writerow([r.get("campaignId"), r.get("templatePath"), r.get("variablesCount"), r.get("ok"), r.get("error")])


if __name__ == "__main__":
    # Mini-Demo im DRY-Style (kein echter POST hier, nur Payload/Datei wenn Service erreichbar)
    example_job = {
        "campaigns": ["https://sonar-eu.amazon.com/#/3/campaigns/1415322561"],
        "template_path": "/LAYOUT-TEMPLATES/f48f86ad-3935-4bc1-895d-b4d05cfa11f3",
        "supported_languages": ["en_GB"],
        "use_json_variables": False,
        "extra_variables": {
            "notificationTitle": "test",
            "notificationText": "test",
            "primaryButtonText": "test",
            "primaryButtonCta": "test",
            "url": "https://example.com",          # wird korrekt als Variable übertragen
            "consolidationKey": "abc",
            "hubImage": "https://img",
            "androidIconImage": "https://icon",
            "iosImageOrVideo": "https://ios",
            "androidBigPicture": "https://big",
        }
    }

    def _status(m: str): print(m)
    def _progress(i: int): pass

    try:
        out = run_update_campaign_content(example_job, _status, _progress, headless=True)
        print("Log:", out)
    except Exception as ex:
        print("Error:", ex)
