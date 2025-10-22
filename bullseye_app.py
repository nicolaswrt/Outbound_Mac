from pathlib import Path
import sys
import subprocess
import os, json, re
import tkinter as tk
import tkinter.font as tkfont
from tkinter import ttk, messagebox, simpledialog, filedialog

from get_sizes import get_segment_sizes
from extract_rules import get_segment_rules_http as run_extract_rules
from queue_segments import queue_segments as run_queue_segments
from sonar_apply import apply_segments_to_sonar_pairs  # optional: alias nicht nötig
from clone_publish import (
    clone_and_publish_segments as run_clone_and_publish_segments,
    mass_clone_fixed as run_mass_clone_fixed,            # falls extern gebraucht
    clone_across_marketplaces as backend_clone_across_mps,  # wird unten lokal importiert, optional
)
from create_os_sonar import run_create_os_sonar
from update_campaign_content import run_update_campaign_content as run_update_content
from preview_campaigns import (
    plan_preview_batches as preview_plan_batches,
    run_preview_batch_for_marketplace as preview_run_batch,
)
from approve_sonar import run_approve_sonar
from create_rc_sonar import create_remote_configs

import time
from datetime import timedelta

import requests
import sqlite3
import shutil
import tempfile

# Selenium wird nur im Cookie-Fallback genutzt:
from selenium import webdriver
from selenium.webdriver.firefox.service import Service as FxService
from selenium.webdriver.firefox.options import Options as FxOptions



MASS_CLONE_FIXED_BASE_BE = "1749101702"
"""
# --- Sonar MP → Language mapping for supportedLanguages ---
SONAR_MP_TO_LANGUAGE = {
    3: "language_en_GB",     # UK
    4: "language_de_DE",     # DE
    5: "language_fr_FR",     # FR
    35691: "language_it_IT", # IT
    44551: "language_es_ES", # ES
}
"""

def format_time(seconds):
    """Format time in seconds to readable string"""
    if isinstance(seconds, str):
        return seconds
    return str(timedelta(seconds=round(seconds)))

def _is_firefox_running() -> bool:
    """Return True if Firefox is currently running."""
    try:
        if os.name == "nt":
            # Windows: parse 'tasklist' output
            res = subprocess.run(["tasklist"], capture_output=True, text=True)
            return "firefox.exe" in (res.stdout or "").lower()
        else:
            # Linux/macOS: use pgrep
            res = subprocess.run(["pgrep", "-x", "firefox"], capture_output=True, text=True)
            return res.returncode == 0
    except Exception:
        # If check fails, assume not running to avoid blocking startup.
        return False


def _kill_firefox() -> bool:
    """Try to force-close Firefox. Returns True if the command ran without raising."""
    try:
        if os.name == "nt":
            subprocess.run(["taskkill", "/IM", "firefox.exe", "/F"], capture_output=True, text=True)
        else:
            subprocess.run(["pkill", "-x", "firefox"], capture_output=True, text=True)
        return True
    except Exception:
        return False


def run_startup_preflight(root) -> bool:
    """
    Preflight before showing the UI:
      1) If Firefox is running, ask to close it. If user refuses, exit.
      2) After it is closed, try to locate the Firefox profile once.
         If none is found, warn the user that they must set this up first.
    Returns True to continue opening the app, False to exit.
    """
    # 1) Ensure Firefox is not running
    while _is_firefox_running():
        choice = messagebox.askyesno(
            "Firefox is running",
            "Firefox must be closed for this app to work.\n\n"
            "Do you want me to close Firefox now?"
        )
        if not choice:
            messagebox.showinfo(
                "Exiting",
                "Please close Firefox yourself and restart the app."
            )
            return False

        _kill_firefox()
        # brief wait then re-check
        time.sleep(1.0)

        if _is_firefox_running():
            retry = messagebox.askretrycancel(
                "Still running",
                "Firefox still seems to be running.\n\n"
                "Close it manually and click Retry, or click Cancel to exit."
            )
            if not retry:
                return False
            # loop continues to re-check

    # 2) Probe Firefox profile availability
    try:
        from utils import get_firefox_profile
        profile_path = get_firefox_profile()
    except Exception:
        profile_path = None

    if not profile_path:
        messagebox.showwarning(
            "No Firefox profile found",
            "We could not find a valid Firefox profile.\n\n"
            "Please open Firefox once on this machine and sign in/set it up first, "
            "then restart the app."
        )
        # We allow the app to open, but many functions will not work until a profile exists.
        # If you prefer to block startup instead, return False here.
        # return False

    return True

def resource_path(rel_path: str) -> str:
    """
    Gibt einen absoluten Pfad auf eine Ressource zurück, egal ob:
    - direkt aus dem Quellbaum gestartet wird, oder
    - aus einem PyInstaller-Bundle (sys._MEIPASS)
    rel_path: z.B. "assets/Icon_outbound.png" (mit Slash)
    """
    base = getattr(sys, "_MEIPASS", None)
    if base:
        return str(Path(base, rel_path))
    return str(Path(__file__).resolve().parent.joinpath(rel_path))



# ---- Sonar Web (GUI) — Cookies & Session ----
SONAR_WEB_DOMAIN = "sonar-eu.amazon.com"

def _copy_sqlite_readonly(src_path: str):
    if not os.path.exists(src_path):
        raise FileNotFoundError(f"cookies.sqlite not found at: {src_path}")
    tmpdir = tempfile.mkdtemp(prefix="ff_cookies_")
    dst = os.path.join(tmpdir, "cookies.sqlite")
    shutil.copy2(src_path, dst)
    return dst, tmpdir

def _load_firefox_cookies_for_domain(profile_path: str, domain_suffix: str):
    cookies_db = os.path.join(profile_path, "cookies.sqlite")
    cleanup_dir = None
    try:
        conn = sqlite3.connect(f"file:{cookies_db}?mode=ro", uri=True)
    except sqlite3.OperationalError:
        copied_path, cleanup_dir = _copy_sqlite_readonly(cookies_db)
        conn = sqlite3.connect(copied_path)

    jar = requests.cookies.RequestsCookieJar()
    try:
        cur = conn.cursor()
        cur.execute(
            "SELECT name, value, host, path, isSecure FROM moz_cookies WHERE host LIKE ?",
            (f"%{domain_suffix}",)
        )
        for name, value, host, path, isSecure in cur.fetchall():
            jar.set(name, value, domain=host, path=path, secure=bool(isSecure))
    finally:
        conn.close()
        if cleanup_dir and os.path.isdir(cleanup_dir):
            shutil.rmtree(cleanup_dir, ignore_errors=True)
    return jar

def _build_sonar_web_session(profile_path: str) -> requests.Session:
    jar = _load_firefox_cookies_for_domain(profile_path, SONAR_WEB_DOMAIN)
    s = requests.Session()
    s.cookies = jar
    s.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:140.0) Gecko/20100101 Firefox/140.0",
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "en-US,en;q=0.5",
        "Referer": f"https://{SONAR_WEB_DOMAIN}/",
        "Origin": f"https://{SONAR_WEB_DOMAIN}",
        "X-Requested-With": "XMLHttpRequest",
        "Connection": "keep-alive",
    })
    return s


def _cookiejar_from_selenium_cookies(cookies_list):
    jar = requests.cookies.RequestsCookieJar()
    for c in cookies_list:
        jar.set(
            c.get("name"),
            c.get("value"),
            domain=c.get("domain"),
            path=c.get("path", "/"),
            secure=bool(c.get("secure", False))
        )
    return jar

def _selenium_refresh_sonar_cookies(profile_path: str, headless: bool = True):
    """Startet Firefox mit Profil, lädt Sonar-Web, gibt Cookies als RequestsCookieJar zurück."""
    opts = FxOptions()
    opts.add_argument("-profile")
    opts.add_argument(profile_path)
    if headless:
        opts.add_argument("--headless")
        opts.add_argument("--width=1920")
        opts.add_argument("--height=1080")

    service = FxService("geckodriver.exe")  # ggf. Pfad anpassen
    driver = None
    try:
        driver = webdriver.Firefox(service=service, options=opts)
        driver.get(f"https://{SONAR_WEB_DOMAIN}/")
        time.sleep(3)
        return _cookiejar_from_selenium_cookies(driver.get_cookies())
    finally:
        if driver:
            try:
                driver.quit()
            except Exception:
                pass


def _fetch_json_from_sonar(url: str, profile_path: str, timeout=(5, 30)) -> dict:
    """
    Holt JSON von Sonar-Web; wenn die Antwort kein JSON ist (z. B. Login-HTML),
    werden Cookies via Selenium aufgefrischt und ein zweiter Versuch gemacht.
    """
    s = _build_sonar_web_session(profile_path)

    def _one_try():
        r = s.get(url, timeout=timeout)
        ct = (r.headers.get("Content-Type") or "").lower()
        txt = r.text or ""
        # JSON-Erkennung: Content-Type oder führendes Zeichen
        looks_json = ("json" in ct) or (txt.strip().startswith("{") or txt.strip().startswith("["))
        if r.status_code in (401, 403) or not looks_json:
            return None, r
        try:
            return r.json(), r
        except Exception:
            try:
                return json.loads(txt), r
            except Exception:
                return None, r

    data, resp = _one_try()
    if data is not None:
        return data

    # Selenium-Refresh
    jar = _selenium_refresh_sonar_cookies(profile_path, headless=True)
    if jar:
        s.cookies = jar
        data, resp = _one_try()
        if data is not None:
            return data

    # Verständliche Fehlermeldung mit Snippet
    status = resp.status_code if resp is not None else "no-status"
    snippet = (resp.text or "")[:240].replace("\n", " ").replace("\r", " ")
    raise RuntimeError(f"Sonar lieferte keine JSON-Antwort (status={status}). "
                       f"Snippet: {snippet!r}")








class BullseyeApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Outbound Automation")
        self.root.geometry("1240x760")
        self.root.minsize(1040, 680)

        try:
            icon_path = resource_path("assets/Icon_outbound.png")
            self._app_icon = tk.PhotoImage(file=icon_path)   # Referenz behalten!
            self.root.iconphoto(True, self._app_icon)
        except Exception as e:
            print(f"Icon konnte nicht geladen werden: {e}")

        # Styles
        self.setup_amazon_style()
        self.METRICS_RESULTS_HEIGHT = 220

        # --- Dataset storage (per-user) ---
        self.datasets_dir = os.path.join(os.path.expanduser("~"), ".bullseye_automation")
        os.makedirs(self.datasets_dir, exist_ok=True)
        self.datasets_file = os.path.join(self.datasets_dir, "datasets.json")
        self.datasets = self.load_datasets()

        # --- Profile storage (exactly one profile) ---
        self.profile_file = os.path.join(self.datasets_dir, "profile.json")
        self.profile = self.load_profile()
        self._prepared_preview_batches = None
        self._preview_mps_sent = set() 
        # Profile-Werte auch als Umgebungsvariablen bereitstellen (für Helper/Fallbacks)
        self._export_profile_to_env()


        # --- Sonar Templates storage ---
        self.templates_file = os.path.join(self.datasets_dir, "sonar_templates.json")
        self.templates = self.load_templates()  # list[dict]




        # ---------- Amazon Header (dark navbar) ----------
        header = tk.Frame(root, bg=self.AMAZON["navy"], height=56)
        header.pack(fill="x")
        header.pack_propagate(False)

        title_wrap = tk.Frame(header, bg=self.AMAZON["navy"])
        title_wrap.pack(side="left", padx=16, pady=8)

        title_font = tkfont.Font(family="Segoe UI", size=13, weight="bold")
        subtitle_font = tkfont.Font(family="Segoe UI", size=9)

        title = tk.Label(
            title_wrap,
            text="Outbound Automation",
            bg=self.AMAZON["navy"],
            fg="white",
            font=title_font
        )
        subtitle = tk.Label(
            title_wrap,
            text="Tools for Outbound • Amazon use only",
            bg=self.AMAZON["navy"],
            fg="#D5D8DC",
            font=subtitle_font
        )
        title.pack(anchor="w")
        subtitle.pack(anchor="w")

        # --- Profile UI (rechts im Header) ---
        profile_wrap = tk.Frame(header, bg=self.AMAZON["navy"])
        profile_wrap.pack(side="right", padx=16, pady=8)

        header_btn_font = tkfont.Font(family="Segoe UI", size=10, weight="bold")

        self.hello_label = tk.Label(
            profile_wrap,
            text="",                         # wird durch refresh_profile_ui gesetzt
            bg=self.AMAZON["navy"],
            fg="#D5D8DC",
            font=subtitle_font
        )
        self.hello_label.pack(side="left", padx=(0, 8))

        self.profile_btn = tk.Button(
            profile_wrap,
            text="Create profile",            # wird dynamisch geändert
            command=self.on_profile_click,
            bg=self.AMAZON["navy"],
            fg="white",
            activebackground=self.AMAZON["navyLight"],
            activeforeground="white",
            relief="flat",
            padx=12, pady=6,
            font=header_btn_font,
            cursor="hand2",
            bd=0
        )
        self.profile_btn.pack(side="left")

        self.templates_btn = tk.Button(
            profile_wrap,
            text="Templates",
            command=self.open_templates_manager,
            bg=self.AMAZON["navy"],
            fg="white",
            activebackground=self.AMAZON["navyLight"],
            activeforeground="white",
            relief="flat",
            padx=12, pady=6,
            font=header_btn_font,
            cursor="hand2",
            bd=0
        )
        self.templates_btn.pack(side="left", padx=(8, 0))










        # UI initial befüllen
        self.refresh_profile_ui()



        # ---------- Page Header (make sure background matches, no white boxes) ----------
        page_hdr = tk.Frame(root, bg=self.AMAZON["bg"])
        page_hdr.pack(fill="x", padx=16, pady=(14, 8))
        ttk.Label(page_hdr, text="Select Action", style="AmazonTitle.TLabel", background=self.AMAZON["bg"]).pack(anchor="w")
        ttk.Label(page_hdr, text="Fetch sizes, extract rules, or queue segments",
                  style="AmazonSubtitle.TLabel", background=self.AMAZON["bg"]).pack(anchor="w", pady=(2, 0))


        # ---------- Row 1: Action (links) + Datasets (rechts) ----------
        row1 = ttk.Frame(root, style="Amazon.TFrame")
        row1.pack(fill="both", padx=16, pady=(6, 8))
        # zwei gleich breite Spalten, gleiche Höhe in der Zeile
        row1.columnconfigure(0, weight=1, uniform="row1")
        row1.columnconfigure(1, weight=1, uniform="row1")
        row1.rowconfigure(0, weight=1)

        # Action card (links)
        fn_card = ttk.LabelFrame(row1, text="Action", style="AmazonCard.TLabelframe", padding=12)
        fn_card.grid(row=0, column=0, sticky="nsew", padx=(0, 8))

        self.function_var = tk.StringVar(value="sizes")

        fn_row = ttk.Frame(fn_card, style="Amazon.TFrame")
        fn_row.pack(fill="x")

        fn_row.columnconfigure(0, weight=1)
        fn_row.columnconfigure(1, weight=1)

        functions = [
            ("Get Segment Sizes", "sizes"),
            ("Queue Now", "queue"),
            ("Extract Rules", "rules"),
            ("Upload BE → Sonar", "sonar"),
            ("Clone BE", "clone_and_publish"),
            ("Create OS Sonar", "create_os_sonar"),
            ("Update Content (Sonar)", "update_content"),
            ("Mass Clone (Fixed)", "mass_clone_fixed"),
            ("Clone across MPs", "clone_across_mps"),
            ("Send Preview (Sonar)", "send_preview"),
            ("Approve Sonar", "approve_sonar"),
            ("Create RC Sonar", "create_rc_sonar"),
        ]

        for idx, (text, value) in enumerate(functions):
            r, c = divmod(idx, 2)
            ttk.Radiobutton(
                fn_row,
                text=text,
                variable=self.function_var,
                value=value,
                style="Amazon.TRadiobutton"
            ).grid(row=r, column=c, sticky="w", padx=(0, 16), pady=2)

        self.headless_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            fn_row,
            text="Run in background (headless)",
            variable=self.headless_var,
            style="Amazon.TCheckbutton"
        ).grid(row=(len(functions)+1)//2, column=0, columnspan=2, sticky="w", pady=(8, 0))

        # Proceed-Button in der Action-Karte (unten rechts)
        # Spacer füllt den verbleibenden Platz, damit der Button wirklich unten sitzt
        spacer = ttk.Frame(fn_card, style="Amazon.TFrame")
        spacer.pack(fill="both", expand=True)

        actions_in_card = ttk.Frame(fn_card, style="Amazon.TFrame", padding=(0, 8, 0, 0))
        actions_in_card.pack(side="bottom", fill="x", anchor="se")
        self.create_amazon_button(actions_in_card, "Proceed", self.proceed).pack(side="right")

        
        # Datasets card (rechts) — Treeview mit moderner Scrollbar + Buttons unten rechts
        datasets_card = ttk.LabelFrame(row1, text="Datasets", style="AmazonCard.TLabelframe", padding=12)
        datasets_card.grid(row=0, column=1, sticky="nsew", padx=(8, 0))

        # Body-Container mit Grid: Treeview (col 0) + Scrollbar (col 1)
        ds_body = ttk.Frame(datasets_card, style="Amazon.TFrame")
        ds_body.pack(fill="both", expand=True)
        ds_body.columnconfigure(0, weight=1)
        ds_body.columnconfigure(1, weight=0)
        ds_body.rowconfigure(0, weight=1)

        # Treeview: Name | Type (mit Amazon-Styles)
        self.datasets_tv = ttk.Treeview(
            ds_body,
            columns=("name", "type"),
            show="headings",
            style="Amazon.Treeview",
            height=8
        )
        self.datasets_tv.heading("name", text="Name")
        self.datasets_tv.heading("type", text="Type")
        self.datasets_tv.column("name", width=200, anchor="w")
        self.datasets_tv.column("type", width=70, anchor="center")

        self.datasets_tv.grid(row=0, column=0, sticky="nsew", padx=(0, 8), pady=(0, 8))

        # Vertikale Scrollbar rechts neben der Liste
        ds_scroll = ttk.Scrollbar(
            ds_body,
            orient="vertical",
            command=self.datasets_tv.yview,
            style="Amazon.Vertical.TScrollbar"
        )
        ds_scroll.grid(row=0, column=1, sticky="ns", pady=(0, 8))
        self.datasets_tv.configure(yscrollcommand=ds_scroll.set)

        # Buttons unten rechts – visuell: Create | Edit | Delete
        ds_actions = ttk.Frame(datasets_card, style="Amazon.TFrame")
        ds_actions.pack(fill="x")

        # pack(side="right") → letzte Pack-Anweisung ist links außen.
        self.btn_ds_delete = self.create_secondary_button(ds_actions, "Delete", self.dataset_delete)
        self.btn_ds_delete.pack(side="right", padx=(8, 0))

        self.btn_ds_edit = self.create_secondary_button(ds_actions, "Edit", self.dataset_edit)
        self.btn_ds_edit.pack(side="right", padx=(8, 0))

        self.btn_ds_create = self.create_amazon_button(ds_actions, "Create", self.dataset_create)
        self.btn_ds_create.pack(side="right")

        # Auswahl-Handling
        self.datasets_tv.bind("<<TreeviewSelect>>", self.on_dataset_select)
        self.datasets_tv.bind("<Double-1>", lambda e: self.dataset_edit())

        # Start: Edit/Delete deaktivieren
        self.btn_ds_edit.configure(state="disabled")
        self.btn_ds_delete.configure(state="disabled")

        # Liste füllen
        self.refresh_datasets_view()


                # ---------- Row 2: Progress (links) + Status (rechts) ----------
        row2 = ttk.Frame(root, style="Amazon.TFrame")
        row2.pack(fill="both", padx=16, pady=8)
        row2.columnconfigure(0, weight=1, uniform="row2")
        row2.columnconfigure(1, weight=1, uniform="row2")
        row2.rowconfigure(0, weight=1)

        # Progress card (links)
        progress_card = ttk.LabelFrame(row2, text="Progress", style="AmazonCard.TLabelframe", padding=12)
        progress_card.grid(row=0, column=0, sticky="nsew", padx=(0, 8))

        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(
            progress_card,
            variable=self.progress_var,
            maximum=100,
            style="Amazon.Horizontal.TProgressbar"
        )
        self.progress_bar.pack(fill="x", pady=(4, 8))

        self.progress_label = ttk.Label(progress_card, text="Ready…", style="AmazonBody.TLabel")
        self.progress_label.pack(anchor="w")

        self.time_label = ttk.Label(progress_card, text="", style="AmazonMuted.TLabel")
        self.time_label.pack(anchor="w", pady=(2, 0))

        # Status card (rechts)
        status_card = ttk.LabelFrame(row2, text="Status", style="AmazonCard.TLabelframe", padding=12)
        status_card.grid(row=0, column=1, sticky="nsew", padx=(8, 0))

        self.status_var = tk.StringVar(value="Choose an action and click Proceed.")
        self.status_label = ttk.Label(
            status_card,
            textvariable=self.status_var,
            style="AmazonBody.TLabel",
            wraplength=460,
            justify="left"
        )
        self.status_label.pack(fill="both", expand=True)

        # wraplength dynamisch an Kartenbreite anpassen (für schöne Zeilenumbrüche)
        def _resize_status_wraplength(event):
            pad = 32  # grobe Innenabzüge
            self.status_label.config(wraplength=max(100, event.width - pad))
        status_card.bind("<Configure>", _resize_status_wraplength)


                # ---------- Row 3: Performance Metrics (links) + Results (rechts) ----------
        row3 = ttk.Frame(root, style="Amazon.TFrame")
        row3.pack(fill="both", padx=16, pady=(8, 16))
        row3.columnconfigure(0, weight=1, uniform="row3")
        row3.columnconfigure(1, weight=1, uniform="row3")
        row3.rowconfigure(0, weight=1)

        # Wrapper mit fixer, gemeinsamer Höhe (Links)
        metrics_wrap = tk.Frame(row3, bg=self.AMAZON["bg"], height=self.METRICS_RESULTS_HEIGHT)
        metrics_wrap.pack_propagate(False)
        metrics_wrap.grid(row=0, column=0, sticky="nsew", padx=(0, 8))

        metrics_card = ttk.LabelFrame(metrics_wrap, text="Performance Metrics", style="AmazonCard.TLabelframe", padding=12)
        metrics_card.pack(fill="both", expand=True)

        self.total_time_label = ttk.Label(metrics_card, text="Total Time: -", style="AmazonBody.TLabel")
        self.total_time_label.pack(anchor="w", pady=2)

        self.avg_batch_time_label = ttk.Label(metrics_card, text="Average Batch Time: -", style="AmazonBody.TLabel")
        self.avg_batch_time_label.pack(anchor="w", pady=2)

        self.success_rate_label = ttk.Label(metrics_card, text="Success Rate: -", style="AmazonBody.TLabel")
        self.success_rate_label.pack(anchor="w", pady=2)

        self.failed_segments_label = ttk.Label(metrics_card, text="Failed Segments: -", style="AmazonBody.TLabel")
        self.failed_segments_label.pack(anchor="w", pady=2)

        self.avg_segment_time_label = ttk.Label(metrics_card, text="Average Time per Segment: -", style="AmazonBody.TLabel")
        self.avg_segment_time_label.pack(anchor="w", pady=2)

        # Wrapper mit fixer, gemeinsamer Höhe (Rechts)
        results_wrap = tk.Frame(row3, bg=self.AMAZON["bg"], height=self.METRICS_RESULTS_HEIGHT)
        results_wrap.pack_propagate(False)
        results_wrap.grid(row=0, column=1, sticky="nsew", padx=(8, 0))

        results_card = ttk.LabelFrame(results_wrap, text="Results", style="AmazonCard.TLabelframe", padding=12)
        results_card.pack(fill="both", expand=True)

        # Nur Dateinamenliste (kein innerer Rahmen), ttk-Scrollbar
        results_frame = ttk.Frame(results_card, style="Amazon.TFrame")
        results_frame.pack(fill="both", expand=True)

        self.results_listbox = tk.Listbox(
            results_frame,
            font=("Segoe UI", 10),
            activestyle="none",
            bg=self.AMAZON["bg"],
            fg=self.AMAZON["text"],
            highlightthickness=0,   # kein Rahmen
            relief="flat",
            bd=0
        )
        self.results_listbox.pack(side="left", fill="both", expand=True)

        results_scroll = ttk.Scrollbar(
            results_frame,
            orient="vertical",
            command=self.results_listbox.yview,
            style="Amazon.Vertical.TScrollbar"
        )
        results_scroll.pack(side="right", fill="y")
        self.results_listbox.configure(yscrollcommand=results_scroll.set)

        self.results_listbox.bind("<FocusOut>", lambda e: self.results_listbox.selection_clear(0, tk.END))

        self.result_files = []  # Vollpfade in gleicher Reihenfolge wie Listbox-Einträge
        self.results_listbox.bind("<Double-1>", self._on_result_double_click)



# ------------------- Styles -------------------


    def setup_amazon_style(self):
        """Amazon-like colors & ttk styles"""
        try:
            ttk.Style().theme_use('clam')
        except Exception:
            pass

        # Fonts set globally via Tk named fonts (avoids 'expected integer' errors)
        default_font = tkfont.nametofont("TkDefaultFont")
        default_font.configure(family="Segoe UI", size=10)
        tkfont.nametofont("TkTextFont").configure(family="Segoe UI", size=10)
        tkfont.nametofont("TkHeadingFont").configure(family="Segoe UI", size=11, weight="bold")

        self.AMAZON = {
            "navy": "#232F3E",     # Header
            "navyLight": "#37475A",
            "bg": "#F3F3F3",       # page background
            "card": "#FFFFFF",
            "border": "#E5E7EB",
            "text": "#111111",
            "muted": "#565959",
            "link": "#007185",
            "orange": "#FF9900",
            "orangeHover": "#F09000",
            "orangeActive": "#E48200",
            "progress": "#FF9900",
        }

        self.root.configure(bg=self.AMAZON["bg"])

        s = ttk.Style()

        # Generic frame background (prevents white boxes)
        s.configure("Amazon.TFrame", background=self.AMAZON["bg"])


        s.configure("Amazon.TCheckbutton",
                    background=self.AMAZON["bg"],
                    foreground=self.AMAZON["text"])

        # Card / Labelframe
        s.configure("AmazonCard.TLabelframe",
                    background=self.AMAZON["bg"],
                    foreground=self.AMAZON["text"],
                    bordercolor=self.AMAZON["border"],
                    relief="solid")
        s.configure("AmazonCard.TLabelframe.Label",
                    background=self.AMAZON["bg"],
                    foreground=self.AMAZON["text"],
                    font=("Segoe UI", 11, "bold"))

        # Labels
        s.configure("AmazonTitle.TLabel",
                    background=self.AMAZON["bg"],
                    foreground=self.AMAZON["text"],
                    font=("Segoe UI", 14, "bold"))
        s.configure("AmazonSubtitle.TLabel",
                    background=self.AMAZON["bg"],
                    foreground=self.AMAZON["muted"],
                    font=("Segoe UI", 10))
        s.configure("AmazonBody.TLabel",
                    background=self.AMAZON["bg"],
                    foreground=self.AMAZON["text"],
                    font=("Segoe UI", 10))
        s.configure("AmazonMuted.TLabel",
                    background=self.AMAZON["bg"],
                    foreground=self.AMAZON["muted"],
                    font=("Segoe UI", 9))

        # Radiobuttons
        s.configure("Amazon.TRadiobutton",
                    background=self.AMAZON["bg"],
                    foreground=self.AMAZON["text"],
                    focuscolor=self.AMAZON["card"])

        # Progressbar
        s.configure("Amazon.Horizontal.TProgressbar",
                    troughcolor=self.AMAZON["border"],
                    background=self.AMAZON["progress"],
                    bordercolor=self.AMAZON["border"])

        # schlanke, dezente ttk-Scrollbar (clam)
        s.configure(
            "Amazon.Vertical.TScrollbar",
            troughcolor=self.AMAZON["border"],
            background=self.AMAZON["bg"],
            bordercolor=self.AMAZON["border"],
            arrowcolor=self.AMAZON["muted"]
        )
        s.map(
            "Amazon.Vertical.TScrollbar",
            background=[("active", self.AMAZON["navyLight"])],
            arrowcolor=[("active", "white")]
        )
        # --- Treeview (Datasets) – kleinere Fonts ---
        try:
            self._tree_font_row = tkfont.Font(family="Segoe UI", size=8)          # Zeilen
            self._tree_font_head = tkfont.Font(family="Segoe UI", size=8, weight="bold")  # Header
        except Exception:
            self._tree_font_row = ("Segoe UI", 8)
            self._tree_font_head = ("Segoe UI", 8, "bold")

        s.configure(
            "Amazon.Treeview",
            background=self.AMAZON["bg"],
            fieldbackground=self.AMAZON["bg"],
            foreground=self.AMAZON["text"],
            bordercolor=self.AMAZON["border"],
            rowheight=20,                        # etwas kleiner passend zur 8-pt-Schrift
            font=self._tree_font_row
        )
        s.configure(
            "Amazon.Treeview.Heading",
            background=self.AMAZON["navy"],
            foreground="white",
            font=self._tree_font_head,
            bordercolor=self.AMAZON["border"]
        )
        s.map(
            "Amazon.Treeview.Heading",
            background=[("active", self.AMAZON["navyLight"])],
            foreground=[("active", "white")]
        )



        # Scrollbars: modernes Layout ohne Pfeile, etwas dicker über arrowsize
        s.layout(
            "Amazon.Vertical.TScrollbar",
            [
                ("Vertical.Scrollbar.trough", {
                    "sticky": "ns",
                    "children": [
                        ("Vertical.Scrollbar.thumb", {"sticky": "nswe"})
                    ]
                })
            ]
        )
        s.configure("Amazon.Vertical.TScrollbar", arrowsize=12)  # Dicke/Größe

        s.layout(
            "Amazon.Horizontal.TScrollbar",
            [
                ("Horizontal.Scrollbar.trough", {
                    "sticky": "ew",
                    "children": [
                        ("Horizontal.Scrollbar.thumb", {"sticky": "nswe"})
                    ]
                })
            ]
        )
        s.configure("Amazon.Horizontal.TScrollbar", arrowsize=12)












        s.configure(
            "Amazon.Vertical.TScrollbar",
            troughcolor=self.AMAZON["bg"],
            background=self.AMAZON["navyLight"],
            bordercolor=self.AMAZON["bg"],
            lightcolor=self.AMAZON["bg"],
            darkcolor=self.AMAZON["bg"],
            arrowsize=0
        )
        s.map(
            "Amazon.Vertical.TScrollbar",
            background=[("active", self.AMAZON["navy"])],
        )



    def create_amazon_button(self, parent, text, command):
        """Orange Amazon-like button with hover (tk.Button)"""
        btn_font = tkfont.Font(family="Segoe UI", size=10, weight="bold")
        btn = tk.Button(
            parent,
            text=text,
            command=command,
            bg=self.AMAZON["orange"],
            fg="white",
            activebackground=self.AMAZON["orangeActive"],
            activeforeground="white",
            relief="flat",
            padx=14, pady=8,
            font=btn_font,
            cursor="hand2",
            bd=0
        )
        def on_enter(_): btn.configure(bg=self.AMAZON["orangeHover"])
        def on_leave(_): btn.configure(bg=self.AMAZON["orange"])
        btn.bind("<Enter>", on_enter)
        btn.bind("<Leave>", on_leave)
        return btn

    def create_secondary_button(self, parent, text, command):
        """Secondary gray button in Amazon style."""
        btn = tk.Button(
            parent,
            text=text,
            command=command,
            bg="#E5E7EB",           # light gray
            fg="#111111",
            activebackground="#D1D5DB",
            activeforeground="#111111",
            relief="flat",
            padx=14, pady=8,
            font=tkfont.Font(family="Segoe UI", size=10),
            cursor="hand2",
            bd=0
        )
        def on_enter(_): btn.configure(bg="#D1D5DB")
        def on_leave(_): btn.configure(bg="#E5E7EB")
        btn.bind("<Enter>", on_enter)
        btn.bind("<Leave>", on_leave)
        return btn

    # ------------------- UI Updates -------------------

    def update_progress(self, index, total):
        """Update progress bar and labels"""
        progress = ((index + 1) / total) * 100
        self.progress_var.set(progress)
        self.progress_label.config(
            text=f"Processing: {index + 1}/{total} segments ({progress:.1f}%)"
        )
        self.root.update()

    def update_metrics(self, total_time, stats):
        """Update performance metrics labels"""
        # Normalize total_time to seconds (int)
        if isinstance(total_time, str) and ':' in total_time:
            parts = total_time.split(':')
            if len(parts) == 3:
                total_seconds = int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
            else:
                total_seconds = int(parts[0]) * 60 + int(parts[1])
        else:
            total_seconds = int(total_time) if isinstance(total_time, (int, float)) else 0

        self.total_time_label.config(text=f"Total Time: {total_seconds}s")
        self.avg_batch_time_label.config(text=f"Average Batch Time: {stats.get('average_batch_time', '-')}")
        self.success_rate_label.config(text=f"Success Rate: {stats.get('average_success_rate', 0)*100:.1f}%")
        self.failed_segments_label.config(text=f"Failed Segments: {stats.get('failed_segments', 0)}")
        if 'avg_segment_time' in stats:
            self.avg_segment_time_label.config(text=f"Average Time per Segment: {stats['avg_segment_time']}")
        self.root.update()




    # ---- Dataset v2: Misch-Datasets ----
    DATASET_TYPES = ("NAME", "BE", "SONAR")  # Anzeige- und Speicher-Reihenfolge

    def ds_types(self, ds):
        """Typen (Keys) eines Datasets in gespeicherter Reihenfolge."""
        return [c.get("key") for c in ds.get("columns", []) if c.get("key")]

    def ds_has(self, ds, key):
        """Hat das Dataset eine Spalte dieses Typs?"""
        return key in self.ds_types(ds)

    def ds_items(self, ds, key):
        """Hole nur die Items einer bestimmten Spalte."""
        for c in ds.get("columns", []):
            if c.get("key") == key:
                return list(c.get("items", []))
        return []


    def filter_datasets(self, key):
        """Alle Datasets, die eine Spalte `key` (z. B. 'BE' oder 'SONAR') haben."""
        return [ds for ds in (self.datasets or []) if self.ds_has(ds, key)]



    # ------------------- Datasets: Storage (v2) -------------------

    def load_datasets(self):
        """Load v2 datasets (mit optionaler Migration von v1)."""
        try:
            if os.path.exists(self.datasets_file):
                with open(self.datasets_file, "r", encoding="utf-8") as f:
                    raw = json.load(f)
                if not isinstance(raw, list):
                    return []
                out = []
                changed = False
                for ds in raw:
                    if "columns" in ds:
                        out.append(ds)
                    elif "type" in ds and "items" in ds:  # v1 → v2 Migration
                        cols = [{"key": ds["type"], "title": ds["type"], "items": ds["items"]}]
                        out.append({"name": ds.get("name", "Unnamed"), "columns": cols})
                        changed = True
                if changed:
                    try:
                        with open(self.datasets_file, "w", encoding="utf-8") as f:
                            json.dump(out, f, ensure_ascii=False, indent=2)
                    except Exception:
                        pass
                return out
        except Exception:
            pass
        return []

    def save_datasets(self):
        """Persist datasets to JSON."""
        try:
            with open(self.datasets_file, "w", encoding="utf-8") as f:
                json.dump(self.datasets, f, ensure_ascii=False, indent=2)
        except Exception as e:
            messagebox.showerror("Error", f"Could not save datasets:\n{e}")


    def refresh_datasets_view(self):
        """Refresh the treeview from self.datasets (v2-aware)."""
        for iid in self.datasets_tv.get_children():
            self.datasets_tv.delete(iid)
        for i, ds in enumerate(self.datasets):
            # Anzeige: vorhandene Spalten in Reihenfolge NAME | BE | SONAR
            types_in_order = [t for t in self.DATASET_TYPES if self.ds_has(ds, t)]
            type_str = " • ".join(types_in_order) if types_in_order else "-"
            self.datasets_tv.insert("", "end", iid=str(i), values=(ds.get("name", ""), type_str))



    def on_dataset_select(self, *_):
        """Enable/disable Edit/Delete based on selection."""
        sel = self.datasets_tv.selection()
        state = "normal" if sel else "disabled"
        self.btn_ds_edit.configure(state=state)
        self.btn_ds_delete.configure(state=state)

    def _get_selected_index(self):
        sel = self.datasets_tv.selection()
        if not sel:
            return None
        try:
            return int(sel[0])
        except Exception:
            return None

    # ------------------- Datasets: Parsing & Type detection -------------------

    def parse_dataset_text(self, raw_text):
        """
        Clever extraction:
        - If text contains 'https://sonar-eu.amazon.com' -> type=SONAR, extract all such URLs.
        - Else -> type=BE, extract all 10-digit IDs anywhere.
        Returns (type_str, items_list). Raises ValueError if nothing valid.
        """
        text = (raw_text or "").strip()
        if not text:
            raise ValueError("Please paste at least one line.")

        sonar_domain = "sonar-eu.amazon.com"
        # SONAR: collect URLs containing the domain
        if sonar_domain in text:
            # very permissive URL regex around the domain
            url_re = re.compile(r'https?://sonar-eu\.amazon\.com[^\s]*', re.IGNORECASE)
            items = url_re.findall(text)
            items = self._unique_preserve_order([it.strip() for it in items if it.strip()])
            if not items:
                raise ValueError("No SONAR URLs found.")
            return ("SONAR", items)

        # else: BE IDs (10 digits)
        ids = re.findall(r'\b\d{10}\b', text)
        ids = self._unique_preserve_order(ids)
        if not ids:
            raise ValueError("No 10-digit BE IDs found.")
        return ("BE", ids)

    @staticmethod
    def _unique_preserve_order(seq):
        seen = set()
        out = []
        for x in seq:
            if x not in seen:
                seen.add(x)
                out.append(x)
        return out

    def _name_exists(self, name, exclude_index=None):
        name_low = (name or "").strip().lower()
        for idx, ds in enumerate(self.datasets):
            if exclude_index is not None and idx == exclude_index:
                continue
            if ds.get("name", "").strip().lower() == name_low:
                return True
        return False

    # ------------------- Datasets: Create / Edit / Delete -------------------

    def dataset_create(self):
        """Open create dialog; save on confirm."""
        self._open_dataset_dialog(mode="create")

    def dataset_edit(self):
        """Open edit dialog for selected dataset."""
        idx = self._get_selected_index()
        if idx is None:
            return
        self._open_dataset_dialog(mode="edit", index=idx, initial=self.datasets[idx])

    def dataset_delete(self):
        """Delete selected dataset after confirmation."""
        idx = self._get_selected_index()
        if idx is None:
            return
        ds = self.datasets[idx]
        name = ds.get("name", "Unnamed")
        if not messagebox.askyesno("Delete Dataset", f"Delete dataset '{name}'?"):
            return
        del self.datasets[idx]
        self.save_datasets()
        self.refresh_datasets_view()
        self.btn_ds_edit.configure(state="disabled")
        self.btn_ds_delete.configure(state="disabled")

    def _open_dataset_dialog(self, mode="create", index=None, initial=None):
        """
        v2: Misch-Datasets mit Spalten NAME/BE/SONAR.
        Spalten liegen nebeneinander (Excel-artig) und können per "Add column" ergänzt/entfernt werden.
        - Keine doppelten Typen im Dataset
        - Alle vorhandenen Spalten müssen gleich viele Zeilen haben
        - BE: 10 Ziffern pro Zeile
        - SONAR: ID (Ziffern) ODER URL beginnend mit https://sonar-eu.amazon.com
        """
        assert mode in ("create", "edit")

        dlg = tk.Toplevel(self.root)
        dlg.title("Create Dataset" if mode == "create" else "Edit Dataset")
        dlg.configure(bg=self.AMAZON["bg"])
        dlg.geometry("880x560")
        dlg.minsize(820, 520)
        dlg.transient(self.root)
        dlg.grab_set()

        # zentrieren
        dlg.update_idletasks()
        x = self.root.winfo_rootx() + (self.root.winfo_width() // 2 - dlg.winfo_width() // 2)
        y = self.root.winfo_rooty() + (self.root.winfo_height() // 2 - dlg.winfo_height() // 2)
        dlg.geometry(f"+{max(0, x)}+{max(0, y)}")

        # Header
        hdr = tk.Frame(dlg, bg=self.AMAZON["bg"])
        hdr.pack(fill="x", padx=16, pady=(14, 6))
        ttk.Label(hdr, text=("Create a new dataset" if mode == "create" else "Edit dataset"),
                  style="AmazonTitle.TLabel").pack(anchor="w")
        ttk.Label(hdr, text="Add one or more columns (NAME, BE, SONAR). Columns must be unique.",
                  style="AmazonSubtitle.TLabel").pack(anchor="w", pady=(2, 0))

        # Body
        body = tk.Frame(dlg, bg=self.AMAZON["bg"])
        body.pack(fill="both", expand=True, padx=16, pady=(8, 0))

        # Dataset Name + Add column oben
        top_row = ttk.Frame(body, style="Amazon.TFrame")
        top_row.pack(fill="x", pady=(0, 8))
        ttk.Label(top_row, text="Dataset Name", style="AmazonBody.TLabel").pack(side="left", anchor="w")
        name_var = tk.StringVar(value=(initial.get("name") if (initial and initial.get("name")) else ""))
        name_entry = ttk.Entry(top_row, textvariable=name_var)
        name_entry.pack(side="left", fill="x", expand=True, padx=(8, 8))
        name_entry.focus_set()

        # Button rechts in derselben Zeile
        add_btn = self.create_secondary_button(top_row, "Add column", lambda: add_column())
        add_btn.pack(side="right")

                # Spalten-Fläche (Canvas) mit H- und V-Scrollbars
        cols_wrap = ttk.Frame(body, style="Amazon.TFrame")
        cols_wrap.pack(fill="both", expand=True)

        canvas = tk.Canvas(cols_wrap, bg=self.AMAZON["bg"], highlightthickness=0, bd=0)
        vbar = ttk.Scrollbar(cols_wrap, orient="vertical",
                             command=canvas.yview, style="Amazon.Vertical.TScrollbar")
        hbar = ttk.Scrollbar(cols_wrap, orient="horizontal",
                             command=canvas.xview)  # horizontal: default style ok

        canvas.configure(xscrollcommand=hbar.set, yscrollcommand=vbar.set)

        # Grid-Layout für sauberes Resize
        cols_wrap.rowconfigure(0, weight=1)
        cols_wrap.columnconfigure(0, weight=1)
        canvas.grid(row=0, column=0, sticky="nsew")
        vbar.grid(row=0, column=1, sticky="ns")
        hbar.grid(row=1, column=0, sticky="ew")

        inner = ttk.Frame(canvas, style="Amazon.TFrame")
        win_id = canvas.create_window((0, 0), window=inner, anchor="nw")

        def _sync_scrollregion(*_):
            canvas.configure(scrollregion=canvas.bbox("all"))

        inner.bind("<Configure>", _sync_scrollregion)
        canvas.bind("<Configure>", _sync_scrollregion)

        # Spalten horizontal in row=0 nebeneinander
        col_blocks = []  # [{'frame','type_var','hint_var','text'}]

        def regrid_columns():
            for i, b in enumerate(col_blocks):
                b["frame"].grid(row=0, column=i, sticky="ns",
                                padx=(0 if i == 0 else 8, 8), pady=4)
            inner.update_idletasks()
            _sync_scrollregion()

        def used_types():
            return {b['type_var'].get() for b in col_blocks}

        def _hint_for(key):
            if key == "BE":
                return "BE IDs: 10 digits, one per line"
            if key == "SONAR":
                return "Sonar: ID (digits) or URL starting with https://sonar-eu.amazon.com"
            return "Names (one per line)"

        def add_column(preset_key=None, preset_items=None):
            # default key = erster erlaubter, der noch nicht benutzt ist (oder NAME)
            default_key = None
            for k in self.DATASET_TYPES:
                if k not in used_types():
                    default_key = k
                    break
            key_init = preset_key or default_key or "NAME"

            frame = ttk.LabelFrame(inner, text=f"Column — {key_init}",
                                   style="AmazonCard.TLabelframe", padding=12)

            frame.grid_propagate(True)



            top = ttk.Frame(frame, style="Amazon.TFrame")
            top.grid(row=0, column=0, sticky="ew")
            top.columnconfigure(0, weight=1)

            ttk.Label(top, text="Type", style="AmazonBody.TLabel").grid(row=0, column=0, sticky="w")
            type_var = tk.StringVar(value=key_init)
            type_box = ttk.Combobox(top, state="readonly", values=list(self.DATASET_TYPES),
                                    textvariable=type_var, width=12)
            type_box.grid(row=0, column=1, sticky="w", padx=(8, 0))

            def remove_block():
                try:
                    col_blocks.remove(block)
                except ValueError:
                    pass
                frame.destroy()
                regrid_columns()

            rm_btn = self.create_secondary_button(top, "Remove", remove_block)
            rm_btn.grid(row=0, column=2, sticky="e", padx=(8, 0))

            # Hinweis passend zum Typ
            hint_var = tk.StringVar(value=_hint_for(type_var.get()))
            ttk.Label(frame, textvariable=hint_var, style="AmazonMuted.TLabel").grid(row=1, column=0, sticky="w", pady=(6, 2))

            # Text + eigene vertikale Scrollbar pro Spalte
            text_wrap = ttk.Frame(frame, style="Amazon.TFrame")
            text_wrap.grid(row=2, column=0, sticky="nsew")
            frame.rowconfigure(2, weight=1)
            text_wrap.columnconfigure(0, weight=1)

            txt = tk.Text(text_wrap, wrap="none", height=16,  width=40, font=("Segoe UI", 10))
            txt.grid(row=0, column=0, sticky="nsew")
            vscroll = ttk.Scrollbar(text_wrap, orient="vertical", command=txt.yview,
                                    style="Amazon.Vertical.TScrollbar")
            vscroll.grid(row=0, column=1, sticky="ns")
            txt.configure(yscrollcommand=vscroll.set)

            if preset_items:
                txt.insert("1.0", "\n".join(preset_items))

            def on_type_change(*_):
                k = type_var.get()
                frame.configure(text=f"Column — {k}")
                hint_var.set(_hint_for(k))

            type_var.trace_add("write", on_type_change)

            block = {"frame": frame, "type_var": type_var, "hint_var": hint_var, "text": txt}
            col_blocks.append(block)
            regrid_columns()

        # Vorbelegen (Edit) oder Standard (Create -> NAME)
        if initial and initial.get("columns"):
            for c in initial["columns"]:
                add_column(preset_key=c.get("key"), preset_items=c.get("items"))
        else:
            add_column(preset_key="NAME")

        # Footer
        actions = ttk.Frame(dlg, style="Amazon.TFrame")
        actions.pack(fill="x", padx=16, pady=12)

        def on_cancel():
            dlg.destroy()

        def on_save():
            name = (name_var.get() or "").strip()
            if not name:
                messagebox.showerror("Error", "Please enter a dataset name.")
                return

            # Name unique
            if mode == "create":
                if self._name_exists(name):
                    messagebox.showerror("Error", "Dataset name must be unique.")
                    return
            else:
                if self._name_exists(name, exclude_index=index):
                    messagebox.showerror("Error", "Dataset name must be unique.")
                    return

            if not col_blocks:
                messagebox.showerror("Error", "Add at least one column.")
                return

            # Sammeln + Validieren
            cols_map = {}  # key -> list[str]
            for b in col_blocks:
                key = b["type_var"].get()
                if key in cols_map:
                    messagebox.showerror("Error", f"Duplicate column type '{key}'. Each type may appear only once.")
                    return
                raw = [l.strip() for l in b["text"].get("1.0", "end").splitlines() if l.strip()]

                if key == "BE":
                    cleaned = []
                    for i, v in enumerate(raw, start=1):
                        digits = "".join(ch for ch in v if ch.isdigit())
                        if len(digits) != 10:
                            messagebox.showerror("Error", f"BE line {i}: must be exactly 10 digits (got '{v}').")
                            return
                        cleaned.append(digits)
                    cols_map[key] = cleaned

                elif key == "SONAR":
                    checked = []
                    for i, v in enumerate(raw, start=1):
                        v = v.strip()
                        if v.isdigit() or v.startswith("https://sonar-eu.amazon.com"):
                            checked.append(v)
                        else:
                            messagebox.showerror(
                                "Error",
                                f"SONAR line {i}: must be an ID (digits) or a URL starting with https://sonar-eu.amazon.com"
                            )
                            return
                    cols_map[key] = checked

                else:  # NAME
                    cols_map[key] = raw  # frei

            # Gleichlange Spalten (nur die vorhandenen vergleichen)
            lengths = [len(v) for v in cols_map.values()]
            if any(l == 0 for l in lengths):
                messagebox.showerror("Error", "Columns cannot be empty. Provide at least one line per column.")
                return
            if len(set(lengths)) > 1:
                messagebox.showerror("Error", f"All columns must have the same number of rows (got {lengths}).")
                return

            # In fester Reihenfolge speichern: NAME, BE, SONAR (nur vorhandene)
            columns = []
            for key in self.DATASET_TYPES:
                if key in cols_map:
                    columns.append({"key": key, "title": key, "items": cols_map[key]})

            dataset = {"name": name, "columns": columns}
            if mode == "create":
                self.datasets.append(dataset)
            else:
                self.datasets[index] = dataset

            self.save_datasets()
            self.refresh_datasets_view()
            dlg.destroy()

        self.create_secondary_button(actions, "Cancel", on_cancel).pack(side="right", padx=(0, 8))
        self.create_amazon_button(actions, "Save", on_save).pack(side="right")

        dlg.bind("<Escape>", lambda e: on_cancel())
        dlg.bind("<Control-Return>", lambda e: (on_save(), "break"))



    # ------------------- Profile: Storage & UI -------------------

    def load_profile(self):
        """Liest das (einzige) Profil oder gibt None zurück. Akzeptiert 'customerId' und 'customer_id'."""
        try:
            if os.path.exists(self.profile_file):
                with open(self.profile_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                if isinstance(data, dict) and data.get("alias") and data.get("email"):
                    prof = {"alias": data["alias"], "email": data["email"]}
                    cid = data.get("customerId", data.get("customer_id"))
                    try:
                        if cid not in (None, ""):
                            cid_int = int(str(cid).strip())
                            prof["customerId"] = cid_int
                            prof["customer_id"] = cid_int  # snake_case für Helper
                    except Exception:
                        pass
                    return prof
        except Exception:
            pass
        return None


    def save_profile(self, alias: str, email: str, customer_id: int | None = None):
        """Speichert das Profil (überschreibt vorhandenes) und exportiert Env Vars."""
        prof = {"alias": alias.strip(), "email": email.strip()}
        if customer_id is not None:
            cid = int(customer_id)
            prof["customerId"] = cid
            prof["customer_id"] = cid   # beide Schreibweisen speichern
        try:
            with open(self.profile_file, "w", encoding="utf-8") as f:
                json.dump(prof, f, ensure_ascii=False, indent=2)
            self.profile = prof
            self._export_profile_to_env()
            self.refresh_profile_ui()
        except Exception as e:
            messagebox.showerror("Profile", f"Could not save profile:\n{e}")

    def _export_profile_to_env(self):
        """Expose current profile via environment variables used by helper scripts."""
        try:
            prof = self.profile or {}
            os.environ["AMZN_ALIAS"] = str(prof.get("alias") or "")
            os.environ["AMZN_EMAIL"] = str(prof.get("email") or "")
            cid = prof.get("customerId") or prof.get("customer_id")
            if cid not in (None, ""):
                os.environ["AMZN_CUSTOMER_ID"] = str(cid)
            else:
                os.environ.pop("AMZN_CUSTOMER_ID", None)
        except Exception:
            pass




    def delete_profile(self):
        """Löscht das Profil von der Platte und aus dem Speicher."""
        try:
            if os.path.exists(self.profile_file):
                os.remove(self.profile_file)
        except Exception:
            pass
        self.profile = None
        self._export_profile_to_env()
        self.refresh_profile_ui()


    def refresh_profile_ui(self):
        """Aktualisiert Label + Button-Text im Header."""
        if self.profile:
            self.hello_label.config(text=f"Hello, {self.profile.get('alias','')}")
            self.profile_btn.config(text="Profile")
        else:
            self.hello_label.config(text="")
            self.profile_btn.config(text="Create profile")

    def on_profile_click(self):
        """
        Wenn kein Profil existiert: Create-Dialog.
        Wenn eins existiert: View-Dialog mit Alias/Email + Edit + Delete.
        """
        if not self.profile:
            data = self._open_profile_dialog(mode="create")
            if data:
                self.save_profile(data["alias"], data["email"], data.get("customerId"))
            return

        # View/Edit/Delete Dialog
        dlg = tk.Toplevel(self.root)
        dlg.title("Profile")
        dlg.configure(bg=self.AMAZON["bg"])
        dlg.geometry("420x240")
        dlg.minsize(400, 220)
        dlg.transient(self.root)
        dlg.grab_set()

        dlg.update_idletasks()
        x = self.root.winfo_rootx() + (self.root.winfo_width() // 2 - dlg.winfo_width() // 2)
        y = self.root.winfo_rooty() + (self.root.winfo_height() // 2 - dlg.winfo_height() // 2)
        dlg.geometry(f"+{max(0, x)}+{max(0, y)}")

        wrap = tk.Frame(dlg, bg=self.AMAZON["bg"])
        wrap.pack(fill="both", expand=True, padx=16, pady=16)

        ttk.Label(wrap, text="Alias", style="AmazonMuted.TLabel").pack(anchor="w")
        ttk.Label(wrap, text=self.profile.get("alias",""), style="AmazonTitle.TLabel").pack(anchor="w", pady=(0, 8))

        ttk.Label(wrap, text="Email", style="AmazonMuted.TLabel").pack(anchor="w")
        ttk.Label(wrap, text=self.profile.get("email",""), style="AmazonBody.TLabel").pack(anchor="w")

        ttk.Label(wrap, text="Customer Id", style="AmazonMuted.TLabel").pack(anchor="w", pady=(8, 0))
        ttk.Label(wrap, text=str(self.profile.get("customerId","")), style="AmazonBody.TLabel").pack(anchor="w")


        actions = ttk.Frame(dlg, style="Amazon.TFrame")
        actions.pack(fill="x", padx=16, pady=(8, 12))

        def do_edit():
            data = self._open_profile_dialog(mode="edit", initial=self.profile)
            if data:
                self.save_profile(data["alias"], data["email"], data.get("customerId"))
            dlg.destroy()

        def do_delete():
            if messagebox.askyesno("Delete profile", "Delete the current profile?"):
                self.delete_profile()
                dlg.destroy()

        self.create_secondary_button(actions, "Delete", do_delete).pack(side="right", padx=(8, 0))
        self.create_amazon_button(actions, "Edit", do_edit).pack(side="right")
        self.create_secondary_button(actions, "Close", dlg.destroy).pack(side="left")

        dlg.bind("<Escape>", lambda e: dlg.destroy())

    def _open_profile_dialog(self, mode="create", initial=None):
        """
        Modal-Dialog zum Anlegen/Bearbeiten des Profils.
        mode: 'create' | 'edit'
        Rückgabe: {'alias': str, 'email': str, 'customerId': int|None} oder None
        """
        assert mode in ("create", "edit")
        dlg = tk.Toplevel(self.root)
        dlg.title("Create profile" if mode == "create" else "Edit profile")
        dlg.configure(bg=self.AMAZON["bg"])
        dlg.geometry("460x320")  # etwas höher wegen neuem Feld
        dlg.minsize(440, 300)
        dlg.transient(self.root)
        dlg.grab_set()

        # zentrieren
        dlg.update_idletasks()
        x = self.root.winfo_rootx() + (self.root.winfo_width() // 2 - dlg.winfo_width() // 2)
        y = self.root.winfo_rooty() + (self.root.winfo_height() // 2 - dlg.winfo_height() // 2)
        dlg.geometry(f"+{max(0, x)}+{max(0, y)}")

        body = ttk.Frame(dlg, style="Amazon.TFrame")
        body.pack(fill="both", expand=True, padx=16, pady=12)

        ttk.Label(body, text="Alias", style="AmazonBody.TLabel").grid(row=0, column=0, sticky="w")
        alias_var = tk.StringVar(value=(initial.get("alias") if (initial and initial.get("alias")) else ""))
        alias_entry = ttk.Entry(body, textvariable=alias_var)
        alias_entry.grid(row=1, column=0, sticky="ew", pady=(0, 8))
        alias_entry.focus_set()

        ttk.Label(body, text="Email", style="AmazonBody.TLabel").grid(row=2, column=0, sticky="w")
        email_var = tk.StringVar(value=(initial.get("email") if (initial and initial.get("email")) else ""))
        email_entry = ttk.Entry(body, textvariable=email_var)
        email_entry.grid(row=3, column=0, sticky="ew", pady=(0, 8))

        # --- neu: Customer Id (Zahl, optional) ---
        ttk.Label(body, text="Customer Id (number)", style="AmazonBody.TLabel").grid(row=4, column=0, sticky="w")
        cust_init = ""
        if initial and initial.get("customerId") is not None:
            cust_init = str(initial.get("customerId"))
        cust_var = tk.StringVar(value=cust_init)
        cust_entry = ttk.Entry(body, textvariable=cust_var)
        cust_entry.grid(row=5, column=0, sticky="ew")

        body.columnconfigure(0, weight=1)







        # ---------- Actions (Save / Cancel) ----------
        actions = ttk.Frame(dlg, style="Amazon.TFrame")
        actions.pack(fill="x", padx=16, pady=12)   # WICHTIG: pack (nicht grid) im selben Parent!

        result = {"alias": None, "email": None, "customerId": None}

        def is_valid_email(s: str) -> bool:
            s = s.strip()
            return bool(re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", s))

        def submit():
            alias = (alias_var.get() or "").strip()
            email = (email_var.get() or "").strip()
            cust  = (cust_var.get() or "").strip()

            if not alias:
                messagebox.showerror("Profile", "Please enter an alias.")
                return
            if not is_valid_email(email):
                messagebox.showerror("Profile", "Please enter a valid email address.")
                return

            customer_id = None
            if cust:
                if not cust.isdigit():
                    messagebox.showerror("Profile", "Customer Id must be a number.")
                    return
                customer_id = int(cust)

            result["alias"] = alias
            result["email"] = email
            result["customerId"] = customer_id
            dlg.destroy()

        def cancel():
            dlg.destroy()

        # Buttons rechts ausrichten
        self.create_secondary_button(actions, "Cancel", cancel).pack(side="right", padx=(0, 8))
        self.create_amazon_button(actions, "Save", submit).pack(side="right")

        # Shortcuts
        dlg.bind("<Escape>", lambda e: cancel())
        dlg.bind("<Return>", lambda e: submit())

        dlg.wait_window()
        if result["alias"] and result["email"]:
            return result
        return None




    # ------------------- Sonar Templates: Storage & UI -------------------

    def load_templates(self):
        """
        Liest gespeicherte Sonar-Templates aus ~/.bullseye_automation/sonar_templates.json
        Schema je Template:
          {
            "name": str,
            "channel": str,
            "teamBindle": str,
            "lobExpression": str,
            "managementType": str,
            "businessGroupId": int,
            "familyId": int,
            "optOuts": list[str],
            "startTimeMinutesOffset": int,   # 0..1439
            "endTimeMinutesOffset": int      # 0..1439, > start
          }
        Fixe (nicht speicherbar, immer gleich): duration=1, reason="OneShot", topic="CAFEP",
          communicationContentType.optOutList=[]
        """
        try:
            if os.path.exists(self.templates_file):
                with open(self.templates_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                if isinstance(data, list):
                    # Mini-Migration: alte Keys entfernen/ignorieren
                    for t in data:
                        if isinstance(t, dict):
                            t.pop("communicationContentType", None)  # optOutList ist fix []
                    return data
        except Exception:
            pass
        return []

    def save_templates(self):
        try:
            with open(self.templates_file, "w", encoding="utf-8") as f:
                json.dump(self.templates, f, ensure_ascii=False, indent=2)
        except Exception as e:
            messagebox.showerror("Templates", f"Could not save templates:\n{e}")

    def _template_name_exists(self, name: str, exclude_index: int = None) -> bool:
        n = (name or "").strip().lower()
        for i, t in enumerate(self.templates or []):
            if exclude_index is not None and i == exclude_index:
                continue
            if (t.get("name") or "").strip().lower() == n:
                return True
        return False

    @staticmethod
    def _minutes_to_hhmm(m: int) -> str:
        try:
            m = int(m)
            if m < 0:
                m = 0
            h = m // 60
            mm = m % 60
            return f"{h:02d}:{mm:02d}"
        except Exception:
            return "--:--"

    @staticmethod
    def _minutes_to_ampm(m: int) -> str:
        """z. B. 540 -> '9:00 AM'"""
        try:
            m = int(m) % 1440
        except Exception:
            return ""
        h = m // 60
        mi = m % 60
        ap = "AM" if h < 12 else "PM"
        h12 = h % 12
        if h12 == 0:
            h12 = 12
        return f"{h12}:{mi:02d} {ap}"

    @staticmethod
    def _ampm_to_minutes(s: str) -> int:
        """
        Akzeptiert:
          - '9', '9 AM', '9:15', '9:15 PM'  (12h)
          - '21:30'                          (24h)
        Liefert Minuten seit 00:00 (0..1439) oder wirft ValueError.
        """
        import re as _re
        if not s:
            raise ValueError("Empty time")
        txt = s.strip().upper()
        # 12h: H(:MM)?(AM|PM)
        m = _re.match(r"^(\d{1,2})(?::(\d{1,2}))?\s*(AM|PM)?$", txt)
        if not m:
            raise ValueError("Invalid time format. Use e.g. '9:00 AM' or '21:30'.")
        h = int(m.group(1))
        mi = int(m.group(2) or 0)
        ap = m.group(3)  # None -> 24h
        if mi < 0 or mi > 59:
            raise ValueError("Minutes must be 0..59.")
        if ap:
            if h < 1 or h > 12:
                raise ValueError("Hour must be 1..12 for AM/PM.")
            if h == 12:
                h = 0
            if ap == "PM":
                h += 12
        else:
            if h < 0 or h > 23:
                raise ValueError("Hour must be 0..23.")
        total = h * 60 + mi
        if total < 0 or total > 1439:
            raise ValueError("Time must be within a single day.")
        return total




    @staticmethod
    def _parse_list_input(text: str):
        """
        Erlaubt:
          - JSON: ["a","b"]
          - Komma/Strichpunkt-getrennt: a, b ; c
          - Leer -> []
        """
        s = (text or "").strip()
        if not s:
            return []
        try:
            val = json.loads(s)
            if isinstance(val, list):
                return [str(x).strip() for x in val if str(x).strip()]
        except Exception:
            pass
        import re as _re
        parts = [p.strip() for p in _re.split(r"[;,]", s) if p.strip()]
        return parts

    def open_templates_manager(self):
        """Hauptdialog: Liste + Create/Edit/Delete"""
        dlg = tk.Toplevel(self.root)
        dlg.title("Templates")
        dlg.configure(bg=self.AMAZON["bg"])
        dlg.geometry("680x420")
        dlg.minsize(640, 380)
        dlg.transient(self.root)
        dlg.grab_set()

        # center
        dlg.update_idletasks()
        x = self.root.winfo_rootx() + (self.root.winfo_width() // 2 - dlg.winfo_width() // 2)
        y = self.root.winfo_rooty() + (self.root.winfo_height() // 2 - dlg.winfo_height() // 2)
        dlg.geometry(f"+{max(0, x)}+{max(0, y)}")

        hdr = tk.Frame(dlg, bg=self.AMAZON["bg"])
        hdr.pack(fill="x", padx=16, pady=(14, 6))
        ttk.Label(hdr, text="Sonar Templates", style="AmazonTitle.TLabel").pack(anchor="w")
        ttk.Label(hdr, text="Create, edit, or delete templates for Sonar creation.",
                  style="AmazonSubtitle.TLabel").pack(anchor="w", pady=(2, 0))

        body = ttk.Frame(dlg, style="Amazon.TFrame")
        body.pack(fill="both", expand=True, padx=16, pady=(8, 0))
        body.columnconfigure(0, weight=1)
        body.rowconfigure(0, weight=1)

        tv = ttk.Treeview(body, columns=("name", "channel", "window"),
                          show="headings", style="Amazon.Treeview")
        tv.heading("name", text="Name")
        tv.heading("channel", text="Channel")
        tv.heading("window", text="Time Window")
        tv.column("name", width=260, anchor="w")
        tv.column("channel", width=130, anchor="center")
        tv.column("window", width=130, anchor="center")
        tv.grid(row=0, column=0, sticky="nsew", padx=(0, 8), pady=(0, 8))

        vs = ttk.Scrollbar(body, orient="vertical", command=tv.yview, style="Amazon.Vertical.TScrollbar")
        vs.grid(row=0, column=1, sticky="ns", pady=(0, 8))
        tv.configure(yscrollcommand=vs.set)

        def refresh():
            for iid in tv.get_children():
                tv.delete(iid)
            for i, t in enumerate(self.templates):
                w = f"{self._minutes_to_hhmm(t.get('startTimeMinutesOffset', 0))}–{self._minutes_to_hhmm(t.get('endTimeMinutesOffset', 0))}"
                tv.insert("", "end", iid=str(i),
                          values=(t.get("name", ""), t.get("channel", ""), w))
            btn_edit.configure(state="normal" if tv.selection() else "disabled")
            btn_del.configure(state="normal" if tv.selection() else "disabled")

        def on_select(_=None):
            btn_edit.configure(state="normal" if tv.selection() else "disabled")
            btn_del.configure(state="normal" if tv.selection() else "disabled")

        tv.bind("<<TreeviewSelect>>", on_select)
        tv.bind("<Double-1>", lambda e: do_edit())

        actions = ttk.Frame(dlg, style="Amazon.TFrame")
        actions.pack(fill="x", padx=16, pady=12)

        def do_create():
            data = self.open_template_editor(mode="create")
            if data:
                self.templates.append(data)
                self.save_templates()
                refresh()

        def do_edit():
            sel = tv.selection()
            if not sel:
                return
            idx = int(sel[0])
            data = self.open_template_editor(mode="edit", initial=self.templates[idx], index=idx)
            if data:
                self.templates[idx] = data
                self.save_templates()
                refresh()

        def do_delete():
            sel = tv.selection()
            if not sel:
                return
            idx = int(sel[0])
            name = self.templates[idx].get("name", "Unnamed")
            if messagebox.askyesno("Delete Template", f"Delete template '{name}'?"):
                del self.templates[idx]
                self.save_templates()
                refresh()

        btn_del = self.create_secondary_button(actions, "Delete", do_delete)
        btn_del.pack(side="right", padx=(8, 0))

        btn_edit = self.create_amazon_button(actions, "Edit", do_edit)
        btn_edit.pack(side="right", padx=(8, 0))

        btn_create = self.create_amazon_button(actions, "Create", do_create)
        btn_create.pack(side="right")

        self.create_secondary_button(actions, "Close", dlg.destroy).pack(side="left")

        refresh()
        dlg.bind("<Escape>", lambda e: dlg.destroy())


    def open_template_editor(self, mode="create", initial=None, index=None):
        """Modaler Editor für 1 Template. Rückgabe: dict oder None"""
        assert mode in ("create", "edit")
        t = initial or {}

        dlg = tk.Toplevel(self.root)
        dlg.title("Create Template" if mode == "create" else "Edit Template")
        dlg.configure(bg=self.AMAZON["bg"])
        dlg.geometry("620x560")
        dlg.minsize(600, 520)
        dlg.transient(self.root)
        dlg.grab_set()

        dlg.update_idletasks()
        x = self.root.winfo_rootx() + (self.root.winfo_width() // 2 - dlg.winfo_width() // 2)
        y = self.root.winfo_rooty() + (self.root.winfo_height() // 2 - dlg.winfo_height() // 2)
        dlg.geometry(f"+{max(0, x)}+{max(0, y)}")

        # Header
        hdr = tk.Frame(dlg, bg=self.AMAZON["bg"])
        hdr.pack(fill="x", padx=16, pady=(14, 6))
        ttk.Label(hdr, text="Template", style="AmazonTitle.TLabel").pack(anchor="w")
        ttk.Label(
            hdr,
            text="Define defaults used when creating Sonar programs/campaigns.",
            style="AmazonSubtitle.TLabel"
        ).pack(anchor="w", pady=(2, 0))

        # Input source (Manual vs Import)
        mode_wrap = ttk.Frame(dlg, style="Amazon.TFrame")
        mode_wrap.pack(fill="x", padx=16, pady=(0, 8))
        mode_var = tk.StringVar(value="manual")
        ttk.Label(mode_wrap, text="Input source:", style="AmazonBody.TLabel").pack(side="left")
        ttk.Radiobutton(mode_wrap, text="Manual", variable=mode_var, value="manual",
                        style="Amazon.TRadiobutton").pack(side="left", padx=(8, 0))
        ttk.Radiobutton(mode_wrap, text="Import from existing Sonar campaign",
                        variable=mode_var, value="import",
                        style="Amazon.TRadiobutton").pack(side="left", padx=(12, 0))

        # Body (Grid)
        body = ttk.Frame(dlg, style="Amazon.TFrame")
        body.pack(fill="both", expand=True, padx=16, pady=(8, 0))
        for c in range(2):
            body.columnconfigure(c, weight=1)

        # Name
        ttk.Label(body, text="Name", style="AmazonBody.TLabel").grid(row=0, column=0, sticky="w")
        name_var = tk.StringVar(value=t.get("name", ""))          # keine Defaults
        name_entry = ttk.Entry(body, textvariable=name_var)
        name_entry.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(0, 8))
        name_entry.focus_set()

        # --- Import-Bereich direkt unter Name -------------------------------
        import_label = ttk.Label(body, text="Sonar link (campaign) or campaign ID:", style="AmazonBody.TLabel")

        import_label.grid(row=2, column=0, columnspan=2, sticky="w")

        url_var = tk.StringVar()
        url_entry = ttk.Entry(body, textvariable=url_var)
        url_entry.grid(row=3, column=0, columnspan=2, sticky="ew")

        import_actions = ttk.Frame(body, style="Amazon.TFrame")
        import_actions.grid(row=4, column=0, columnspan=2, sticky="e", pady=(6, 8))
        self.create_amazon_button(import_actions, "Fetch & Prefill",
                                  lambda: do_fetch_import()).pack(side="right")

        # --- Restliche Template-Felder -------------------------------------
        # Channel + Management Type (keine Defaults)
        ttk.Label(body, text="Channel", style="AmazonBody.TLabel").grid(row=5, column=0, sticky="w")
        channel_var = tk.StringVar(value=t.get("channel", ""))    # leer
        cb_channel = ttk.Combobox(body, state="readonly",
                                  values=["MOBILE_PUSH", "EMAIL"],
                                  textvariable=channel_var)
        cb_channel.grid(row=6, column=0, sticky="ew", pady=(0, 8))

        ttk.Label(body, text="Management Type", style="AmazonBody.TLabel").grid(row=5, column=1, sticky="w")
        mgmt_var = tk.StringVar(value=t.get("managementType", ""))  # frei & leer
        ttk.Entry(body, textvariable=mgmt_var).grid(row=6, column=1, sticky="ew", pady=(0, 8))

        # teamBindle + lobExpression
        ttk.Label(body, text="teamBindle", style="AmazonBody.TLabel").grid(row=7, column=0, sticky="w")
        team_var = tk.StringVar(value=t.get("teamBindle", ""))
        ttk.Entry(body, textvariable=team_var).grid(row=8, column=0, sticky="ew", pady=(0, 8))

        ttk.Label(body, text="lobExpression", style="AmazonBody.TLabel").grid(row=7, column=1, sticky="w")
        lob_var = tk.StringVar(value=t.get("lobExpression", ""))
        ttk.Entry(body, textvariable=lob_var).grid(row=8, column=1, sticky="ew", pady=(0, 8))

        # businessGroupId + familyId
        ttk.Label(body, text="businessGroupId", style="AmazonBody.TLabel").grid(row=9, column=0, sticky="w")
        bgid_var = tk.StringVar(value=str(t.get("businessGroupId", "")).strip())
        ttk.Entry(body, textvariable=bgid_var).grid(row=10, column=0, sticky="ew", pady=(0, 8))

        ttk.Label(body, text="familyId", style="AmazonBody.TLabel").grid(row=9, column=1, sticky="w")
        fam_var = tk.StringVar(value=str(t.get("familyId", "")).strip())
        ttk.Entry(body, textvariable=fam_var).grid(row=10, column=1, sticky="ew", pady=(0, 8))

        # optOuts – leer
        ttk.Label(body, text="optOuts (JSON oder a,b,c)", style="AmazonBody.TLabel").grid(
            row=11, column=0, sticky="w", columnspan=2
        )
        opt_text = tk.Text(body, height=3, wrap="none")
        opt_text.grid(row=12, column=0, columnspan=2, sticky="nsew", pady=(0, 8))

        # Zeitfenster – leer
        # Zeitfenster – Anzeige als '9:00 AM/PM' (intern speichern wir Minuten)
        ttk.Label(body, text="Start time (e.g., 9:00 AM or 21:30)", style="AmazonBody.TLabel").grid(
            row=13, column=0, sticky="w"
        )
        # vorhandene Minuten -> hübsch anzeigen, sonst leer
        _st_init = t.get("startTimeMinutesOffset", "")
        start_time_str_var = tk.StringVar(
            value=(self._minutes_to_ampm(_st_init) if str(_st_init).strip().isdigit() else "")
        )
        ttk.Entry(body, textvariable=start_time_str_var).grid(row=14, column=0, sticky="ew", pady=(0, 8))

        ttk.Label(body, text="End time (e.g., 9:00 PM or 21:00)", style="AmazonBody.TLabel").grid(
            row=13, column=1, sticky="w"
        )
        _en_init = t.get("endTimeMinutesOffset", "")
        end_time_str_var = tk.StringVar(
            value=(self._minutes_to_ampm(_en_init) if str(_en_init).strip().isdigit() else "")
        )
        ttk.Entry(body, textvariable=end_time_str_var).grid(row=14, column=1, sticky="ew", pady=(0, 8))



        
        # Import-Funktion (benötigt oben definierte Vars)
        def do_fetch_import():
            link = (url_var.get() or "").strip()
            if not link:
                messagebox.showerror("Import", "Bitte einen Sonar-Campaign-Link oder eine Campaign-ID einfügen.")
                return
            try:
                # updated Variante verwenden
                prog = self._templates_import_from_campaign(link)
            except Exception as e:
                messagebox.showerror("Import", f"Konnte nicht laden:\n{e}")
                return

            # Minimal-Validierung
            if not isinstance(prog, dict) or not prog:
                messagebox.showerror("Import", "Antwort ungültig oder leer.")
                return

            # Felder aus der Response übernehmen
            ch = (prog.get("channel") or "").strip().upper()
            if ch in ("MOBILE_PUSH", "EMAIL"):
                channel_var.set(ch)

            mgmt_var.set(str(prog.get("managementType") or ""))
            team_var.set(str(prog.get("teamBindle") or ""))
            lob_var.set(str(prog.get("lobExpression") or ""))

            if prog.get("businessGroupId") is not None:
                bgid_var.set(str(prog.get("businessGroupId") or "").strip())
            if prog.get("familyId") is not None:
                fam_var.set(str(prog.get("familyId") or "").strip())

            if prog.get("startTimeMinutesOffset") is not None:
                start_time_str_var.set(self._minutes_to_ampm(int(prog["startTimeMinutesOffset"])))
            if prog.get("endTimeMinutesOffset") is not None:
                end_time_str_var.set(self._minutes_to_ampm(int(prog["endTimeMinutesOffset"])))

            # optOuts bleiben absichtlich leer
            messagebox.showinfo("Import", "Template-Felder wurden aus der Campaign übernommen.")



        # Import-Bereich ein-/ausblenden
        def _toggle_import_panel(*_):
            widgets = (import_label, url_entry, import_actions)
            if mode_var.get() == "import":
                for w in widgets:
                    w.grid()  # letzte Grid-Optionen verwenden
                url_entry.focus_set()
            else:
                for w in widgets:
                    w.grid_remove()
        mode_var.trace_add("write", _toggle_import_panel)
        _toggle_import_panel()

        # Footer
        actions = ttk.Frame(dlg, style="Amazon.TFrame")
        actions.pack(fill="x", padx=16, pady=12)

        result = {"value": None}

        def do_save():
            name = (name_var.get() or "").strip()
            if not name:
                messagebox.showerror("Template", "Please enter a template name.")
                return
            if mode == "create":
                if self._template_name_exists(name):
                    messagebox.showerror("Template", "Template name must be unique.")
                    return
            else:
                if self._template_name_exists(name, exclude_index=index):
                    messagebox.showerror("Template", "Template name must be unique.")
                    return

            # Zahlenfelder validieren
            # Zahlenfelder validieren (bg/family als int, Zeiten aus '9:00 AM/PM' -> Minuten)
            try:
                bgid = int((bgid_var.get() or "").strip())
                fam = int((fam_var.get() or "").strip())
                st = self._ampm_to_minutes(start_time_str_var.get())
                en = self._ampm_to_minutes(end_time_str_var.get())
            except Exception as ex:
                messagebox.showerror("Template", f"Invalid input:\n{ex}")
                return

            if not (0 <= st <= 1439 and 0 <= en <= 1439):
                messagebox.showerror("Template", "Time offsets must be between 0 and 1439.")
                return
            if not (st < en):
                messagebox.showerror("Template", "End time must be greater than start time.")
                return

            # optOuts parsen (leer erlaubt)
            opt_outs = self._parse_list_input(opt_text.get("1.0", "end"))

            data = {
                "name": name,
                "channel": (channel_var.get() or "").strip(),
                "teamBindle": (team_var.get() or "").strip(),
                "lobExpression": (lob_var.get() or "").strip(),
                "managementType": (mgmt_var.get() or "").strip(),
                "businessGroupId": bgid,
                "familyId": fam,
                "optOuts": opt_outs,
                "startTimeMinutesOffset": st,
                "endTimeMinutesOffset": en,
            }
            result["value"] = data
            dlg.destroy()

        def do_cancel():
            dlg.destroy()

        self.create_secondary_button(actions, "Cancel", do_cancel).pack(side="right", padx=(0, 8))
        self.create_amazon_button(actions, "Save", do_save).pack(side="right")

        dlg.bind("<Escape>", lambda e: do_cancel())
        dlg.bind("<Control-Return>", lambda e: (do_save(), "break"))

        dlg.wait_window()
        return result["value"]





    # ------------------- Core Actions -------------------

    def proceed(self):
        selected_function = self.function_var.get()
        self.status_var.set(f"Selected function: {selected_function}")


        if selected_function == "sonar":
            self.upload_be_to_sonar()
            return
        elif selected_function == "create_os_sonar":
            self.create_os_sonar()
            return
        elif selected_function == "clone_and_publish":
            self.clone_and_publish()
            return
        elif selected_function == "mass_clone_fixed":
            self.mass_clone_fixed()   # <--- Handler unten
            return
        elif selected_function == "clone_across_mps":           
            self.clone_across_mps()                               
            return
        elif selected_function == "update_content":
            self.update_campaign_content()
            return
        elif selected_function == "send_preview":
            self.send_preview()
            return
        elif selected_function == "approve_sonar":
            self.approve_sonar()
            return
        elif selected_function == "create_rc_sonar":
            self.create_rc_sonar()
            return

        be_ids = self.get_be_ids()
        if not be_ids:
            self.status_var.set("No BE IDs entered.")
            return

        # Reset progress
        self.progress_var.set(0)
        self.progress_label.config(text="Starting...")
        start_time = time.time()

        if selected_function == "sizes":
            self.get_sizes(be_ids)
        elif selected_function == "queue":
            self.queue_segments(be_ids)
        elif selected_function == "rules":
            self.extract_rules(be_ids)

    def show_mass_clone_fixed_dialog(self):
        """
        Dialog: Basis-BE-ID wählen (Default ODER Custom), Anzahl, und Namen (eine Zeile pro Clone).
        Rückgabe: {"base_id": "##########", "count": int, "names": [str, ...]} oder None bei Abbruch.
        """
        import tkinter as tk
        from tkinter import ttk, messagebox

        
        dlg = tk.Toplevel(self.root)
        dlg.title("Mass Clone — Base ID, Count, Names")
        dlg.configure(bg=self.AMAZON["bg"])
        dlg.geometry("700x520")
        dlg.minsize(660, 480)
        dlg.transient(self.root)
        dlg.grab_set()

        dlg.update_idletasks()
        x = self.root.winfo_rootx() + (self.root.winfo_width() // 2 - dlg.winfo_width() // 2)
        y = self.root.winfo_rooty() + (self.root.winfo_height() // 2 - dlg.winfo_height() // 2)
        dlg.geometry(f"+{max(0, x)}+{max(0, y)}")

        # Header
        hdr = tk.Frame(dlg, bg=self.AMAZON["bg"])
        hdr.pack(fill="x", padx=16, pady=(14, 6))
        ttk.Label(hdr, text="Bulk Clone from Base Segment", style="AmazonTitle.TLabel").pack(anchor="w")
        ttk.Label(
            hdr,
            text="Wähle die Basis-BE-ID (Default oder eigene), gib die Anzahl der Kopien an und liste die Namen (eine Zeile pro Clone).",
            style="AmazonSubtitle.TLabel"
        ).pack(anchor="w", pady=(2, 0))

        # Body
        body = tk.Frame(dlg, bg=self.AMAZON["bg"])
        body.pack(fill="both", expand=True, padx=16, pady=(8, 0))

        # --- Base-ID Auswahl (Radio + Entry) ---
        base_group = tk.LabelFrame(body, text="Base Segment", bg=self.AMAZON["bg"])
        base_group.pack(fill="x", pady=(0, 10))

        choice_var = tk.StringVar(value="default")
        ttk.Radiobutton(base_group, text=f"Default: {MASS_CLONE_FIXED_BASE_BE}", variable=choice_var, value="default",
                        style="Amazon.TRadiobutton").grid(row=0, column=0, sticky="w", padx=8, pady=6)


        ttk.Radiobutton(base_group, text="Eigene Base BE-ID:", variable=choice_var, value="custom",
                        style="Amazon.TRadiobutton").grid(row=1, column=0, sticky="w", padx=8, pady=(0, 8))

        custom_var = tk.StringVar()
        custom_entry = ttk.Entry(base_group, textvariable=custom_var, width=24)
        custom_entry.grid(row=1, column=1, sticky="w", padx=(0, 8), pady=(0, 8))

        def toggle_custom_state(*_):
            if choice_var.get() == "custom":
                custom_entry.configure(state="normal")
                custom_entry.focus_set()
            else:
                custom_entry.configure(state="disabled")
        toggle_custom_state()
        choice_var.trace_add("write", lambda *_: toggle_custom_state())

        # --- Anzahl ---
        count_group = tk.Frame(body, bg=self.AMAZON["bg"])
        count_group.pack(fill="x", pady=(0, 8))
        ttk.Label(count_group, text="Anzahl der Clones:", style="AmazonBody.TLabel").grid(row=0, column=0, sticky="w")
        count_var = tk.StringVar(value="1")
        try:
            count_spin = ttk.Spinbox(count_group, from_=1, to=999, textvariable=count_var, width=8)
        except Exception:
            count_spin = ttk.Entry(count_group, textvariable=count_var, width=8)
        count_spin.grid(row=0, column=1, sticky="w", padx=(8, 0))

        # --- Namen (eine Zeile pro Clone) ---
        names_group = tk.Frame(body, bg=self.AMAZON["bg"])
        names_group.pack(fill="both", expand=True, pady=(8, 0))
        ttk.Label(names_group, text="Namen (eine Zeile pro Clone):", style="AmazonBody.TLabel").pack(anchor="w")
        names_text = tk.Text(names_group, wrap="none", height=14, font=("Segoe UI", 10))
        names_text.pack(fill="both", expand=True)

        # Actions
        actions = tk.Frame(dlg, bg=self.AMAZON["bg"])
        actions.pack(fill="x", padx=16, pady=12)

        result = {"base_id": None, "count": None, "names": None}

        def validate_and_collect():
            # Base-ID
            if choice_var.get() == "default":
                base_id = MASS_CLONE_FIXED_BASE_BE
            else:
                base_id = "".join(ch for ch in custom_var.get().strip() if ch.isdigit())
                if len(base_id) != 10:
                    messagebox.showerror("Error", "Eigene Base BE-ID muss 10-stellig sein.")
                    return

            # Count
            try:
                n = int(count_var.get().strip())
                if n < 1:
                    raise ValueError
            except Exception:
                messagebox.showerror("Error", "Anzahl muss eine positive Zahl sein.")
                return

            # Names
            raw_lines = names_text.get("1.0", "end").splitlines()
            names = [ln.strip() for ln in raw_lines if ln.strip()]
            if len(names) != n:
                messagebox.showerror("Error", f"Bitte genau {n} Name(n) eingeben (eine Zeile pro Clone).")
                return

            result["base_id"] = base_id
            result["count"] = n
            result["names"] = names
            dlg.destroy()

        def cancel():
            result["base_id"] = None
            result["count"] = None
            result["names"] = None
            dlg.destroy()

        self.create_secondary_button(actions, "Cancel", cancel).pack(side="right", padx=(0, 8))
        self.create_amazon_button(actions, "OK", validate_and_collect).pack(side="right")

        # Shortcuts: Esc = Cancel, Ctrl+Enter = OK (Enter bleibt Zeilenumbruch)
        dlg.bind("<Escape>", lambda e: cancel())

        def on_ctrl_enter(e):
            validate_and_collect()
            return "break"
        dlg.bind("<Control-Return>", on_ctrl_enter)

        dlg.wait_window()
        if result["base_id"] and result["count"] and result["names"]:
            return result
        return None


    def show_single_be_dialog(self):
        """Amazon-style Dialog: genau eine 10-stellige BE-ID abfragen. Enter wechselt NICHT automatisch."""
        dlg = tk.Toplevel(self.root)
        dlg.title("Clone across MPs — Source BE ID")
        dlg.configure(bg=self.AMAZON["bg"])
        dlg.geometry("460x220")
        dlg.minsize(440, 200)
        dlg.transient(self.root)
        dlg.grab_set()

        # center on parent
        dlg.update_idletasks()
        x = self.root.winfo_rootx() + (self.root.winfo_width() // 2 - dlg.winfo_width() // 2)
        y = self.root.winfo_rooty() + (self.root.winfo_height() // 2 - dlg.winfo_height() // 2)
        dlg.geometry(f"+{max(0, x)}+{max(0, y)}")

        hdr = tk.Frame(dlg, bg=self.AMAZON["bg"])
        hdr.pack(fill="x", padx=16, pady=(14, 6))
        ttk.Label(hdr, text="Enter the source BE ID (10 digits)", style="AmazonTitle.TLabel").pack(anchor="w")
        ttk.Label(
            hdr,
            text="This segment will be cloned to all other marketplaces (UK/DE/FR/IT/ES).",
            style="AmazonSubtitle.TLabel"
        ).pack(anchor="w", pady=(2, 0))

        body = tk.Frame(dlg, bg=self.AMAZON["bg"])
        body.pack(fill="both", expand=True, padx=16, pady=(8, 0))

        entry = tk.Entry(body, font=("Segoe UI", 11))
        entry.pack(fill="x")
        entry.focus_set()

        actions = tk.Frame(dlg, bg=self.AMAZON["bg"])
        actions.pack(fill="x", padx=16, pady=12)

        result = {"be_id": None}

        def submit():
            raw = (entry.get() or "").strip()
            digits = "".join(ch for ch in raw if ch.isdigit())
            if len(digits) != 10:
                messagebox.showerror("Error", "BE ID must be exactly 10 digits.")
                return
            result["be_id"] = digits
            dlg.destroy()

        def cancel():
            result["be_id"] = None
            dlg.destroy()

        self.create_secondary_button(actions, "Cancel", cancel).pack(side="right", padx=(0, 8))
        self.create_amazon_button(actions, "Confirm", submit).pack(side="right")

        # Wichtig: Enter NICHT auf OK binden
        dlg.bind("<Escape>", lambda e: cancel())

        dlg.wait_window()
        return result["be_id"]




    def show_be_input_dialog(self):
        """
        Dialog für BE-IDs (v2): Links manuell, rechts Auswahl aus BE-Datasets (columns['BE']).
        Rückgabe: kombinierter Rohtext (manuell + ausgewählte Datasets) oder None bei Abbruch.
        """
        dlg = tk.Toplevel(self.root)
        dlg.title("Enter BE IDs")
        dlg.configure(bg=self.AMAZON["bg"])
        dlg.geometry("560x380")
        dlg.minsize(520, 320)
        dlg.transient(self.root)
        dlg.grab_set()

        # zentrieren
        dlg.update_idletasks()
        x = self.root.winfo_rootx() + (self.root.winfo_width() // 2 - dlg.winfo_width() // 2)
        y = self.root.winfo_rooty() + (self.root.winfo_height() // 2 - dlg.winfo_height() // 2)
        dlg.geometry(f"+{max(0,x)}+{max(0,y)}")

        # Header
        hdr = tk.Frame(dlg, bg=self.AMAZON["bg"])
        hdr.pack(fill="x", padx=16, pady=(14, 6))
        ttk.Label(hdr, text="Enter BE IDs (any format)", style="AmazonTitle.TLabel").pack(anchor="w")
        ttk.Label(
            hdr,
            text="Left: paste any text — digits will be extracted into 10-digit IDs.  Right: pick from saved BE datasets.",
            style="AmazonSubtitle.TLabel"
        ).pack(anchor="w", pady=(2, 0))

        # Body – 2 Spalten
        body = ttk.Frame(dlg, style="Amazon.TFrame")
        body.pack(fill="both", expand=True, padx=16, pady=(8, 0))
        body.columnconfigure(0, weight=1, uniform="cols")
        body.columnconfigure(1, weight=1, uniform="cols")
        body.rowconfigure(0, weight=0)
        body.rowconfigure(1, weight=1)
        body.grid_rowconfigure(1, minsize=120)

        # Links: Manuell
        left = ttk.Frame(body, style="Amazon.TFrame")
        left.grid(row=0, column=0, rowspan=2, sticky="nsew", padx=(0, 8))
        ttk.Label(left, text="Manual input", style="AmazonBody.TLabel").pack(anchor="w", pady=(0, 4))
        text_box = tk.Text(left, wrap="word", height=10, font=("Segoe UI", 10), bg=self.AMAZON["bg"])
        text_box.pack(fill="both", expand=True)
        text_box.focus_set()

        # Rechts: Datasets (nur Typ BE)
        right = ttk.Frame(body, style="Amazon.TFrame")
        right.grid(row=0, column=1, rowspan=2, sticky="nsew", padx=(8, 0))
        ttk.Label(right, text="From BE Datasets", style="AmazonBody.TLabel").pack(anchor="w", pady=(0, 4))

        list_wrap = ttk.Frame(right, style="Amazon.TFrame")
        list_wrap.pack(fill="both", expand=True)
        list_wrap.columnconfigure(0, weight=1)
        list_wrap.columnconfigure(1, weight=0)
        list_wrap.rowconfigure(0, weight=1)

        be_datasets = self.filter_datasets("BE")

        ds_list = tk.Listbox(
            list_wrap,
            selectmode="extended",
            activestyle="none",
            font=("Segoe UI", 10),
            bg=self.AMAZON["bg"],
            fg=self.AMAZON["text"],
            highlightthickness=0,
            relief="flat",
            bd=0
        )
        if be_datasets:
            for ds in be_datasets:
                ds_list.insert("end", ds.get("name", "Unnamed"))
        else:
            ds_list.insert("end", "No BE datasets yet")
            ds_list.configure(state="disabled")
        ds_list.grid(row=0, column=0, sticky="nsew")

        ds_scroll = ttk.Scrollbar(
            list_wrap,
            orient="vertical",
            command=ds_list.yview,
            style="Amazon.Vertical.TScrollbar"
        )
        ds_scroll.grid(row=0, column=1, sticky="ns")
        ds_list.configure(yscrollcommand=ds_scroll.set)

        # Actions
        actions = ttk.Frame(dlg, style="Amazon.TFrame")
        actions.pack(side="bottom", fill="x", padx=16, pady=12)

        result = {"value": None}

        def submit():
            manual_text = (text_box.get("1.0", "end") or "").strip()

            selected_items = []
            if be_datasets and ds_list.curselection():
                for idx in ds_list.curselection():
                    selected_items.extend(self.ds_items(be_datasets[idx], "BE"))

            if not manual_text and not selected_items:
                messagebox.showerror("Error", "Please paste BE IDs or select at least one dataset.")
                return

            combined_parts = []
            if manual_text:
                combined_parts.append(manual_text)
            if selected_items:
                combined_parts.append("\n".join(selected_items))

            result["value"] = "\n".join(combined_parts)
            dlg.destroy()

        def cancel():
            result["value"] = None
            dlg.destroy()

        self.create_amazon_button(actions, "Confirm", submit).pack(side="right")
        self.create_secondary_button(actions, "Cancel", cancel).pack(side="right", padx=(8, 0))

        dlg.bind("<Escape>", lambda e: cancel())
        dlg.wait_window()
        return result["value"]


    def show_be_sonar_mapping_dialog(self):
        """
        Upload BE → Sonar (v2):
        Links (BE) und rechts (SONAR) jeweils 'Manual' oder 'From Dataset'.
        Bei Dataset werden die Items aus den jeweiligen Spalten ('BE' / 'SONAR') zusammengeführt.
        Rückgabe: Liste von Paaren [(be_id, sonar_entry), ...] oder None.
        """
        dlg = tk.Toplevel(self.root)
        dlg.title("Upload BE to Sonar — Mapping")
        dlg.configure(bg=self.AMAZON["bg"])
        dlg.geometry("780x420")
        dlg.minsize(720, 380)
        dlg.transient(self.root)
        dlg.grab_set()

        # Grid-Layout
        dlg.grid_rowconfigure(0, weight=0)
        dlg.grid_rowconfigure(1, weight=1)
        dlg.grid_rowconfigure(2, weight=0)
        dlg.grid_columnconfigure(0, weight=1)

        # Header
        hdr = tk.Frame(dlg, bg=self.AMAZON["bg"])
        hdr.grid(row=0, column=0, sticky="ew", padx=16, pady=(14, 6))
        ttk.Label(hdr, text="Paste BE IDs and matching Sonar Campaigns", style="AmazonTitle.TLabel").pack(anchor="w")
        ttk.Label(
            hdr,
            text="Left: BE IDs (10 digits). Right: Sonar (URL or ID). Choose Manual or From Dataset on each side.",
            style="AmazonSubtitle.TLabel"
        ).pack(anchor="w", pady=(2, 0))

        # Body
        body = ttk.Frame(dlg, style="Amazon.TFrame")
        body.grid(row=1, column=0, sticky="nsew", padx=16, pady=(8, 0))
        body.columnconfigure(0, weight=1, uniform="cols")
        body.columnconfigure(1, weight=1, uniform="cols")
        body.rowconfigure(0, weight=0)
        body.rowconfigure(1, weight=0)
        body.rowconfigure(2, weight=1)
        body.rowconfigure(3, weight=0)
        body.rowconfigure(4, weight=1)
        body.grid_rowconfigure(4, minsize=120)

        # LEFT (BE)
        ttk.Label(body, text="Bullseye IDs", style="AmazonBody.TLabel").grid(row=0, column=0, sticky="w")
        left_mode = tk.StringVar(value="manual")
        left_modes = ttk.Frame(body, style="Amazon.TFrame")
        left_modes.grid(row=1, column=0, sticky="w", pady=(2, 4))
        ttk.Radiobutton(left_modes, text="Manual", variable=left_mode, value="manual", style="Amazon.TRadiobutton").pack(side="left")
        ttk.Radiobutton(left_modes, text="From Dataset", variable=left_mode, value="dataset", style="Amazon.TRadiobutton").pack(side="left", padx=(12, 0))

        left_manual = tk.Text(body, wrap="none", height=8, font=("Segoe UI", 10), bg=self.AMAZON["bg"])
        left_manual.grid(row=2, column=0, sticky="nsew", padx=(0, 8))
        left_manual.focus_set()

        ttk.Label(body, text="Select BE datasets", style="AmazonMuted.TLabel").grid(row=3, column=0, sticky="w", pady=(6, 2))
        be_datasets = self.filter_datasets("BE")
        be_names = [ds.get("name", "Unnamed") for ds in be_datasets] or ["No BE datasets yet"]
        left_checks_frame, left_vars = self._make_scrollable_checks(body, be_names)
        left_checks_frame.grid(row=4, column=0, sticky="nsew", padx=(0, 8))

        # RIGHT (SONAR)
        ttk.Label(body, text="Sonar Campaigns (URL or ID)", style="AmazonBody.TLabel").grid(row=0, column=1, sticky="w")
        right_mode = tk.StringVar(value="manual")
        right_modes = ttk.Frame(body, style="Amazon.TFrame")
        right_modes.grid(row=1, column=1, sticky="w", pady=(2, 4))
        ttk.Radiobutton(right_modes, text="Manual", variable=right_mode, value="manual", style="Amazon.TRadiobutton").pack(side="left")
        ttk.Radiobutton(right_modes, text="From Dataset", variable=right_mode, value="dataset", style="Amazon.TRadiobutton").pack(side="left", padx=(12, 0))

        right_manual = tk.Text(body, wrap="none", height=8, font=("Segoe UI", 10), bg=self.AMAZON["bg"])
        right_manual.grid(row=2, column=1, sticky="nsew", padx=(8, 0))

        ttk.Label(body, text="Select SONAR datasets", style="AmazonMuted.TLabel").grid(row=3, column=1, sticky="w", pady=(6, 2))
        sonar_datasets = self.filter_datasets("SONAR")
        sonar_names = [ds.get("name", "Unnamed") for ds in sonar_datasets] or ["No SONAR datasets yet"]
        right_checks_frame, right_vars = self._make_scrollable_checks(body, sonar_names)
        right_checks_frame.grid(row=4, column=1, sticky="nsew", padx=(8, 0))

        # Sichtbarkeit je Modus
        def _toggle_left(*_):
            if left_mode.get() == "manual":
                left_checks_frame.grid_remove()
                left_manual.grid()
                left_manual.focus_set()
            else:
                left_manual.grid_remove()
                left_checks_frame.grid()

        def _toggle_right(*_):
            if right_mode.get() == "manual":
                right_checks_frame.grid_remove()
                right_manual.grid()
                right_manual.focus_set()
            else:
                right_manual.grid_remove()
                right_checks_frame.grid()

        left_mode.trace_add("write", _toggle_left)
        right_mode.trace_add("write", _toggle_right)
        _toggle_left()
        _toggle_right()

        # Footer
        actions = ttk.Frame(dlg, style="Amazon.TFrame")
        actions.grid(row=2, column=0, sticky="ew", padx=16, pady=12)

        result = {"pairs": None}

        def _lines(text_widget):
            return [l.strip() for l in text_widget.get("1.0", "end").splitlines() if l.strip()]

        def _collect_be():
            if left_mode.get() == "manual":
                raw = _lines(left_manual)
                cleaned = []
                for i, be in enumerate(raw, start=1):
                    digits = "".join(ch for ch in be if ch.isdigit())
                    if len(digits) != 10:
                        messagebox.showerror("Error", f"BE line {i}: must be 10 digits (got: '{be}').")
                        return None
                    cleaned.append(digits)
                return cleaned
            else:
                if not be_datasets:
                    messagebox.showerror("Error", "No BE datasets available.")
                    return None
                items = []
                for i, v in enumerate(left_vars):
                    if v.get():
                        items.extend(self.ds_items(be_datasets[i], "BE"))
                if not items:
                    messagebox.showerror("Error", "Select at least one BE dataset or switch to Manual.")
                    return None
                cleaned = []
                for i, be in enumerate(items, start=1):
                    digits = "".join(ch for ch in be if ch.isdigit())
                    if len(digits) != 10:
                        messagebox.showerror("Error", f"BE dataset item {i}: must be 10 digits (got: '{be}').")
                        return None
                    cleaned.append(digits)
                return cleaned

        def _collect_sonar():
            if right_mode.get() == "manual":
                items = _lines(right_manual)
                if not items:
                    messagebox.showerror("Error", "Please enter at least one Sonar URL or ID, or select a dataset.")
                    return None
                return items
            else:
                if not sonar_datasets:
                    messagebox.showerror("Error", "No SONAR datasets available.")
                    return None
                items = []
                for i, v in enumerate(right_vars):
                    if v.get():
                        items.extend(self.ds_items(sonar_datasets[i], "SONAR"))
                if not items:
                    messagebox.showerror("Error", "Select at least one SONAR dataset or switch to Manual.")
                    return None
                return items

        def parse_pairs():
            left_items = _collect_be()
            if left_items is None:
                return
            right_items = _collect_sonar()
            if right_items is None:
                return
            if len(left_items) != len(right_items):
                messagebox.showerror(
                    "Error",
                    f"Line counts do not match:\nBE: {len(left_items)}  vs  Sonar: {len(right_items)}.\nEach BE must have a matching Sonar entry."
                )
                return
            result["pairs"] = list(zip(left_items, right_items))
            dlg.destroy()

        def cancel():
            result["pairs"] = None
            dlg.destroy()

        self.create_amazon_button(actions, "Confirm", parse_pairs).pack(side="right")
        self.create_secondary_button(actions, "Cancel", cancel).pack(side="right", padx=(8, 0))

        dlg.bind("<Escape>", lambda e: cancel())
        dlg.bind("<Return>", lambda e: parse_pairs())

        dlg.wait_window()
        return result["pairs"]

    
    
        

    def show_confirm_ids_dialog(self, segment_ids):
        """Amazon-style modal confirmation with the list of IDs. Returns True if user confirms."""
        dlg = tk.Toplevel(self.root)
        dlg.title("Confirm IDs")
        dlg.configure(bg=self.AMAZON["bg"])
        dlg.geometry("440x420")
        dlg.minsize(420, 360)
        dlg.transient(self.root)
        dlg.grab_set()

        # center on parent
        dlg.update_idletasks()
        x = self.root.winfo_rootx() + (self.root.winfo_width() // 2 - dlg.winfo_width() // 2)
        y = self.root.winfo_rooty() + (self.root.winfo_height() // 2 - dlg.winfo_height() // 2)
        dlg.geometry(f"+{max(0,x)}+{max(0,y)}")

        # header
        hdr = tk.Frame(dlg, bg=self.AMAZON["bg"])
        hdr.pack(fill="x", padx=16, pady=(14, 8))
        ttk.Label(hdr, text=f"Found {len(segment_ids)} valid Segment IDs:",
                  style="AmazonTitle.TLabel").pack(anchor="w")

        # list area (read-only Text with scrollbar)
        body = tk.Frame(dlg, bg=self.AMAZON["bg"])
        body.pack(fill="both", expand=True, padx=16)

        scroll = tk.Scrollbar(body)
        scroll.pack(side="right", fill="y")

        txt = tk.Text(
            body,
            wrap="none",
            height=10,
            font=("Segoe UI", 10),
            bg="white",
            relief="solid",
            bd=1
        )
        txt.pack(fill="both", expand=True, side="left")
        txt.configure(yscrollcommand=scroll.set)
        scroll.configure(command=txt.yview)

        txt.insert("1.0", "\n".join(segment_ids))
        txt.configure(state="disabled")

        # Hinweis separat anzeigen
        lbl = ttk.Label(dlg, text="Proceed?", style="AmazonSubtitle.TLabel")
        lbl.pack(anchor="e", padx=16, pady=(6, 0))


        # actions
        actions = tk.Frame(dlg, bg=self.AMAZON["bg"])
        actions.pack(fill="x", padx=16, pady=12)
        user_choice = {"ok": False}

        def on_yes():
            user_choice["ok"] = True
            dlg.destroy()

        def on_no():
            user_choice["ok"] = False
            dlg.destroy()

        self.create_secondary_button(actions, "No", on_no).pack(side="right", padx=(0, 8))
        self.create_amazon_button(actions, "Yes", on_yes).pack(side="right")

        dlg.bind("<Escape>", lambda e: on_no())
        dlg.bind("<Return>", lambda e: on_yes())

        dlg.wait_window()
        return user_choice["ok"]



    def show_send_preview_dialog(self):
        """
        Dialog: Kampagnen für Preview auswählen.
        Entweder manuell (eine pro Zeile: ID oder Sonar-URL) ODER aus Datasets (nur SONAR-Spalte).
        Rückgabe: {"campaigns": [str, ...]} oder None bei Abbruch/Fehler.
        """
        dlg = tk.Toplevel(self.root)
        dlg.title("Send Preview — Select Campaigns")
        dlg.configure(bg=self.AMAZON["bg"])
        dlg.geometry("780x420")
        dlg.minsize(720, 380)
        dlg.transient(self.root)
        dlg.grab_set()

        # zentrieren
        dlg.update_idletasks()
        x = self.root.winfo_rootx() + (self.root.winfo_width() // 2 - dlg.winfo_width() // 2)
        y = self.root.winfo_rooty() + (self.root.winfo_height() // 2 - dlg.winfo_height() // 2)
        dlg.geometry(f"+{max(0, x)}+{max(0, y)}")

        # Header
        hdr = tk.Frame(dlg, bg=self.AMAZON["bg"])
        hdr.pack(fill="x", padx=16, pady=(14, 6))
        ttk.Label(hdr, text="Choose Sonar Campaigns for Preview", style="AmazonTitle.TLabel").pack(anchor="w")
        ttk.Label(
            hdr,
            text="Option A: Paste campaigns (ID or https://sonar-eu.amazon.com/… per line).  Option B: Select from SONAR datasets.",
            style="AmazonSubtitle.TLabel"
        ).pack(anchor="w", pady=(2, 0))

        # Body 2-Spalten
        body = ttk.Frame(dlg, style="Amazon.TFrame")
        body.pack(fill="both", expand=True, padx=16, pady=(8, 0))
        body.columnconfigure(0, weight=1, uniform="cols")
        body.columnconfigure(1, weight=1, uniform="cols")
        body.rowconfigure(1, weight=1)

        # Links: Manuelle Eingabe
        left = ttk.Frame(body, style="Amazon.TFrame")
        left.grid(row=0, column=0, rowspan=2, sticky="nsew", padx=(0, 8))
        ttk.Label(left, text="Manual (one per line)", style="AmazonBody.TLabel").pack(anchor="w", pady=(0, 4))
        manual_text = tk.Text(left, wrap="none", height=14, font=("Segoe UI", 10), bg=self.AMAZON["bg"])
        manual_text.pack(fill="both", expand=True)
        manual_text.focus_set()

        # Rechts: SONAR Datasets (Checkbox-Liste)
        right = ttk.Frame(body, style="Amazon.TFrame")
        right.grid(row=0, column=1, rowspan=2, sticky="nsew", padx=(8, 0))
        ttk.Label(right, text="From SONAR datasets", style="AmazonBody.TLabel").pack(anchor="w", pady=(0, 4))

        sonar_datasets = self.filter_datasets("SONAR")
        sonar_names = [ds.get("name", "Unnamed") for ds in sonar_datasets] or ["No SONAR datasets yet"]
        ds_frame, ds_vars = self._make_scrollable_checks(right, sonar_names)
        ds_frame.pack(fill="both", expand=True)

        # Footer / Actions
        actions = ttk.Frame(dlg, style="Amazon.TFrame")
        actions.pack(fill="x", padx=16, pady=12)

        result = {"campaigns": None}

        def _collect():
            # 1) Manuelle Zeilen
            lines = [l.strip() for l in manual_text.get("1.0", "end").splitlines() if l.strip()]

            # 2) Ausgewählte SONAR-Datasets
            if sonar_datasets and any(v.get() for v in ds_vars):
                for i, v in enumerate(ds_vars):
                    if v.get():
                        lines.extend(self.ds_items(sonar_datasets[i], "SONAR"))

            # Validierung
            if not lines:
                messagebox.showerror("Send Preview", "Please provide at least one campaign (manual or from dataset).")
                return

            # Akzeptiert: reine Ziffern ODER echte Sonar-URL (wie im Dataset-Editor)
            cleaned = []
            for idx, val in enumerate(lines, start=1):
                v = str(val).strip()
                if v.isdigit() or v.startswith("https://sonar-eu.amazon.com"):
                    cleaned.append(v)
                else:
                    messagebox.showerror(
                        "Send Preview",
                        f"Line {idx}: must be a campaign ID (digits) or a Sonar URL starting with https://sonar-eu.amazon.com"
                    )
                    return

            # Duplikate entfernen (Reihenfolge beibehalten)
            cleaned = self._unique_preserve_order(cleaned)
            result["campaigns"] = cleaned
            dlg.destroy()

        def _cancel():
            result["campaigns"] = None
            dlg.destroy()

        self.create_secondary_button(actions, "Cancel", _cancel).pack(side="right", padx=(0, 8))
        self.create_amazon_button(actions, "Confirm", _collect).pack(side="right")

        dlg.bind("<Escape>", lambda e: _cancel())
        dlg.bind("<Return>", lambda e: _collect())

        dlg.wait_window()
        return result["campaigns"]

    def _show_preview_plan_dialog(self, batches: dict):
        """Zeigt vorbereitete Batches an, erlaubt Mehrfachauswahl der MPs
        und (falls vorhanden) die Zuordnung von 'unknown'-Jobs zu einem MP.
        Rückgabe:
          - (selected_mps: list[int], unknown_target_mp: int|None), wenn 'Senden' gedrückt
          - None, wenn nur geschlossen (nur vorbereitet)
        """
        import tkinter as tk
        from tkinter import ttk, messagebox

        country = {3: "UK", 4: "DE", 5: "FR", 35691: "IT", 44551: "ES"}

        dlg = tk.Toplevel(self.root)
        dlg.title("Preview – Batches vorbereitet")
        dlg.configure(bg=self.AMAZON["bg"])
        dlg.geometry("460x420")
        dlg.transient(self.root)
        dlg.grab_set()

        ttk.Label(dlg, text="Batches vorbereitet", style="AmazonTitle.TLabel")\
            .pack(anchor="w", padx=16, pady=(14, 6))
        ttk.Label(
            dlg,
            text="Wähle einen oder mehrere Marketplaces zum Senden. Unknown-IDs können einem MP zugeordnet werden.",
            style="AmazonSubtitle.TLabel"
        ).pack(anchor="w", padx=16, pady=(0, 8))

        body = ttk.Frame(dlg, style="Amazon.TFrame")
        body.pack(fill="both", expand=True, padx=16, pady=0)

        # Mehrfachauswahl (Checkboxen)
        checks_wrap = ttk.Frame(body, style="Amazon.TFrame")
        checks_wrap.pack(fill="x", pady=(0, 8))
        var = tk.BooleanVar(value=False)  # Start: nichts ausgewählt
        vars_by_mp: dict[int, tk.BooleanVar] = {}
        available_mps = [k for k in batches.keys() if isinstance(k, int)]
        for mp in sorted(available_mps):
            jobs = batches.get(mp, []) or []
            already = mp in getattr(self, "_preview_mps_sent", set())
            suffix = "  ✓ gesendet" if already else ""
            label = f"{country.get(mp, mp)} (MP {mp}) – {len(jobs)} Kampagne(n){suffix}"

            var = tk.BooleanVar(value=False)  # NICHT vorselektiert
            ttk.Checkbutton(checks_wrap, text=label, variable=var, style="Amazon.TCheckbutton")\
                .pack(anchor="w", pady=2)
            vars_by_mp[mp] = var


        unknown_jobs = batches.get("unknown") or []
        assign_frame = ttk.Frame(body, style="Amazon.TFrame")
        assign_frame.pack(fill="x")

        unknown_target_var = tk.StringVar(value="")
        if unknown_jobs:
            ttk.Label(
                assign_frame,
                text=f"IDs ohne MP: {len(unknown_jobs)} – bitte einem MP zuordnen:",
                style="AmazonBody.TLabel"
            ).pack(anchor="w", pady=(4, 2))

            mp_labels = [f"{country.get(mp, mp)} (MP {mp})" for mp in sorted(vars_by_mp.keys())]
            cb = ttk.Combobox(assign_frame, state="readonly", values=mp_labels, textvariable=unknown_target_var)
            cb.pack(fill="x")
            # Default: erster MP mit Haken (falls vorhanden), sonst erster Eintrag
            preselect = ""
            for mp in sorted(vars_by_mp.keys()):
                if vars_by_mp[mp].get():
                    preselect = f"{country.get(mp, mp)} (MP {mp})"
                    break
            unknown_target_var.set(preselect or (mp_labels[0] if mp_labels else ""))

        # Aktionen
        actions = ttk.Frame(dlg, style="Amazon.TFrame")
        actions.pack(fill="x", padx=16, pady=12)
        result = {"value": None}

        def parse_unknown_target() -> int | None:
            if not unknown_jobs:
                return None
            txt = (unknown_target_var.get() or "").strip()
            if not txt:
                return None
            # ... "(MP 4)" am Ende herausparsen
            m = re.search(r"\(MP\s+(\d+)\)\s*$", txt)
            return int(m.group(1)) if m else None

        def send_now():
            selected_mps = [mp for mp, v in vars_by_mp.items() if v.get()]
            if not selected_mps:
                messagebox.showerror("Send Preview", "Bitte mindestens einen Marketplace auswählen.")
                return
            ut = parse_unknown_target()
            if unknown_jobs and (ut is None):
                messagebox.showerror("Send Preview", "Bitte einen Marketplace für die 'unknown'-IDs auswählen.")
                return
            result["value"] = (selected_mps, ut)
            dlg.destroy()

        def close_only():
            result["value"] = None
            dlg.destroy()

        self.create_secondary_button(actions, "Schließen (nur vorbereiten)", close_only)\
            .pack(side="right", padx=(0, 8))
        self.create_amazon_button(actions, "Ausgewählte senden", send_now).pack(side="right")

        import re  # lokal für parse_unknown_target
        dlg.bind("<Escape>", lambda e: close_only())
        dlg.wait_window()
        return result["value"]



    def show_approve_sonar_dialog(self):
        """
        Dialog: Kampagnen zum Approven auswählen.
        Modus A: Manuell (eine pro Zeile: ID ODER Sonar-URL)
        Modus B: Aus gespeicherten Datasets (nur SONAR-Spalten).
        Rückgabe: {"campaigns": [str, ...]} oder None.
        """
        dlg = tk.Toplevel(self.root)
        dlg.title("Approve Sonar — Select Campaigns")
        dlg.configure(bg=self.AMAZON["bg"])
        dlg.geometry("780x420")
        dlg.minsize(720, 380)
        dlg.transient(self.root)
        dlg.grab_set()

        # zentrieren
        dlg.update_idletasks()
        x = self.root.winfo_rootx() + (self.root.winfo_width() // 2 - dlg.winfo_width() // 2)
        y = self.root.winfo_rooty() + (self.root.winfo_height() // 2 - dlg.winfo_height() // 2)
        dlg.geometry(f"+{max(0, x)}+{max(0, y)}")

        # Header
        hdr = tk.Frame(dlg, bg=self.AMAZON["bg"])
        hdr.pack(fill="x", padx=16, pady=(14, 6))
        ttk.Label(hdr, text="Choose Sonar Campaigns to Approve", style="AmazonTitle.TLabel").pack(anchor="w")
        ttk.Label(
            hdr,
            text="Option A: Paste campaigns (ID or https://sonar-eu.amazon.com/… per line).  Option B: Select from SONAR datasets.",
            style="AmazonSubtitle.TLabel"
        ).pack(anchor="w", pady=(2, 0))

        # Body 2-Spalten
        body = ttk.Frame(dlg, style="Amazon.TFrame")
        body.pack(fill="both", expand=True, padx=16, pady=(8, 0))
        body.columnconfigure(0, weight=1, uniform="cols")
        body.columnconfigure(1, weight=1, uniform="cols")
        body.rowconfigure(1, weight=1)

        # Links: Manuell
        left = ttk.Frame(body, style="Amazon.TFrame")
        left.grid(row=0, column=0, rowspan=2, sticky="nsew", padx=(0, 8))
        ttk.Label(left, text="Manual (one per line)", style="AmazonBody.TLabel").pack(anchor="w", pady=(0, 4))
        manual_text = tk.Text(left, wrap="none", height=14, font=("Segoe UI", 10), bg=self.AMAZON["bg"])
        manual_text.pack(fill="both", expand=True)
        manual_text.focus_set()

        # Rechts: SONAR Datasets (Checkboxen)
        right = ttk.Frame(body, style="Amazon.TFrame")
        right.grid(row=0, column=1, rowspan=2, sticky="nsew", padx=(8, 0))
        ttk.Label(right, text="From SONAR datasets", style="AmazonBody.TLabel").pack(anchor="w", pady=(0, 4))

        sonar_datasets = self.filter_datasets("SONAR")
        sonar_names = [ds.get("name", "Unnamed") for ds in sonar_datasets] or ["No SONAR datasets yet"]
        ds_frame, ds_vars = self._make_scrollable_checks(right, sonar_names)
        ds_frame.pack(fill="both", expand=True)

        # Footer / Actions
        actions = ttk.Frame(dlg, style="Amazon.TFrame")
        actions.pack(fill="x", padx=16, pady=12)

        result = {"campaigns": None}

        def _collect():
            lines = [l.strip() for l in manual_text.get("1.0", "end").splitlines() if l.strip()]
            if sonar_datasets and any(v.get() for v in ds_vars):
                for i, v in enumerate(ds_vars):
                    if v.get():
                        lines.extend(self.ds_items(sonar_datasets[i], "SONAR"))

            if not lines:
                messagebox.showerror("Approve Sonar", "Please provide at least one campaign (manual or from dataset).")
                return

            cleaned = []
            for idx, val in enumerate(lines, start=1):
                v = str(val).strip()
                if v.isdigit() or v.startswith("https://sonar-eu.amazon.com"):
                    cleaned.append(v)
                else:
                    messagebox.showerror(
                        "Approve Sonar",
                        f"Line {idx}: must be a campaign ID (digits) or a Sonar URL starting with https://sonar-eu.amazon.com"
                    )
                    return

            cleaned = self._unique_preserve_order(cleaned)
            result["campaigns"] = cleaned
            dlg.destroy()

        def _cancel():
            result["campaigns"] = None
            dlg.destroy()

        self.create_secondary_button(actions, "Cancel", _cancel).pack(side="right", padx=(0, 8))
        self.create_amazon_button(actions, "Confirm", _collect).pack(side="right")

        dlg.bind("<Escape>", lambda e: _cancel())
        dlg.bind("<Return>", lambda e: _collect())

        dlg.wait_window()
        return result["campaigns"]


    def _choose_marketplace_dialog(self, title="Select marketplace"):
        import tkinter as tk
        from tkinter import ttk, messagebox

        # bekannte MPs
        known_mps = [3, 4, 5, 35691, 44551]
        country = {3: "UK", 4: "DE", 5: "FR", 35691: "IT", 44551: "ES"}

        dlg = tk.Toplevel(self.root)
        dlg.title(title)
        dlg.configure(bg=self.AMAZON["bg"])
        dlg.geometry("360x180")
        dlg.minsize(340, 160)
        dlg.transient(self.root)
        dlg.grab_set()

        ttk.Label(dlg, text="Choose marketplace for previews", style="AmazonTitle.TLabel")\
            .pack(anchor="w", padx=16, pady=(14, 6))

        wrap = ttk.Frame(dlg, style="Amazon.TFrame")
        wrap.pack(fill="x", padx=16, pady=(0, 8))
        ttk.Label(wrap, text="Marketplace", style="AmazonBody.TLabel").pack(anchor="w")

        values = [f"{country.get(mp, mp)} (MP {mp})" for mp in known_mps]
        var = tk.StringVar(value=values[1])  # Default DE(4)
        cb = ttk.Combobox(wrap, state="readonly", values=values, textvariable=var)
        cb.pack(fill="x", pady=(2, 0))
        cb.focus_set()

        result = {"mp": None}

        def ok():
            sel = var.get()
            for mp in known_mps:
                if sel.endswith(f"(MP {mp})"):
                    result["mp"] = mp
                    break
            dlg.destroy()

        def cancel():
            result["mp"] = None
            dlg.destroy()

        actions = ttk.Frame(dlg, style="Amazon.TFrame")
        actions.pack(fill="x", padx=16, pady=12)
        self.create_secondary_button(actions, "Cancel", cancel).pack(side="right", padx=(0, 8))
        self.create_amazon_button(actions, "OK", ok).pack(side="right")
        dlg.bind("<Escape>", lambda e: cancel())

        dlg.wait_window()
        return result["mp"]

        
            


    
    
    def send_preview(self):
        """
        Kampagnen nur einmal einsammeln, Batches planen und senden.
        Nach jedem Senden öffnet sich der MP-Dialog erneut, bis nichts mehr übrig ist
        oder du ihn schließt.
        """
        try:
            if preview_run_batch is None:
                
                messagebox.showerror("Send Preview", "preview_campaigns.py nicht gefunden.\nBitte neben bullseye_app.py ablegen.")
                return

            customer_id = (self.profile or {}).get("customerId")
            if customer_id in (None, ""):
                
                messagebox.showerror("Send Preview", "Bitte zuerst eine numerische Customer Id im Profil hinterlegen.")
                return

            # Erstaufruf: Kampagnen einsammeln und Batches vorbereiten
            if not self._prepared_preview_batches:
                campaigns = self.show_send_preview_dialog()
                if not campaigns:
                    self.status_var.set("No campaigns selected.")
                    return

                jobs = [{"campaign": c} for c in campaigns]

                def detect_mp(job):
                    if isinstance(job.get("marketplaceId"), int):
                        return job["marketplaceId"]
                    return self._extract_mp_from_line(job.get("campaign", ""))

                by_mp, unknown_jobs = {}, []
                for j in jobs:
                    mp = detect_mp(j)
                    if mp is None:
                        unknown_jobs.append(j)
                    else:
                        by_mp.setdefault(int(mp), []).append(j)

                self._prepared_preview_batches = {**by_mp, "unknown": unknown_jobs}
                self._preview_mps_sent = set()

            def remaining(batches):
                return sum(len(v) for k, v in batches.items() if isinstance(k, int)) + len(batches.get("unknown") or [])

            def ask_and_send():
                batches = self._prepared_preview_batches or {}
                if remaining(batches) == 0:
                    self.status_var.set("All prepared previews have been sent.")
                    return

                selection = self._show_preview_plan_dialog(batches)
                if selection is None:
                    self.status_var.set(f"Prepared {remaining(batches)} preview(s).")
                    return

                selected_mps, unknown_target_mp = selection

                def jobs_for_mp(mp: int) -> list:
                    base = list(batches.get(mp, []))
                    if unknown_target_mp == mp and batches.get("unknown"):
                        base += list(batches["unknown"])
                    return base

                per_mp = {mp: jobs_for_mp(mp) for mp in selected_mps}
                total_jobs = sum(len(v) for v in per_mp.values())
                if total_jobs == 0:
                    
                    messagebox.showerror("Send Preview", "Keine Jobs für die ausgewählten Marketplaces.")
                    return

                self._set_busy(True)
                self.progress_var.set(0)
                self.progress_label.config(text="Starting previews…")
                self.status_var.set(f"Sending {total_jobs} preview(s) across {len(selected_mps)} MP(s)…")

                import threading, time as _time
                def worker():
                    out_paths = []
                    processed = 0
                    start_ts = _time.time()
                    try:
                        for mp in selected_mps:
                            batch_jobs = per_mp.get(mp, [])
                            if not batch_jobs:
                                continue

                            def progress_cb(i, _processed=processed):
                                self.root.after(0, lambda: self.update_progress(_processed + i, total_jobs))

                            def status_cb(msg):
                                self.root.after(0, lambda m=msg: (self.status_var.set(m), self.root.update_idletasks()))

                            results, out_path = preview_run_batch(
                                batch_jobs,
                                marketplace_id=mp,
                                status_callback=status_cb,
                                progress_callback=progress_cb,
                                headless=self.headless_var.get(),
                            )
                            if out_path:
                                out_paths.append(out_path)
                            processed += len(batch_jobs)

                            # Verbrauchte Jobs entfernen + MP markieren
                            self._preview_mps_sent.add(mp)
                            if mp in self._prepared_preview_batches:
                                self._prepared_preview_batches.pop(mp, None)
                            if unknown_target_mp == mp and self._prepared_preview_batches.get("unknown"):
                                self._prepared_preview_batches["unknown"].clear()

                        def done():
                            self._set_busy(False)
                            for p in out_paths:
                                self._add_result_file(p)
                            elapsed = _time.time() - start_ts
                            self.progress_label.config(
                                text=f"Sent {total_jobs} preview(s) across {len(selected_mps)} MP(s) in {elapsed:.1f}s"
                            )
                            # solange noch was übrig ist -> erneut öffnen
                            if remaining(self._prepared_preview_batches or {}) > 0:
                                ask_and_send()
                            else:
                                self.status_var.set("All prepared previews have been sent.")
                        self.root.after(0, done)

                    except Exception as ex:
                        def fail():
                            self._set_busy(False)
                            from tkinter import messagebox
                            messagebox.showerror("Send Preview", str(ex))
                            self.status_var.set(f"Error: {ex}")
                        self.root.after(0, fail)

                threading.Thread(target=worker, daemon=True).start()

            ask_and_send()

        except Exception as outer_ex:
            self._set_busy(False)
            from tkinter import messagebox
            self.status_var.set(f"Error: {outer_ex}")
            messagebox.showerror("Send Preview", str(outer_ex))



    def approve_sonar(self):
        """
        Approves Sonar campaigns.
        Benötigt:
          - requester alias (aus Profile)
          - Kampagnen (ID oder Sonar-URL), manuell oder aus SONAR-Datasets
        Läuft im Hintergrund-Thread, UI bleibt responsiv.
        """
        try:
            alias = ((self.profile or {}).get("alias") or "").strip()
            if not alias:
                
                messagebox.showerror("Approve Sonar", "Bitte zuerst im Profile deinen Alias hinterlegen.")
                return

            campaigns = self.show_approve_sonar_dialog()
            if not campaigns:
                self.status_var.set("No campaigns selected.")
                return

            # Progress UI reset
            self.progress_var.set(0)
            self.progress_label.config(text="Starting approval…")
            start_time = time.time()
            total = len(campaigns)

            def ui_progress(i):
                # i = 0-based Index
                self.root.after(0, lambda: self.update_progress(i, total))

            def ui_status(message):
                self.root.after(0, lambda: (self.status_var.set(message), self.root.update_idletasks()))

            def on_done(out_path):
                # Ergebnisdatei aufnehmen (falls vorhanden, sonst heuristisch raten)
                if out_path:
                    self._add_result_file(out_path)
                else:
                    guessed = self._guess_new_xlsx(start_time)
                    if guessed:
                        self._add_result_file(guessed)

                elapsed = time.time() - start_time
                self.progress_label.config(text=f"Approved {total} campaign(s) in {elapsed:.1f}s")
                messagebox.showinfo("Approve Sonar", "Finished. Results saved.")
                self._set_busy(False)

            def on_error(err):
                self.status_var.set(f"Error: {err}")
                messagebox.showerror("Approve Sonar", str(err))
                self._set_busy(False)

            import threading
            def worker():
                try:
                    # Backend-Signatur siehe Abschnitt 5 unten
                    from approve_sonar import run_approve_sonar
                except Exception as e:
                    self.root.after(0, lambda: on_error(f"Import failed: {e}"))
                    return

                try:
                    # Erwartete Rückgabe: (results_list, out_path_str) ODER nur out_path_str
                    ret = run_approve_sonar(
                        campaigns=campaigns,
                        requester_alias=alias,
                        status_callback=ui_status,
                        progress_callback=lambda i: self.root.after(0, lambda: self.update_progress(i, total)),
                        headless=self.headless_var.get(),
                        parallel=True,              # Backend darf parallelisieren
                    )
                    # Pfad robust herausziehen
                    out_path = None
                    if isinstance(ret, (list, tuple)):
                        for itm in ret:
                            if isinstance(itm, (str, bytes, os.PathLike)):
                                out_path = itm
                                break
                    elif isinstance(ret, (str, bytes, os.PathLike)):
                        out_path = ret
                    self.root.after(0, lambda: on_done(out_path))
                except Exception as e:
                    self.root.after(0, lambda err=e: on_error(err))

            self._set_busy(True)
            threading.Thread(target=worker, daemon=True).start()

        except Exception as e:
            self._set_busy(False)
            self.status_var.set(f"Error: {str(e)}")
            
            messagebox.showerror("Approve Sonar", str(e))



    # =========================
    # Create RC Sonar (Program + Recurring Version)
    # =========================

    def create_rc_sonar(self):
        """
        GUI-Wrapper für 'Create RC Sonar':
        - wählt Template, Template-Path, Excel/CSV
        - liest die Datei in ein DataFrame
        - ruft das Backend (create_remote_configs) auf
        - zeigt Ergebnisse an und hängt die Datei im UI an
        """
        import os
        from tkinter import messagebox
        import pandas as pd

        # 0) Templates vorhanden?
        if not self.templates:
            if messagebox.askyesno("Templates", "No templates yet. Create one now?"):
                self.open_templates_manager()
            return

        # 1) Dialog (Template + Template-Path + Excel/CSV)
        dlg = self._show_create_rc_dialog()
        if not dlg:
            return

        template = dlg["template"]           # dict mit channel, lobExpression, managementType, businessGroupId,
                                             # teamBindle, startTimeMinutesOffset, endTimeMinutesOffset, ...
        template_path = dlg["template_path"] # z. B. /LAYOUT-TEMPLATES/<uuid>
        file_path = dlg["xlsx_path"]

        # 2) Alias prüfen (wird im Backend in requestContext.userName genutzt)
        alias = ((self.profile or {}).get("alias") or "").strip()
        if not alias:
            messagebox.showerror("Create RC Sonar", "Bitte zuerst im Profile deinen Alias hinterlegen.")
            return

        # 3) Datei lesen → DataFrame (Excel oder CSV)
        try:
            ext = os.path.splitext(file_path)[1].lower()
            if ext in (".xlsx", ".xls"):
                df = pd.read_excel(file_path)
            elif ext == ".csv":
                df = pd.read_csv(file_path)
            else:
                raise ValueError(f"Unsupported file type: {ext}")
            df.columns = [str(c).strip() for c in df.columns]
        except Exception as e:
            messagebox.showerror("Create RC Sonar", f"Could not read file:\n{file_path}\n\n{e}")
            return

        # 4) Schnelle Pflichtfeld-Validierung
        def _has(candidates):
            lc = [c.lower() for c in df.columns]
            return any(c.lower() in lc for c in candidates)

        missing = []
        if not _has(["Name", "Program Name", "ProgramName"]):
            missing.append("Name")
        if not _has(["Marketplace", "marketplaceId"]):
            missing.append("Marketplace")
        if not _has(["BE ID", "BEID", "be id", "beid"]):
            missing.append("BE ID")
        if not _has(["Schedule Start Date", "ScheduleStartDate"]):
            missing.append("Schedule Start Date")
        if not _has(["Schedule End Date", "ScheduleEndDate"]):
            missing.append("Schedule End Date")

        if missing:
            messagebox.showerror("Create RC Sonar", "Missing required columns:\n - " + "\n - ".join(missing))
            return

        # 5) Status-/Progress-Callbacks fürs UI
        def status_cb(msg: str):
            try:
                self.status_var.set(str(msg))
                self.root.update_idletasks()
            except Exception:
                pass

        total_rows = len(df)

        def progress_cb(done: int):
            try:
                self.status_var.set(f"[{done}/{total_rows}] processed …")
                self.root.update_idletasks()
            except Exception:
                pass

        # 6) Backend-Funktion aufrufen
        self._set_busy(True)
        try:
            from create_rc_sonar import create_remote_configs  # Backend

            headless = True  # oder aus Settings übernehmen
            result_df, saved_path = create_remote_configs(
                df=df,
                template=template,
                template_path=template_path,
                alias=alias,
                out_dir=self.datasets_dir,
                status_callback=status_cb,
                progress_callback=progress_cb,
                headless=headless
            )
        except Exception as e:
            self._set_busy(False)
            messagebox.showerror("Create RC Sonar", f"Backend error:\n{e}")
            return
        finally:
            self._set_busy(False)

        # 7) Ergebnisdatei im UI registrieren + kurze Zusammenfassung
        if saved_path and os.path.exists(saved_path):
            try:
                self._add_result_file(saved_path)
            except Exception:
                pass

        try:
            ok_program = (result_df["Program Success"] == True).sum() if "Program Success" in result_df.columns else 0
            ok_campaign = (result_df["Campaign Success"] == True).sum() if "Campaign Success" in result_df.columns else 0
            ok_pairs = min(ok_program, ok_campaign)
            fail_cnt = len(result_df) - ok_pairs
            msg = f"Create RC Sonar: {ok_pairs} success, {fail_cnt} failed."
        except Exception:
            msg = "Create RC Sonar finished."

        self.status_var.set(msg)


    def _show_create_rc_dialog(self):
        """
        UI-Dialog für Create RC Sonar:
          - Template-Auswahl (aus gespeicherten Templates)
          - Template Path (Pflicht, z.B. /LAYOUT-TEMPLATES/<uuid>)
          - Excel/CSV wählen
        Rückgabe: {"template": dict, "template_path": str, "xlsx_path": str} oder None
        """
        import tkinter as tk
        from tkinter import ttk, filedialog, messagebox

        dlg = tk.Toplevel(self.root)
        dlg.title("Create RC Sonar (Program + Version)")
        dlg.configure(bg=self.AMAZON["bg"])
        dlg.geometry("620x420")
        dlg.minsize(600, 380)
        dlg.transient(self.root)
        dlg.grab_set()

        # center
        dlg.update_idletasks()
        x = self.root.winfo_rootx() + (self.root.winfo_width() // 2 - dlg.winfo_width() // 2)
        y = self.root.winfo_rooty() + (self.root.winfo_height() // 2 - dlg.winfo_height() // 2)
        dlg.geometry(f"+{max(0, x)}+{max(0, y)}")

        body = ttk.Frame(dlg, style="Amazon.TFrame")
        body.pack(fill="both", expand=True, padx=16, pady=12)
        for c in range(2):
            body.columnconfigure(c, weight=1)

        # Template Auswahl
        ttk.Label(body, text="Template", style="AmazonBody.TLabel").grid(row=0, column=0, sticky="w")
        tpl_names = [t.get("name", f"Template {i+1}") for i, t in enumerate(self.templates)]
        tpl_var = tk.StringVar(value=(tpl_names[0] if tpl_names else ""))
        cb = ttk.Combobox(body, state="readonly", values=tpl_names, textvariable=tpl_var)
        cb.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(0, 10))

        # Template Path
        ttk.Label(body, text="Template Path (z. B. /LAYOUT-TEMPLATES/<uuid>)", style="AmazonBody.TLabel").grid(row=2, column=0, sticky="w")
        tpl_path_var = tk.StringVar(value="")
        tpl_path_entry = ttk.Entry(body, textvariable=tpl_path_var)
        tpl_path_entry.grid(row=3, column=0, columnspan=2, sticky="ew", pady=(0, 10))
        tpl_path_entry.focus_set()

        # Excel/CSV
        ttk.Label(body, text="Excel/CSV mit Eingaben", style="AmazonBody.TLabel").grid(row=4, column=0, sticky="w")
        path_var = tk.StringVar(value="")
        rowp = ttk.Frame(body, style="Amazon.TFrame")
        rowp.grid(row=5, column=0, columnspan=2, sticky="ew", pady=(0, 8))
        rowp.columnconfigure(0, weight=1)
        e = ttk.Entry(rowp, textvariable=path_var)
        e.grid(row=0, column=0, sticky="ew")
        def pick():
            p = filedialog.askopenfilename(
                title="Choose Excel/CSV",
                filetypes=[("Excel/CSV", "*.xlsx *.xls *.csv"), ("All files", "*.*")]
            )
            if p:
                path_var.set(p)
        self.create_secondary_button(rowp, "Browse…", pick).grid(row=0, column=1, padx=(8, 0))

        # Hinweis
        hint = (
            "Required columns:\n"
            "  - Name, Marketplace, BE ID, Schedule Start Date, Schedule End Date\n"
            "Optional:\n"
            "  - Description\n"
            "Variables (optional, any subset):\n"
            "  notificationTitle, notificationText, primaryButtonText, primaryButtonCta, url,\n"
            "  consolidationKey, hubImage, androidIconImage, iosImageOrVideo, androidBigPicture"
        )
        ttk.Label(body, text=hint, style="AmazonMuted.TLabel", justify="left").grid(row=6, column=0, columnspan=2, sticky="w", pady=(8, 0))

        # Actions
        actions = ttk.Frame(dlg, style="Amazon.TFrame")
        actions.pack(fill="x", padx=16, pady=(8, 12))

        result = {"template": None, "template_path": None, "xlsx_path": None}
        def submit():
            if not tpl_names:
                messagebox.showerror("Create RC Sonar", "No templates.")
                return
            name = tpl_var.get().strip()
            try:
                idx = tpl_names.index(name)
            except ValueError:
                messagebox.showerror("Create RC Sonar", "Please select a template.")
                return
            tpath = (tpl_path_var.get() or "").strip()
            if not tpath or not tpath.startswith("/"):
                messagebox.showerror("Create RC Sonar", "Please enter a valid Template Path (e.g., /LAYOUT-TEMPLATES/<uuid>).")
                return
            p = (path_var.get() or "").strip()
            if not p:
                messagebox.showerror("Create RC Sonar", "Please choose an Excel/CSV file.")
                return
            if not os.path.exists(p):
                messagebox.showerror("Create RC Sonar", f"File not found:\n{p}")
                return

            result["template"] = self.templates[idx]
            result["template_path"] = tpath
            result["xlsx_path"] = p
            dlg.destroy()

        def cancel():
            dlg.destroy()

        self.create_secondary_button(actions, "Cancel", cancel).pack(side="right", padx=(0, 8))
        self.create_amazon_button(actions, "Proceed", submit).pack(side="right")

        dlg.bind("<Escape>", lambda e: cancel())
        dlg.wait_window()
        if result["template"] and result["template_path"] and result["xlsx_path"]:
            return result
        return None




    def get_be_ids(self):
        """BE-IDs robust via Regex extrahieren (v2-Dialog liefert Rohtext)."""
        raw_input = self.show_be_input_dialog()
        if not raw_input:
            return None

        # 10-stellige IDs extrahieren, Reihenfolge beibehalten, Duplikate entfernen
        ids = re.findall(r'\b\d{10}\b', raw_input)
        seen = set()
        segment_ids = []
        for i in ids:
            if i not in seen:
                seen.add(i)
                segment_ids.append(i)

        if not segment_ids:
            messagebox.showerror("Error", "No valid segment IDs found!\nIDs must be 10 digits.")
            return None

        if self.show_confirm_ids_dialog(segment_ids):
            return segment_ids
        return None




    def _make_scrollable_checks(self, parent, labels):
        """
        Erzeugt eine scrollbare Liste von Checkbuttons.
        Rückgabe: (frame, vars) – frame in Grid/Pack einsetzen, vars ist Liste[tk.BooleanVar].
        """
        # Canvas + Scrollbar
        wrap = ttk.Frame(parent, style="Amazon.TFrame")
        canvas = tk.Canvas(wrap, bg=self.AMAZON["bg"], highlightthickness=0, bd=0)
        vsb = ttk.Scrollbar(wrap, orient="vertical", style="Amazon.Vertical.TScrollbar", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)

        inner = ttk.Frame(canvas, style="Amazon.TFrame")
        inner_id = canvas.create_window((0, 0), window=inner, anchor="nw")

        def _on_configure(_=None):
            canvas.configure(scrollregion=canvas.bbox("all"))
            # Breite anpassen, damit keine abgeschnittenen Checkbuttons
            canvas.itemconfigure(inner_id, width=canvas.winfo_width())

        inner.bind("<Configure>", _on_configure)
        wrap.bind("<Configure>", _on_configure)

        canvas.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        vars_ = []
        for text in labels:
            v = tk.BooleanVar(value=False)
            cb = ttk.Checkbutton(inner, text=text, variable=v, style="Amazon.TCheckbutton")
            cb.pack(anchor="w", pady=2)
            vars_.append(v)

        return wrap, vars_


    def get_sizes(self, segment_ids):
        """Run get_sizes in a worker thread; UI stays responsive."""
        try:
            self._set_busy(True)
            start_time = time.time()
            self.progress_var.set(0)
            self.progress_label.config(text="Starting...")
            self.status_var.set("Getting segment sizes...")

            # collected stats
            self.collected_stats = {
                'average_batch_time': "-",
                'average_success_rate': 0,
                'total_segments': len(segment_ids),
                'failed_segments': 0,
                'avg_segment_time': "-"
            }

            total = len(segment_ids)

            def ui_progress(i):
                self.root.after(0, lambda: self.update_progress(i, total))

            def ui_status(message):
                def _update():
                    self.status_var.set(message)
                    stats = self.parse_performance_stats(message)
                    if stats:
                        self.collected_stats.update(stats)
                        self.update_metrics(total_time=time.time() - start_time, stats=self.collected_stats)
                self.root.after(0, _update)

            def on_done(result):
                self._set_busy(False)
                if result is not None:
                    df, filename = result
                    elapsed = time.time() - start_time
                    self.progress_label.config(text=f"Completed! Processed {len(segment_ids)} segments in {elapsed:.1f} seconds")
                    self.status_var.set(f"Segment sizes saved to {filename}")
                    # --> Datei in Results aufnehmen
                    self._add_result_file(filename)
                    messagebox.showinfo("Success", f"Processed {len(segment_ids)} segments!\nResults saved to {filename}")
                else:
                    messagebox.showerror("Error", "Operation failed.")

            def on_error(err):
                self._set_busy(False)
                self.status_var.set(f"Error: {err}")
                messagebox.showerror("Error", str(err))

            import threading
            def run():
                try:
                    result = get_segment_sizes(
                        segment_ids,
                        status_callback=ui_status,
                        progress_callback=ui_progress,
                        headless=self.headless_var.get()
                    )
                    self.root.after(0, lambda: on_done(result))
                except Exception as e:
                    self.root.after(0, lambda err=e: on_error(err))

            threading.Thread(target=run, daemon=True).start()

        except Exception as e:
            self._set_busy(False)
            self.status_var.set(f"Error: {str(e)}")
            messagebox.showerror("Error", str(e))






    def extract_rules(self, segment_ids):
        """Run extract_rules in a worker thread; UI stays responsive."""
        try:
            self._set_busy(True)
            start_time = time.time()
            self.progress_var.set(0)
            self.progress_label.config(text="Starting...")
            self.status_var.set("Extracting rules...")

            self.collected_stats = {
                'average_batch_time': "-",
                'average_success_rate': 0,
                'total_segments': len(segment_ids),
                'failed_segments': 0
            }
            total = len(segment_ids)

            def ui_progress(i):
                self.root.after(0, lambda: self.update_progress(i, total))

            def ui_status(message):
                def _update():
                    self.status_var.set(message)
                    stats = self.parse_performance_stats(message)
                    if stats:
                        self.collected_stats.update(stats)
                        self.update_metrics(total_time=time.time() - start_time, stats=self.collected_stats)
                self.root.after(0, _update)

            def on_done(results):
                self._set_busy(False)
                if results is not None:
                    dfs, filename = results
                    elapsed = time.time() - start_time
                    self.progress_label.config(text=f"Completed! Processed {len(segment_ids)} segments in {elapsed:.1f} seconds")
                    self.status_var.set(f"Rules saved to {filename}")
                    # --> Datei in Results aufnehmen
                    self._add_result_file(filename)
                    messagebox.showinfo("Success", f"Extracted rules for {len(segment_ids)} segments!\nSaved to {filename}")
                else:
                    messagebox.showerror("Error", "Operation failed.")

            def on_error(err):
                self._set_busy(False)
                self.status_var.set(f"Error: {err}")
                messagebox.showerror("Error", str(err))

            import threading
            def run():
                try:
                    results = run_extract_rules(
                        segment_ids,
                        status_callback=ui_status,
                        progress_callback=ui_progress,
                        headless=self.headless_var.get()
                    )
                    self.root.after(0, lambda: on_done(results))
                except Exception as e:
                    self.root.after(0, lambda err=e: on_error(err))

            threading.Thread(target=run, daemon=True).start()

        except Exception as e:
            self._set_busy(False)
            self.status_var.set(f"Error: {str(e)}")
            messagebox.showerror("Error", str(e))
   



    def queue_segments(self, segment_ids):
        """Run queue_segments in a worker thread; UI stays responsive."""
        try:
            self._set_busy(True)
            start_time = time.time()
            self.progress_var.set(0)
            self.progress_label.config(text="Starting...")
            self.status_var.set("Queueing segments...")

            self.collected_stats = {
                'average_batch_time': "-",
                'average_success_rate': 0,
                'total_segments': len(segment_ids),
                'failed_segments': 0
            }
            total = len(segment_ids)

            def ui_progress(i):
                self.root.after(0, lambda: self.update_progress(i, total))

            def ui_status(message):
                def _update():
                    self.status_var.set(message)
                    stats = self.parse_performance_stats(message)
                    if stats:
                        self.collected_stats.update(stats)
                        self.update_metrics(total_time=time.time() - start_time, stats=self.collected_stats)
                self.root.after(0, _update)

            def on_done(results):
                self._set_busy(False)
                if results is not None:
                    df, filename = results
                    elapsed = time.time() - start_time
                    self.progress_label.config(text=f"Completed! Processed {len(segment_ids)} segments in {elapsed:.1f} seconds")
                    self.status_var.set(f"Queue results saved to {filename}")
                    # --> Datei in Results aufnehmen
                    self._add_result_file(filename)
                    messagebox.showinfo("Success", f"Queued {len(segment_ids)} segments!\nResults saved to {filename}")
                else:
                    messagebox.showerror("Error", "Operation failed.")

            def on_error(err):
                self._set_busy(False)
                self.status_var.set(f"Error: {err}")
                messagebox.showerror("Error", str(err))

            import threading
            def run():
                try:
                    results = run_queue_segments(
                        segment_ids,
                        status_callback=ui_status,
                        progress_callback=ui_progress,
                        headless=self.headless_var.get()
                    )
                    self.root.after(0, lambda: on_done(results))
                except Exception as e:
                    self.root.after(0, lambda err=e: on_error(err))

            threading.Thread(target=run, daemon=True).start()

        except Exception as e:
            self._set_busy(False)
            self.status_var.set(f"Error: {str(e)}")
            messagebox.showerror("Error", str(e))




    def update_campaign_content(self):
        """
        Öffnet den 'Update Content (Sonar)'-Dialog; der Dialog sammelt die Eingaben,
        baut die Jobs und führt sie in einem Worker-Thread aus.
        """
        try:
            self._show_update_content_dialog()
        except Exception as e:
            self._set_busy(False)
            self.status_var.set(f"Error: {e}")
            messagebox.showerror("Update Content", str(e))


    def _show_update_content_dialog(self):
        """
        Update Content (Sonar):
        - Template path (einmal, gilt für alle)
        - OPTIONAL: Excel (.xlsx) mit Spalten:
            * 'Sonar Link'  (Pflicht je Zeile)
            * beliebige Variable-Spalten (z. B. notificationTitle, url, ...)
            * optional: supportedLanguages (JSON-Array), marketplaceId, useJsonVariables
          Reihenfolge egal. Pro Zeile wird ein Job gebaut.
        - Falls KEINE Excel ausgewählt ist, bleibt der manuelle Modus (Kampagnen-Textbox + Datasets).
        """
        import tkinter as tk
        from tkinter import ttk, messagebox, filedialog
        import json as _json

        dlg = tk.Toplevel(self.root)
        dlg.title("Update Content (Sonar)")
        dlg.configure(bg=self.AMAZON["bg"])
        dlg.geometry("860x620")
        dlg.minsize(820, 560)
        dlg.transient(self.root)
        dlg.grab_set()

        # center
        dlg.update_idletasks()
        x = self.root.winfo_rootx() + (self.root.winfo_width() // 2 - dlg.winfo_width() // 2)
        y = self.root.winfo_rooty() + (self.root.winfo_height() // 2 - dlg.winfo_height() // 2)
        dlg.geometry(f"+{max(0, x)}+{max(0, y)}")

        # Top-Level auf Grid umstellen – Row 1 (Body) wächst, Row 2 (Buttons) bleibt fix
        dlg.grid_rowconfigure(1, weight=1)
        dlg.grid_columnconfigure(0, weight=1)

        # Header
        hdr = tk.Frame(dlg, bg=self.AMAZON["bg"])
        hdr.grid(row=0, column=0, sticky="ew", padx=16, pady=(14, 6))
        ttk.Label(hdr, text="Update Content for Sonar Campaigns", style="AmazonTitle.TLabel").pack(anchor="w")
        ttk.Label(
            hdr,
            text="Option A: Excel (.xlsx) mit 'Sonar Link' + Variablen.  Option B: manuell Kampagnen auflisten.",
            style="AmazonSubtitle.TLabel"
        ).pack(anchor="w", pady=(2, 0))

        # Body
        body = ttk.Frame(dlg, style="Amazon.TFrame")
        body.grid(row=1, column=0, sticky="nsew", padx=16, pady=(8, 0))
        body.columnconfigure(0, weight=1)

        # ---- Row 1: Template path
        ttk.Label(body, text="Template path (e.g., /LAYOUT-TEMPLATES/<uuid>)", style="AmazonBody.TLabel") \
            .grid(row=0, column=0, sticky="w")
        tpl_var = tk.StringVar(value="")
        tpl_entry = ttk.Entry(body, textvariable=tpl_var)
        tpl_entry.grid(row=1, column=0, sticky="ew", pady=(0, 10))
        tpl_entry.focus_set()

        # ---- Row 2: Excel chooser (sichtbar bleibt)
        row2 = ttk.Frame(body, style="Amazon.TFrame")
        row2.grid(row=2, column=0, sticky="ew")
        row2.columnconfigure(0, weight=1)
        ttk.Label(row2, text="Excel (.xlsx)", style="AmazonBody.TLabel").grid(row=0, column=0, sticky="w")

        xwrap = ttk.Frame(row2, style="Amazon.TFrame")
        xwrap.grid(row=1, column=0, sticky="ew", pady=(2, 10))
        xwrap.columnconfigure(0, weight=1)
        excel_path_var = tk.StringVar(value="")
        ttk.Entry(xwrap, textvariable=excel_path_var).grid(row=0, column=0, sticky="ew")
        self.create_secondary_button(
            xwrap, "Browse…", lambda: excel_path_var.set(filedialog.askopenfilename(
                title="Choose Excel file",
                filetypes=[("Excel", "*.xlsx"), ("All files", "*.*")]
            ) or excel_path_var.get())
        ).grid(row=0, column=1, padx=(8, 0))

        # ---- Advanced settings (alles andere eingeklappt)
        adv_frame = ttk.Frame(body, style="Amazon.TFrame")
        adv_frame.grid(row=4, column=0, sticky="nsew")
        adv_frame.grid_remove()  # Start: versteckt
        body.rowconfigure(4, weight=1)  # wenn offen, darf wachsen

        adv_open = tk.BooleanVar(value=False)

        def toggle_adv():
            adv_open.set(not adv_open.get())
            if adv_open.get():
                adv_frame.grid()
                adv_btn.config(text="Advanced settings ▾")
            else:
                adv_frame.grid_remove()
                adv_btn.config(text="Advanced settings ▸")

        adv_btn = self.create_secondary_button(body, "Advanced settings ▸", toggle_adv)
        adv_btn.grid(row=3, column=0, sticky="w", pady=(0, 6))

        # --- Advanced: Use JSON variables
        use_json_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            adv_frame,
            text="Use JSON variables (Correios.use-json-variables=true)",
            variable=use_json_var,
            style="Amazon.TCheckbutton"
        ).grid(row=0, column=0, sticky="w", pady=(0, 10))

        # --- Advanced: Manuell – Kampagnen + Datasets
        two_col = ttk.Frame(adv_frame, style="Amazon.TFrame")
        two_col.grid(row=1, column=0, sticky="nsew", pady=(0, 10))
        two_col.columnconfigure(0, weight=1, uniform="cols")
        two_col.columnconfigure(1, weight=1, uniform="cols")
        two_col.rowconfigure(1, weight=1)

        ttk.Label(two_col, text="Campaigns (one per line: ID or Sonar URL)",
                  style="AmazonBody.TLabel").grid(row=0, column=0, sticky="w", padx=(0, 8))
        ttk.Label(two_col, text="From SONAR datasets",
                  style="AmazonBody.TLabel").grid(row=0, column=1, sticky="w", padx=(8, 0))

        camp_text = tk.Text(two_col, wrap="none", height=12, font=("Segoe UI", 10), bg=self.AMAZON["bg"])
        camp_text.grid(row=1, column=0, sticky="nsew", padx=(0, 8))

        sonar_sets = self.filter_datasets("SONAR")
        sonar_names = [ds.get("name", "Unnamed") for ds in sonar_sets] or ["No SONAR datasets yet"]
        ds_frame, ds_vars = self._make_scrollable_checks(two_col, sonar_names)
        ds_frame.grid(row=1, column=1, sticky="nsew", padx=(8, 0))

        # --- Advanced: supportedLanguages + Detected
        ttk.Label(adv_frame,
                  text="supportedLanguages (auto) — Fallback als JSON, falls Links keine MP enthalten",
                  style="AmazonBody.TLabel").grid(row=2, column=0, sticky="w")
        langs_var = tk.StringVar(value="[]")
        langs_entry = ttk.Entry(adv_frame, textvariable=langs_var, state="readonly")
        langs_entry.grid(row=3, column=0, sticky="ew")
        detected_var = tk.StringVar(value="Detected: –")
        ttk.Label(adv_frame, textvariable=detected_var,
                  style="AmazonMuted.TLabel").grid(row=4, column=0, sticky="w", pady=(4, 0))

        # Helpers für Detection
        def _collect_campaign_lines() -> list[str]:
            lines = [l.strip() for l in camp_text.get("1.0", "end").splitlines() if l.strip()]
            if sonar_sets and any(v.get() for v in ds_vars):
                for i, v in enumerate(ds_vars):
                    if v.get():
                        lines.extend(self.ds_items(sonar_sets[i], "SONAR"))
            return lines

        def _refresh_detection(*_):
            # Wenn Excel gewählt ist, Detection nicht erzwingen
            if (excel_path_var.get() or "").strip():
                detected_var.set("Detected: (Excel mode)")
                langs_entry.configure(state="normal")
                langs_var.set("[]")
                langs_entry.configure(state="readonly")
                return

            lines = _collect_campaign_lines()
            groups = self._group_campaigns_by_mp(lines)

            chips = []
            for mp, data in groups.items():
                n = len(data["items"])
                if mp is None:
                    chips.append(f"Unknown × {n}")
                else:
                    lang = data["lang"] or "?"
                    country = {3: "UK", 4: "DE", 5: "FR", 35691: "IT", 44551: "ES"}.get(mp, str(mp))
                    chips.append(f"{country} ({lang}) × {n}")

            detected_var.set("Detected: " + (", ".join(chips) if chips else "–"))

            known_mps = [mp for mp in groups.keys() if mp is not None]
            has_unknown = any(mp is None for mp in groups.keys())
            if len(known_mps) == 1 and not has_unknown:
                only_mp = known_mps[0]
                lang = self._language_for_mp(only_mp)
                langs_entry.configure(state="normal")
                langs_var.set(_json.dumps([lang] if lang else []))
                langs_entry.configure(state="readonly")
            else:
                langs_entry.configure(state="normal")
                langs_var.set("[]")
                langs_entry.configure(state="readonly")

        camp_text.bind("<KeyRelease>", _refresh_detection)
        for v in ds_vars:
            v.trace_add("write", lambda *_: _refresh_detection())
        _refresh_detection()

        # ---- Footer / Actions – eigene feste Zeile unten
        actions = ttk.Frame(dlg, style="Amazon.TFrame")
        actions.grid(row=2, column=0, sticky="ew", padx=16, pady=12)
        actions.columnconfigure(0, weight=1)  # Spacer

        def cancel():
            try:
                dlg.destroy()
            except Exception:
                pass

        def proceed():
            template_path = (tpl_var.get() or "").strip()
            if not template_path:
                messagebox.showerror("Update Content", "Please enter a template path.")
                return

            xlsx = (excel_path_var.get() or "").strip()
            use_json_vars = bool(use_json_var.get())

            if xlsx:
                # A) Excel: dynamische Variablen, nur Kampagnen-Spalte ist speziell
                try:
                    jobs = self._jobs_from_update_excel_any_vars(xlsx, template_path, use_json_vars)
                except Exception as e:
                    messagebox.showerror("Update Content", str(e))
                    return
            else:
                # B) Kein Excel: manueller Modus
                lines = _collect_campaign_lines()
                if not lines:
                    messagebox.showerror(
                        "Update Content",
                        "Please provide campaigns (ID or Sonar URL) or choose an Excel file."
                    )
                    return

                groups = self._group_campaigns_by_mp(lines)
                try:
                    fallback_langs = _json.loads((langs_var.get() or "[]").strip() or "[]")
                    if not isinstance(fallback_langs, list):
                        raise ValueError
                except Exception:
                    messagebox.showerror(
                        "Update Content",
                        'supportedLanguages must be a JSON array, e.g. ["language_en_GB"].'
                    )
                    return

                jobs = []
                for mp, data in groups.items():
                    if mp is None:
                        if not fallback_langs:
                            messagebox.showerror(
                                "Update Content",
                                "Some campaigns have no marketplace in the link. "
                                "Either paste Sonar URLs with '#/<mp>/…' or provide a fallback in 'supportedLanguages'."
                            )
                            return
                        langs = fallback_langs
                    else:
                        lang = self._language_for_mp(mp)
                        if not lang:
                            messagebox.showerror("Update Content", f"No language mapping for MP {mp}.")
                            return
                        langs = [lang]

                    for item in data["items"]:
                        jobs.append({
                            "campaigns": [item["raw"]],
                            "template_path": template_path,
                            "supported_languages": langs,
                            "use_json_variables": use_json_vars,
                            "extra_variables": {},  # manueller Modus ohne Excel
                        })

            if not jobs:
                messagebox.showerror("Update Content", "Nothing to do.")
                return

            # --- Ausführen (ein POST pro Job) ---
            self._set_busy(True)
            self.status_var.set("Updating campaign content…")

            import threading, time as _time
            start_ts = _time.time()
            total_items = len(jobs)

            def status_cb(msg):
                self.root.after(0, lambda: self.status_var.set(msg))

            def worker():
                try:
                    from update_campaign_content import run_update_campaign_content
                except Exception as e:
                    err = str(e)
                    self.root.after(0, lambda: (
                        self._set_busy(False),
                        messagebox.showerror("Update Content", f"Import failed:\n{err}")
                    ))
                    return

                out_paths = []
                try:
                    for idx, j in enumerate(jobs):
                        status_cb(f"Updating {idx+1}/{total_items} … {j['supported_languages']}")
                        out_path_tuple = run_update_campaign_content(
                            job=j,
                            status_callback=status_cb,
                            progress_callback=lambda _i: None,
                            headless=self.headless_var.get()
                        )
                        out_paths.append(out_path_tuple[1] if isinstance(out_path_tuple, (list, tuple)) else out_path_tuple)
                        self.root.after(0, lambda k=idx: self.update_progress(k, total_items))

                    def done():
                        self._set_busy(False)
                        for p in out_paths:
                            if p:
                                self._add_result_file(p)
                        elapsed = _time.time() - start_ts
                        self.progress_label.config(text=f"Updated {total_items} campaign(s) in {elapsed:.1f}s")
                        messagebox.showinfo("Update Content", "Finished. Results saved.")
                        try:
                            if dlg.winfo_exists():
                                dlg.destroy()
                        except Exception:
                            pass

                    self.root.after(0, done)
                except Exception as e:
                    err_msg = str(e)
                    self.root.after(0, lambda m=err_msg: (
                        self._set_busy(False),
                        messagebox.showerror("Update Content", m)
                    ))

            threading.Thread(target=worker, daemon=True).start()

            # Dialog sofort schließen – Fortschritt im Hauptfenster
            try:
                dlg.destroy()
            except Exception:
                pass

        # Buttons unten rechts, immer sichtbar
        self.create_secondary_button(actions, "Cancel", cancel).grid(row=0, column=1, padx=(0, 8), sticky="e")
        self.create_amazon_button(actions, "Proceed", proceed).grid(row=0, column=2, sticky="e")

        dlg.bind("<Escape>", lambda e: cancel())






    def _parse_extra_variables(self, raw_text: str):
        """
        Akzeptiert:
          - JSON-Objekt: {"key":"val","obj":{"a":1},"arr":[1,2]}
          - ODER Zeilen im Format: name=value
            (value bleibt als String; JSON-erkennbare Werte wie {…} oder […] werden geparst)
        Liefert dict[str, (str|dict|list)] zurück.
        """
        text = (raw_text or "").strip()
        if not text:
            return {}

        try:
            val = json.loads(text)
            if isinstance(val, dict):
                return val
        except Exception:
            pass

        out = {}
        for ln in text.splitlines():
            if not ln.strip():
                continue
            if "=" not in ln:
                raise ValueError(f"Zeile ohne '=' gefunden: {ln!r}")
            name, value = ln.split("=", 1)
            name = name.strip()
            value = value.strip()
            if not name:
                raise ValueError(f"Leerer Variablenname in Zeile: {ln!r}")
            if (value.startswith("{") and value.endswith("}")) or (value.startswith("[") and value.endswith("]")):
                try:
                    out[name] = json.loads(value)
                    continue
                except Exception:
                    pass
            out[name] = value
        return out



    def _read_update_excel(self, path: str) -> list[dict]:
        """
        Liest .xlsx → list[dict] mit ORIGINALEN Spaltennamen (Case bleibt erhalten).
        Pandas bevorzugt, openpyxl als Fallback.
        """
        rows = []
        # 1) pandas
        try:
            import pandas as pd
            df = pd.read_excel(path)
            for _, row in df.iterrows():
                d = {}
                rd = row.to_dict()
                for k, v in rd.items():
                    key = (str(k) if k is not None else "").strip()  # KEIN .lower()
                    if not key:
                        continue
                    if hasattr(pd, "isna") and pd.isna(v):
                        v = ""
                    d[key] = v
                rows.append(d)
            return rows
        except Exception:
            pass

        # 2) openpyxl
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filename=path, read_only=True, data_only=True)
            ws = wb.active
            headers = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [(str(c).strip() if c is not None else "") for c in row]
                    continue
                d = {}
                for j, val in enumerate(row):
                    key = headers[j] if j < len(headers) else f"col{j+1}"
                    if key:
                        d[key] = "" if val is None else val
                rows.append(d)
            wb.close()
            return rows
        except Exception as e:
            raise RuntimeError(f"Could not read Excel:\n{path}\n\n{e}")

    def _get_ci(self, row: dict, *names):
        """
        Case-insensitive Feldzugriff: gibt den ersten Treffer zurück (Wert),
        wenn einer der 'names' als Key (egal in welchem Case) existiert.
        """
        lowmap = {str(k).strip().lower(): k for k in row.keys()}
        for n in names:
            k = lowmap.get(str(n).strip().lower())
            if k is not None:
                return row.get(k)
        return None

    def _pick_campaign_cell(self, row: dict) -> str:
        """
        Akzeptierte Spaltennamen (case-insensitiv) für die Kampagne:
        'Sonar Link', 'campaign', 'campaignId', 'url', 'sonarlink', 'link'
        """
        val = self._get_ci(row, "Sonar Link", "campaign", "campaignId", "url", "sonarlink", "link")
        return "" if val in (None, "") else str(val).strip()

    def _derive_langs_from_row(self, row: dict, ui_fallback_langs: list[str]) -> list[str]:
        """
        Ermittelt supportedLanguages für einen Excel-Datensatz:
        1) Spalte 'supportedLanguages' (JSON-Array) – falls vorhanden
        2) 'marketplaceId' → Mapping zu Sprache
        3) MP aus Sonar-Link '#/<mp>/...' → Sprache
        4) UI-Fallback (langs_var aus Dialog)
        """
        # 1) explicit supportedLanguages
        sl = self._get_ci(row, "supportedLanguages")
        if sl not in (None, ""):
            try:
                if isinstance(sl, list):
                    return [str(x) for x in sl]
                if isinstance(sl, str):
                    val = json.loads(sl)
                    if isinstance(val, list):
                        return [str(x) for x in val]
            except Exception:
                pass

        # 2) marketplaceId
        mp_raw = self._get_ci(row, "marketplaceId")
        try:
            if mp_raw not in (None, ""):
                mp = int(str(mp_raw).strip())
                lang = self._language_for_mp(mp)
                if lang:
                    return [lang]
        except Exception:
            pass

        # 3) extract from Sonar link
        camp = self._pick_campaign_cell(row)
        mp = self._extract_mp_from_line(camp)
        if mp:
            lang = self._language_for_mp(mp)
            if lang:
                return [lang]

        # 4) UI fallback
        return list(ui_fallback_langs or [])

    def _build_jobs_from_update_excel(
        self,
        rows: list[dict],
        template_path: str,
        ui_fallback_langs: list[str],
        use_json_vars: bool
    ) -> list[dict]:
        """
        Baut pro Excel-Zeile genau EINEN Job.
        Reservierte Spalten (case-insensitiv) werden NICHT als Variable an die API geschickt:
          - 'Sonar Link', 'campaign', 'campaignId', 'url', 'sonarlink', 'link'
          - 'supportedLanguages', 'marketplaceId', 'useJsonVariables'
        Alle anderen Spalten werden 1:1 als Variables übernommen (Key = Original-Spaltenname).
        JSON in Zellen (beginnend mit '{' oder '[') wird gepaart; Zahlen/Booleans bleiben Zahlen/Booleans.
        """
        reserved_lower = {
            "sonar link", "campaign", "campaignid", "url", "sonarlink", "link",
            "supportedlanguages", "marketplaceid", "usejsonvariables"
        }

        def _should_skip(col_name: str) -> bool:
            return str(col_name).strip().lower() in reserved_lower

        def _coerce_value(v):
            # schon richtige Typen durchlassen
            if isinstance(v, (dict, list, bool, int, float)):
                return v
            s = str(v).strip()
            if not s:
                return ""
            # JSON?
            if (s.startswith("{") and s.endswith("}")) or (s.startswith("[") and s.endswith("]")):
                try:
                    return json.loads(s)
                except Exception:
                    pass
            # Zahlen?
            try:
                if s.isdigit() or (s.startswith("-") and s[1:].isdigit()):
                    return int(s)
                # float?
                if any(ch in s for ch in (".", ",")):
                    s2 = s.replace(",", ".")
                    return float(s2)
            except Exception:
                pass
            # true/false?
            if s.lower() in ("true", "false"):
                return s.lower() == "true"
            return s  # als String

        jobs: list[dict] = []
        for i, r in enumerate(rows, start=1):
            campaign_entry = self._pick_campaign_cell(r)
            if not campaign_entry:
                raise ValueError(f"Row {i}: missing 'Sonar Link' (or campaign/url/campaignId).")

            langs = self._derive_langs_from_row(r, ui_fallback_langs)

            extra: dict = {}
            for col, val in r.items():
                if not str(col).strip():
                    continue
                if _should_skip(col):
                    continue
                extra[str(col)] = _coerce_value(val)

            job = {
                "campaigns": [campaign_entry],     # genau 1 Kampagne pro Zeile
                "template_path": template_path,
                "supported_languages": langs,
                "use_json_variables": bool(use_json_vars),
                "extra_variables": extra,
            }
            jobs.append(job)
        return jobs








    # ------------------- Create OS Sonar (Template + Excel) -------------------

    def create_os_sonar(self):
        """
        Orchestriert den Flow:
          1) Dialog: Template wählen + Excel-Datei wählen
          2) Excel laden und validieren
          3) Jobs aus Zeilen + Template + Fixwerten bauen
          4) an Stub-Runner übergeben (wird später durch echte Sonar-Calls ersetzt)
        """
        if not self.templates:
            if messagebox.askyesno("Templates", "No templates yet. Create one now?"):
                self.open_templates_manager()
            return

        dlg_data = self._show_create_os_dialog()
        if not dlg_data:
            return

        template = dlg_data["template"]
        xlsx_path = dlg_data["xlsx_path"]

        try:
            rows = self._read_campaign_excel(xlsx_path)
        except Exception as e:
            messagebox.showerror("Create OS Sonar", f"Could not read Excel:\n{xlsx_path}\n\n{e}")
            return

        try:
            rows_ok = self._validate_campaign_rows(rows)
        except Exception as e:
            messagebox.showerror("Create OS Sonar", str(e))
            return

        # Jobs bauen
        jobs = self._build_jobs_from_rows(rows_ok, template)

        # Ausführen (Stub; ersetzt wir im nächsten Schritt durch echte API)
        self._run_create_os_sonar(jobs)

    def _show_create_os_dialog(self):
        """
        UI-Dialog:
         - Template-Auswahl (Combobox)
         - Excel-Datei auswählen
         - Kurze Spaltenanforderung anzeigen
        Rückgabe: {"template": dict, "xlsx_path": str} oder None
        """
        dlg = tk.Toplevel(self.root)
        dlg.title("Create OS Sonar")
        dlg.configure(bg=self.AMAZON["bg"])
        dlg.geometry("560x360")
        dlg.minsize(520, 320)
        dlg.transient(self.root)
        dlg.grab_set()

        # center
        dlg.update_idletasks()
        x = self.root.winfo_rootx() + (self.root.winfo_width() // 2 - dlg.winfo_width() // 2)
        y = self.root.winfo_rooty() + (self.root.winfo_height() // 2 - dlg.winfo_height() // 2)
        dlg.geometry(f"+{max(0, x)}+{max(0, y)}")

        body = ttk.Frame(dlg, style="Amazon.TFrame")
        body.pack(fill="both", expand=True, padx=16, pady=12)
        for c in range(1):
            body.columnconfigure(c, weight=1)

        # Template
        ttk.Label(body, text="Template", style="AmazonBody.TLabel").grid(row=0, column=0, sticky="w")
        tpl_names = [t.get("name", f"Template {i+1}") for i, t in enumerate(self.templates)]
        tpl_var = tk.StringVar(value=(tpl_names[0] if tpl_names else ""))
        tpl_cb = ttk.Combobox(body, state="readonly", values=tpl_names, textvariable=tpl_var)
        tpl_cb.grid(row=1, column=0, sticky="ew", pady=(0, 10))

        # Excel chooser
        ttk.Label(body, text="Excel (.xlsx) with inputs", style="AmazonBody.TLabel").grid(row=2, column=0, sticky="w")
        path_var = tk.StringVar(value="")
        row3 = ttk.Frame(body, style="Amazon.TFrame")
        row3.grid(row=3, column=0, sticky="ew", pady=(0, 8))
        row3.columnconfigure(0, weight=1)
        e = ttk.Entry(row3, textvariable=path_var)
        e.grid(row=0, column=0, sticky="ew")
        def pick():
            p = filedialog.askopenfilename(
                title="Choose Excel file",
                filetypes=[("Excel", "*.xlsx"), ("All files", "*.*")]
            )
            if p:
                path_var.set(p)
        self.create_secondary_button(row3, "Browse…", pick).grid(row=0, column=1, padx=(8, 0))

        # Hint
        hint = (
            "Required columns (case-insensitive):\n"
            "  - programName (str)\n"
            "  - startDate (YYYY-MM-DD)\n"
            "Optional:\n"
            "  - programDescription (str; default = programName)\n"
            "  - marketplaceId (int; default = 4)\n"
        )
        ttk.Label(body, text=hint, style="AmazonMuted.TLabel", justify="left").grid(
            row=4, column=0, sticky="w", pady=(8, 0)
        )

        actions = ttk.Frame(dlg, style="Amazon.TFrame")
        actions.pack(fill="x", padx=16, pady=(8, 12))

        result = {"template": None, "xlsx_path": None}
        def submit():
            if not tpl_names:
                messagebox.showerror("Create OS Sonar", "No templates.")
                return
            name = tpl_var.get().strip()
            try:
                idx = tpl_names.index(name)
            except ValueError:
                messagebox.showerror("Create OS Sonar", "Please select a template.")
                return
            p = (path_var.get() or "").strip()
            if not p:
                messagebox.showerror("Create OS Sonar", "Please choose an Excel (.xlsx) file.")
                return
            if not os.path.exists(p):
                messagebox.showerror("Create OS Sonar", f"File not found:\n{p}")
                return
            result["template"] = self.templates[idx]
            result["xlsx_path"] = p
            dlg.destroy()

        def cancel():
            dlg.destroy()

        self.create_secondary_button(actions, "Cancel", cancel).pack(side="right", padx=(0, 8))
        self.create_amazon_button(actions, "Proceed", submit).pack(side="right")

        dlg.bind("<Escape>", lambda e: cancel())
        dlg.wait_window()
        if result["template"] and result["xlsx_path"]:
            return result
        return None

    def _read_campaign_excel(self, path):
        """
        Liest .xlsx → list[dict] mit Spaltennamen (lowercased keys).
        Versucht zuerst pandas, danach openpyxl (manual).
        """
        rows = []
        # 1) pandas
        try:
            import pandas as pd
            df = pd.read_excel(path)
            for _, row in df.iterrows():
                d = {str(k).strip().lower(): ("" if pd.isna(v) else v) for k, v in row.to_dict().items()}
                rows.append(d)
            return rows
        except Exception:
            pass

        # 2) openpyxl
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filename=path, read_only=True, data_only=True)
            ws = wb.active
            headers = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c).strip().lower() if c is not None else "" for c in row]
                    continue
                d = {}
                for j, val in enumerate(row):
                    key = headers[j] if j < len(headers) else f"col{j+1}"
                    d[key] = "" if val is None else val
                rows.append(d)
            wb.close()
            return rows
        except Exception as e:
            raise RuntimeError(
                f"Could not read Excel. Please install 'pandas' or 'openpyxl'.\n\n{e}"
            )


    def _validate_campaign_rows(self, rows):
        """
        Prüft Pflichtfelder & normalisiert:
          - programName: str (non-empty)
          - startDate: YYYY-MM-DD (tolerant: akzeptiert auch 'YYYY-MM-DD 00:00:00',
            Excel-Datumswerte, '01.10.2025', '01/10/2025', '10/01/2025' usw.)
        Optional:
          - programDescription: str (default = programName)
          - marketplaceId: int (default 4)
        Gibt normalisierte list[dict] zurück.
        """
        if not rows:
            raise ValueError("Excel is empty.")

        out = []
        for i, r in enumerate(rows, start=1):
            # Keys vereinheitlichen
            r = {str(k).strip().lower(): r[k] for k in r.keys()}

            # programName
            pname = str(r.get("programname", "")).strip()
            if not pname:
                raise ValueError(f"Row {i}: 'programName' is required.")
            pdesc = str(r.get("programdescription", "")).strip() or pname

            # startDate robust in 'YYYY-MM-DD' bringen
            sdate_raw = r.get("startdate", "")
            try:
                sdate = self._coerce_date_yyyy_mm_dd(sdate_raw)
            except Exception:
                raise ValueError(f"Row {i}: 'startDate' must be a valid date (got {sdate_raw!r}).")

            # marketplaceId (optional, Default 4)
            mp_raw = r.get("marketplaceid", "")
            try:
                mp = int(str(mp_raw).strip()) if str(mp_raw).strip() else 4
            except Exception:
                raise ValueError(f"Row {i}: 'marketplaceId' must be an integer.")

            # Normalisierte Zeile aufbauen
            out.append({
                "programName": pname,
                "programDescription": pdesc,
                "marketplaceId": mp,
                "startDate": sdate,
                # abgeleitet:
                "campaignName": pname,
                "campaignDescription": pdesc,
                "endDate": sdate,  # OneShot
            })

        return out




    def _build_jobs_from_rows(self, rows, template_dict):
        """
        Kombiniert Zeilen + Template + Fixwerte zu übergabefertigen Jobs.
        Gibt list[dict] zurück, jede Zeile ein Job.
        """
        # fixe Werte laut Vorgabe
        fixed = {
            "duration": 1,
            "reason": "OneShot",
            "topic": "CAFEP",
            "optOuts": [],
            "communicationContentType": {"optOutList": []},
        }

        # Template-Felder (nur die, die wir definiert hatten)
        tpl = {
            "channel": (template_dict.get("channel") or "").strip(),
            "teamBindle": template_dict.get("teamBindle"),
            "lobExpression": template_dict.get("lobExpression"),
            "managementType": template_dict.get("managementType"),
            "businessGroupId": template_dict.get("businessGroupId"),
            "familyId": template_dict.get("familyId"),
            "optOuts": template_dict.get("optOuts") or [],
            "startTimeMinutesOffset": template_dict.get("startTimeMinutesOffset"),
            "endTimeMinutesOffset": template_dict.get("endTimeMinutesOffset"),
        }

        # ggf. Profil-Infos (BusinessOwner etc.)
        owner = (self.profile or {}).get("alias") or ""

        jobs = []
        for r in rows:
            job = {
                "program": {
                    "name": r["programName"],
                    "description": r["programDescription"],
                    "marketplaceId": r["marketplaceId"],
                    "businessOwner": owner,
                    "creator": owner,
                    "type": "MANUAL",
                },
                "campaign": {
                    "name": r["campaignName"],
                    "description": r["campaignDescription"],
                    "startDate": r["startDate"],
                    "endDate": r["endDate"],  # OneShot
                },
                "template": tpl,
                "fixed": fixed,
            }
            jobs.append(job)
        return jobs



    



    def _run_create_os_sonar(self, jobs):
        """
        Create OS Sonar (real):
        - führt die Erstellung via create_os_sonar.run_create_os_sonar aus
        - zeigt Fortschritt/Status live an
        - nimmt die Ergebnisdatei in die Results-Liste auf
        """
        try:
            self._set_busy(True)
            self.progress_var.set(0)
            self.progress_label.config(text="Creating Sonar items…")
            self.status_var.set("Connecting to Sonar…")
            start_ts = time.time()
            total = max(1, len(jobs))

            def ui_status(msg: str):
                try:
                    self.status_var.set(msg)
                    self.root.update_idletasks()
                except Exception:
                    pass

            def ui_progress(i: int):
                try:
                    self.update_progress(i, total)
                except Exception:
                    pass

            def worker():
                try:
                    from create_os_sonar import run_create_os_sonar
                except Exception as e:
                    self.root.after(0, lambda: (
                        self._set_busy(False),
                        messagebox.showerror("Create OS Sonar", f"Import failed:\n{e}")
                    ))
                    return
                try:
                    results, out_path = run_create_os_sonar(
                        jobs=jobs,
                        status_callback=lambda m: self.root.after(0, lambda: ui_status(m)),
                        progress_callback=lambda i: self.root.after(0, lambda: ui_progress(i)),
                        headless=self.headless_var.get(),
                        requester_alias=(self.profile or {}).get("alias")
                    )
                    def on_done():
                        if out_path:
                            self._add_result_file(out_path)
                        elapsed = time.time() - start_ts
                        self.progress_label.config(text=f"Created {len(results)} item(s) in {elapsed:.1f}s")
                        self.status_var.set("Done.")
                        self._set_busy(False)
                        messagebox.showinfo("Create OS Sonar", "Finished. Results saved.")
                    self.root.after(0, on_done)
                except Exception as e:
                    self.root.after(0, lambda: (
                        self._set_busy(False),
                        messagebox.showerror("Create OS Sonar", str(e))
                    ))

            import threading
            threading.Thread(target=worker, daemon=True).start()

        except Exception as e:
            self._set_busy(False)
            messagebox.showerror("Create OS Sonar", str(e))





    def upload_be_to_sonar(self):
        """
        Bulk mode: multiple BE IDs mapped to matching Sonar Campaigns (one-to-one by line order).
        Läuft in einem Worker-Thread; UI-Updates gehen thread-sicher via root.after(...).
        """
        try:
            pairs = self.show_be_sonar_mapping_dialog()
            if not pairs:
                self.status_var.set("No mapping provided.")
                return

            # Reset Progress UI
            self.progress_var.set(0)
            self.progress_label.config(text="Starting...")
            start_time = time.time()
            total = len(pairs)

            # Thread-sichere Callback-Wrapper
            def ui_progress(i):
                # i = 0-based index
                self.root.after(0, lambda: self.update_progress(i, total))

            def ui_status(message):
                self.root.after(0, lambda: (self.status_var.set(message), self.root.update_idletasks()))

            # Abschluss-/Fehler-Handler im Tk-Thread
            def on_done(_df):
                # Versuche den Dateinamen zu ermitteln (Fallback: neueste .xlsx seit Start)
                out = self._guess_new_xlsx(start_time)
                if out:
                    self._add_result_file(out)

                elapsed = time.time() - start_time
                self.progress_label.config(text=f"Completed {total} pair(s) in {elapsed:.1f}s")
                messagebox.showinfo("Success", "Finished applying BE → Sonar (results saved).")

            def on_error(err):
                self.status_var.set(f"Error: {err}")
                messagebox.showerror("Error", str(err))
                self._set_busy(False)

            # Worker-Thread starten
            import threading
            def run():
                try:
                    from sonar_apply import apply_segments_to_sonar_pairs
                    df = apply_segments_to_sonar_pairs(
                        pairs=pairs,
                        status_callback=ui_status,
                        progress_callback=ui_progress,
                        headless=self.headless_var.get()
                    )
                    self.root.after(0, lambda: (self._set_busy(False), on_done(df)))
                except Exception as e:
                    self.root.after(0, lambda err=e: on_error(err))

            self._set_busy(True)
            threading.Thread(target=run, daemon=True).start()

        except Exception as e:
            self.status_var.set(f"Error: {str(e)}")
            messagebox.showerror("Error", str(e))



    
    def clone_and_publish(self):
        """
        Bulk clone & publish: prompts for mapping (BE -> Segment Name) und führt
        clone_publish im Hintergrund-Thread aus. UI bleibt responsiv.
        """
        try:
            pairs = self.show_be_name_mapping_dialog()
            if not pairs:
                self.status_var.set("No BE/name mapping provided.")
                return

            # Progress reset
            self.progress_var.set(0)
            self.progress_label.config(text="Starting clone & publish...")
            start_time = time.time()
            total = len(pairs)

            from clone_publish import clone_and_publish_segments

            def ui_progress(i):
                self.root.after(0, lambda: self.update_progress(i, total))

            def ui_status(message):
                self.root.after(0, lambda: (self.status_var.set(message), self.root.update_idletasks()))

            def on_done(_df):
                # Vermutete Ausgabedatei bestimmen
                out = self._guess_new_xlsx(start_time)
                if out:
                    self._add_result_file(out)

                elapsed = time.time() - start_time
                self.progress_label.config(text=f"Completed {total} pair(s) in {elapsed:.1f}s")
                messagebox.showinfo("Success", "Finished. Results saved.")

            def on_error(err):
                self.status_var.set(f"Error: {err}")
                messagebox.showerror("Error", str(err))
                self._set_busy(False)

            import threading
            def run():
                try:
                    df = run_clone_and_publish_segments(
                        pairs=pairs,
                        status_callback=ui_status,
                        progress_callback=ui_progress,
                        headless=self.headless_var.get(),
                        owner_alias=(self.profile or {}).get("alias")
                    )
                    self.root.after(0, lambda: (self._set_busy(False), on_done(df)))
                except Exception as e:
                    self.root.after(0, lambda err=e: on_error(err))

            self._set_busy(True)
            threading.Thread(target=run, daemon=True).start()

        except Exception as e:
            self.status_var.set(f"Error: {str(e)}")
            messagebox.showerror("Error", str(e))


    def clone_across_mps(self):
        """Clone a source BE to all other marketplaces (UK/DE/FR/IT/ES) with MP + hygiene rule adjusted."""
        try:
            source_be = self.show_single_be_dialog()
            if not source_be:
                self.status_var.set("No source BE ID provided.")
                return

            # Reset progress UI
            self.progress_var.set(0)
            self.progress_label.config(text="Starting clone across marketplaces...")
            start_time = time.time()

            from clone_publish import (
                clone_across_marketplaces as backend_clone_across_mps,
                MP_CODE_BY_ID as _MP_MAP,
            )

            total_targets = max(0, len(_MP_MAP) - 1)

            def progress_callback(i):
                self.root.after(0, lambda: self.update_progress(i, total_targets))

            def status_callback(message):
                self.root.after(0, lambda: (self.status_var.set(message), self.root.update_idletasks()))

            def on_done(result):
                if result is not None:
                    _results, out_main = result
                    # --> Datei in Results aufnehmen
                    if out_main:
                        self._add_result_file(out_main)
                    else:
                        # Fallback raten
                        out = self._guess_new_xlsx(start_time)
                        if out:
                            self._add_result_file(out)

                    elapsed = time.time() - start_time
                    self.progress_label.config(text=f"Completed in {elapsed:.1f}s")
                    self.status_var.set("Cloning across marketplaces finished.")
                    messagebox.showinfo("Success", "Finished. Results saved.")
                else:
                    self.status_var.set("Operation failed.")
                    messagebox.showerror("Error", "Operation failed.")
                self._set_busy(False)

            def on_error(err):
                self.status_var.set(f"Error: {err}")
                messagebox.showerror("Error", str(err))
                self._set_busy(False)

            import threading
            def run():
                try:
                    result = backend_clone_across_mps(
                        source_be_id=source_be,
                        status_callback=status_callback,
                        progress_callback=progress_callback,
                        headless=self.headless_var.get()

                    )
                    self.root.after(0, lambda: on_done(result))
                except Exception as e:
                    self.root.after(0, lambda err=e: on_error(err))

            self._set_busy(True)
            threading.Thread(target=run, daemon=True).start()

        except Exception as e:
            self.status_var.set(f"Error: {str(e)}")
            messagebox.showerror("Error", str(e))


    def mass_clone_fixed(self):
        """
        Startet den Mass-Clone-Flow im Hintergrund-Thread:
        - Dialog: Base-ID (Default oder Custom), Anzahl, Namen
        - Klont Base-Segment 'count' Mal mit den angegebenen Namen
        """
        try:
            data = self.show_mass_clone_fixed_dialog()
            if not data:
                self.status_var.set("Abgebrochen.")
                return

            base_id = data["base_id"]
            names = data["names"]
            pairs = [(base_id, nm) for nm in names]

            # Progress reset
            self.progress_var.set(0)
            self.progress_label.config(text="Starting mass clone...")
            start_time = time.time()
            total = len(pairs)

            from clone_publish import clone_and_publish_segments

            def ui_progress(i):
                self.root.after(0, lambda: self.update_progress(i, total))

            def ui_status(message):
                self.root.after(0, lambda: (self.status_var.set(message), self.root.update_idletasks()))

            def on_done(_df):
                out = self._guess_new_xlsx(start_time)
                if out:
                    self._add_result_file(out)

                elapsed = time.time() - start_time
                self.progress_label.config(text=f"Completed {total} clone(s) in {elapsed:.1f}s")
                messagebox.showinfo("Success", "Finished. Results saved.")

            def on_error(err):
                self.status_var.set(f"Error: {err}")
                messagebox.showerror("Error", str(err))
                self._set_busy(False)

            import threading
            def run():
                try:
                    df = run_clone_and_publish_segments(
                        pairs=pairs,
                        status_callback=ui_status,
                        progress_callback=ui_progress,
                        headless=self.headless_var.get(),
                        owner_alias=(self.profile or {}).get("alias")
                    )
                    self.root.after(0, lambda: (self._set_busy(False), on_done(df)))
                except Exception as e:
                    self.root.after(0, lambda err=e: on_error(err))

            self._set_busy(True)
            threading.Thread(target=run, daemon=True).start()

        except Exception as e:
            self.status_var.set(f"Error: {str(e)}")
            messagebox.showerror("Error", str(e))





    def show_be_name_mapping_dialog(self):
        """
        Amazon-style modal dialog to input BE IDs (left) and matching Segment Names (right).
        Returns list of tuples [(be_id_10digits, segment_name), ...] or None on cancel/validation error.
        """
        dlg = tk.Toplevel(self.root)
        dlg.title("Clone & Publish — BE → Segment Name")
        dlg.configure(bg=self.AMAZON["bg"])
        dlg.geometry("780x420")
        dlg.minsize(720, 380)
        dlg.transient(self.root)
        dlg.grab_set()

        # center on parent
        dlg.update_idletasks()
        x = self.root.winfo_rootx() + (self.root.winfo_width() // 2 - dlg.winfo_width() // 2)
        y = self.root.winfo_rooty() + (self.root.winfo_height() // 2 - dlg.winfo_height() // 2)
        dlg.geometry(f"+{max(0, x)}+{max(0, y)}")

        # Header
        hdr = tk.Frame(dlg, bg=self.AMAZON["bg"])
        hdr.pack(fill="x", padx=16, pady=(14, 6))
        ttk.Label(hdr, text="Paste BE IDs and matching Segment Names", style="AmazonTitle.TLabel").pack(anchor="w")
        ttk.Label(
            hdr,
            text="Enter one pair per line. Left box: BE IDs (10 digits). Right box: desired Segment Names. Lines are matched by order.",
            style="AmazonSubtitle.TLabel"
        ).pack(anchor="w", pady=(2, 0))

        # Body with two text areas
        body = tk.Frame(dlg, bg=self.AMAZON["bg"])
        body.pack(fill="both", expand=True, padx=16, pady=(8, 0))
        body.columnconfigure(0, weight=1)
        body.columnconfigure(1, weight=1)

        # Left (BE IDs)
        left_wrap = tk.Frame(body, bg=self.AMAZON["bg"])
        left_wrap.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        ttk.Label(left_wrap, text="Bullseye IDs (one per line)", style="AmazonBody.TLabel").pack(anchor="w", pady=(0, 4))
        be_text = tk.Text(left_wrap, wrap="none", height=12, font=("Segoe UI", 10))
        be_text.pack(fill="both", expand=True)

        # Right (Segment Names)
        right_wrap = tk.Frame(body, bg=self.AMAZON["bg"])
        right_wrap.grid(row=0, column=1, sticky="nsew", padx=(8, 0))
        ttk.Label(right_wrap, text="Segment Names (one per line)", style="AmazonBody.TLabel").pack(anchor="w", pady=(0, 4))
        name_text = tk.Text(right_wrap, wrap="none", height=12, font=("Segoe UI", 10))
        name_text.pack(fill="both", expand=True)

        # Actions
        actions = tk.Frame(dlg, bg=self.AMAZON["bg"])
        actions.pack(fill="x", padx=16, pady=12)

        result = {"pairs": None}

        def parse_pairs():
            # collect & clean lines
            be_lines_raw = be_text.get("1.0", "end").splitlines()
            name_lines_raw = name_text.get("1.0", "end").splitlines()

            be_lines = [l.strip() for l in be_lines_raw if l.strip()]
            name_lines = [l.strip() for l in name_lines_raw if l.strip()]

            if not be_lines or not name_lines:
                messagebox.showerror("Error", "Both columns must contain at least one line.")
                return

            if len(be_lines) != len(name_lines):
                messagebox.showerror("Error", "Line counts do not match. Each BE must have a matching Segment Name.")
                return

            # validate BE IDs → only 10 digits each
            cleaned_be = []
            for i, be in enumerate(be_lines, start=1):
                digits = "".join(ch for ch in be if ch.isdigit())
                if len(digits) != 10:
                    messagebox.showerror("Error", f"Line {i}: BE must be 10 digits (got: '{be}').")
                    return
                cleaned_be.append(digits)

            # validate names → non-empty
            for i, nm in enumerate(name_lines, start=1):
                if not nm:
                    messagebox.showerror("Error", f"Line {i}: Segment Name cannot be empty.")
                    return

            pairs = list(zip(cleaned_be, name_lines))
            result["pairs"] = pairs
            dlg.destroy()

        def cancel():
            result["pairs"] = None
            dlg.destroy()

        self.create_secondary_button(actions, "Cancel", cancel).pack(side="right", padx=(0, 8))
        self.create_amazon_button(actions, "Confirm", parse_pairs).pack(side="right")

        dlg.bind("<Escape>", lambda e: cancel())
        dlg.bind("<Return>", lambda e: parse_pairs())

        dlg.wait_window()
        return result["pairs"]


    def _read_update_excel_rows(self, path: str):
        """
        Liest XLSX/CSV und liefert eine Liste von Zeilen-Dicts mit
        *originalen* Spaltennamen (nicht lowercased). Leere Zellen -> "".
        """
        import os
        ext = os.path.splitext(path)[1].lower()
        try:
            import pandas as pd
        except Exception as e:
            raise RuntimeError("pandas wird zum Einlesen der Excel benötigt.") from e

        if ext in (".xlsx", ".xls"):
            df = pd.read_excel(path, dtype=str)
        else:
            # Fallback: CSV (Delimiter auto-erkennen)
            df = pd.read_csv(path, dtype=str, sep=None, engine="python")

        df = df.fillna("")
        records = df.to_dict(orient="records")

        # Spaltennamen trimmen, Werte zu Strings
        cleaned = []
        for rec in records:
            row = {}
            for k, v in rec.items():
                key = str(k).strip()
                val = "" if v is None else str(v)
                row[key] = val
            cleaned.append(row)
        return cleaned


    def _jobs_from_update_excel_any_vars(self, xlsx_path: str, template_path: str, use_json_vars: bool):
        """
        Baut 1 Job pro Zeile:
          - Kampagne aus 'Sonar Link' / 'campaign*' (nicht 'url')
          - Alle übrigen Spalten werden als Variablen übernommen
            (Variablen-Name = exakter Spaltenname in der Excel).
        """
        rows = self._read_update_excel_rows(xlsx_path)
        if not rows:
            raise ValueError("Excel ist leer.")

        jobs = []
        for idx, row in enumerate(rows, start=1):
            # Mapping: normalisierte Header -> Original-Header
            norm_to_orig = {self._norm_hdr(h): h for h in row.keys()}

            # Kampagne finden
            camp_val = None
            for key_norm in self.UPDATE_CAMPAIGN_HEADERS:
                if key_norm in norm_to_orig:
                    orig = norm_to_orig[key_norm]
                    val = str(row.get(orig, "")).strip()
                    if val:
                        camp_val = val
                        break
            if not camp_val:
                raise ValueError(f"Row {idx}: missing 'Sonar Link' (or campaign/url/campaignId).")

            # Variablen: alle Spalten außer den Kampagnen-Spalten
            extra_vars = {}
            for orig_hdr, value in row.items():
                if not str(value).strip():
                    continue
                if self._norm_hdr(orig_hdr) in self.UPDATE_CAMPAIGN_HEADERS:
                    continue  # Kampagnenfeld nicht als Variable senden
                # exakter Spaltenname = Variablenname
                extra_vars[orig_hdr.strip()] = value

            # Sprache aus Link ableiten (falls MP im Link)
            mp = self._extract_mp_from_line(camp_val)
            langs = [self._language_for_mp(mp)] if mp else []

            jobs.append({
                "campaigns": [camp_val],
                "template_path": template_path,
                "supported_languages": langs,
                "use_json_variables": use_json_vars,
                "extra_variables": extra_vars,
            })

        return jobs



    # ------------------- Helpers -------------------


    # ---- Update-Content (Excel) – dynamische Variablen ----
    def _norm_hdr(self, s: str) -> str:
        import re as _re
        return _re.sub(r'[^a-z0-9]+', '', (s or '').strip().lower())

    # Welche Spalten dürfen die Kampagne enthalten? (WICHTIG: KEIN nacktes 'url'!)
    UPDATE_CAMPAIGN_HEADERS = {
        "sonarlink", "sonar", "campaign", "campaignurl", "campaignlink", "campaignid"
    }


    MP_TO_LANG = {
        3: "en_GB",      # UK
        4: "de_DE",      # DE
        5: "fr_FR",      # FR
        35691: "it_IT",  # IT
        44551: "es_ES",  # ES
    }

    def _language_for_mp(self, mp: int | None) -> str | None:
        try:
            return self.MP_TO_LANG.get(int(mp)) if mp is not None else None
        except Exception:
            return None

    def _extract_mp_from_line(self, s: str) -> int | None:
        """
        Erwartet z. B. https://sonar-eu.amazon.com/#/3/campaigns/1416358261
        oder ...#/4/programs/123; liefert MP als int. Reine IDs -> None.
        """
        import re
        if not s:
            return None
        m = re.search(r"#/(\d+)/(?:campaigns|programs)/\d+", s)
        if m:
            try:
                return int(m.group(1))
            except Exception:
                return None
        return None



    def _group_campaigns_by_mp(self, lines: list[str]) -> dict:
        """
        Gibt {mp:int|None: {"items":[str,...], "lang": str|None}} zurück.
        None-Schlüssel = unbekannter MP (z. B. reine ID).
        """
        groups: dict[int | None, dict] = {}
        for line in lines:
            mp = self._extract_mp_from_line(line)
            if mp not in groups:
                groups[mp] = {"items": [], "lang": self._language_for_mp(mp)}
            groups[mp]["items"].append(line)
        return groups



    def _set_busy(self, busy: bool):
        """
        Schaltet einen einfachen Busy-Modus:
        - Wartencursor an/aus
        - (falls vorhanden) Dataset-Buttons kurz deaktivieren
        """
        try:
            # Cursor
            self.root.config(cursor="watch" if busy else "")
            self.root.update_idletasks()

            # Optional ein paar Buttons sperren/entsperren
            widgets = [
                getattr(self, "btn_ds_create", None),
                getattr(self, "btn_ds_edit", None),
                getattr(self, "btn_ds_delete", None),
            ]
            for w in widgets:
                if w and w.winfo_exists():
                    w.configure(state="disabled" if busy else "normal")
        except Exception:
            # Busy-Modus ist rein kosmetisch – im Zweifel stillschweigend weiter
            pass


    def parse_performance_stats(self, message):
        """Parse performance statistics from status message"""
        try:
            stats = {}
            if "Average batch time:" in message:
                stats['average_batch_time'] = message.split(': ')[1].strip()
            elif "Success rate:" in message:
                rate_str = message.split(': ')[1].strip()
                stats['average_success_rate'] = float(rate_str.strip('%')) / 100
            elif "Failed segments:" in message:
                stats['failed_segments'] = int(message.split(': ')[1].strip())
            elif "Average time per segment:" in message:
                stats['avg_segment_time'] = message.split(': ')[1].strip()
            return stats
        except Exception:
            return None

    def format_time(self, seconds):
        """Format time in seconds to readable string"""
        return str(timedelta(seconds=round(seconds)))

    # (bullseye) replaces: def _add_result_file(self, path):
    def _add_result_file(self, path):
        """
        Fügt eine Ergebnisdatei in die Results-Liste (Anzeige=Basename, Speicherung=Vollpfad).
        Akzeptiert zusätzlich versehentlich übergebene Rückgaben im Format (results, out_path)
        und entpackt automatisch den Pfadanteil.
        """
        import os
        from tkinter import messagebox

        try:
            # --- robustes Unwrapping ---
            # Falls ein Tuple wie (results, out_path) reinkommt, den Pfad extrahieren.
            if isinstance(path, tuple):
                for itm in path:
                    if isinstance(itm, (str, bytes, os.PathLike)):
                        path = itm
                        break

            # Falls eine 1-Element-Liste mit Pfad kommt, nimm das Element
            if isinstance(path, list) and len(path) == 1 and isinstance(path[0], (str, bytes, os.PathLike)):
                path = path[0]

            # Ab hier muss es ein Pfad sein
            if not isinstance(path, (str, bytes, os.PathLike)):
                raise TypeError(f"expected a file path, got {type(path).__name__}")

            # In String wandeln, falls bytes/PathLike
            path = os.fspath(path)

            abs_path = os.path.abspath(path)
            # Datei kann evtl. noch nicht existieren (asynchron) – wir tragen sie trotzdem ein.
            self.result_files.append(abs_path)
            self.results_listbox.insert("end", os.path.basename(abs_path))

        except Exception as e:
            messagebox.showerror("Results", f"Could not add result file:\n{e}")



    def _on_result_double_click(self, _event=None):
        """Öffnet die doppelt angeklickte Ergebnisdatei mit dem Standardprogramm (Excel)."""
        try:
            sel = self.results_listbox.curselection()
            if not sel:
                return
            idx = sel[0]
            if 0 <= idx < len(self.result_files):
                self._open_file(self.result_files[idx])
        except Exception as e:
            messagebox.showerror("Open file", f"Could not open the selected file:\n{e}")

    def _open_file(self, path):
        """Plattform-agnostisch Datei öffnen (Windows: os.startfile, macOS: open, Linux: xdg-open)."""
        try:
            if not os.path.exists(path):
                messagebox.showerror("Open file", f"File not found:\n{path}")
                return
            import sys, subprocess
            if os.name == "nt":
                os.startfile(path)  # type: ignore[attr-defined]
            elif sys.platform == "darwin":
                subprocess.Popen(["open", path])
            else:
                subprocess.Popen(["xdg-open", path])
        except Exception as e:
            messagebox.showerror("Open file", f"Could not open:\n{path}\n\n{e}")

    def _guess_new_xlsx(self, since_ts):
        """
        Versucht, die neueste .xlsx-Datei seit 'since_ts' im aktuellen Arbeitsordner zu finden.
        Nutzt als Fallback, wenn das Backend keinen Dateinamen zurückgibt.
        """
        try:
            folder = os.getcwd()
            candidates = []
            for name in os.listdir(folder):
                if name.lower().endswith(".xlsx"):
                    p = os.path.join(folder, name)
                    try:
                        if os.path.getmtime(p) >= (since_ts - 2):  # 2s Toleranz
                            candidates.append(p)
                    except Exception:
                        pass
            if candidates:
                candidates.sort(key=lambda p: os.path.getmtime(p), reverse=True)
                return candidates[0]
        except Exception:
            pass
        return None

    def _coerce_date_yyyy_mm_dd(self, val):
        """
        Nimmt Strings, pandas.Timestamp, datetime/date etc. und gibt 'YYYY-MM-DD' zurück.
        Akzeptiert z.B. '2025-10-01 00:00:00', '01.10.2025', '01/10/2025', '10/01/2025'.
        """
        from datetime import datetime, date
        # pandas.Timestamp?
        try:
            if hasattr(val, "to_pydatetime"):
                return val.to_pydatetime().strftime("%Y-%m-%d")
        except Exception:
            pass
        # datetime / date?
        if isinstance(val, (datetime, date)):
            return val.strftime("%Y-%m-%d")

        s = str(val or "").strip()
        if not s:
            raise ValueError("empty date")

        # Zeitteil abtrennen: 'YYYY-MM-DD 00:00:00' oder ISO 'YYYY-MM-DDTHH:MM:SS'
        s = s.replace("T", " ")
        if " " in s:
            s = s.split(" ", 1)[0].strip()

        # direktes ISO
        try:
            return datetime.strptime(s, "%Y-%m-%d").strftime("%Y-%m-%d")
        except Exception:
            pass

        # alternative Formate (eu/us)
        for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            except Exception:
                continue

        raise ValueError(f"bad date: {val!r}")




    # ------------------- Sonar Import (updated) -------------------

    def _templates_parse_ids_from_link(self, text, default_mp=4):
        """
        Akzeptiert:
          - reine Zahl  → (id, mp)
          - '#/<mp>/campaigns/<id>' → (id, mp)
          - '#/<mp>/programs/<id>'  → (id, mp)
          - sonst: letzte längere Zahl im String als ID
        """
        s = (text or "").strip()
        if not s:
            return None, None

        if s.isdigit():
            return s, int(default_mp)

        m = re.search(r"#/(\d+)/(?:campaigns|programs)/(\d+)", s, re.IGNORECASE)
        if m:
            return m.group(2), int(m.group(1))

        m2 = re.search(r"(\d{6,})\D*$", s)
        if m2:
            return m2.group(1), int(default_mp)

        return None, None





    def _templates_copy_sqlite_readonly(self, src_path):
        tmpdir = tempfile.mkdtemp(prefix="ff_cookies_")
        dst = os.path.join(tmpdir, "cookies.sqlite")
        shutil.copy2(src_path, dst)
        return dst, tmpdir

    def _templates_load_firefox_cookies_for_domain(self, profile_path, domain_suffix):
        """
        Lädt Cookies aus Firefox (cookies.sqlite) für die gegebene Domain.
        Gibt eine RequestsCookieJar zurück.
        """
        cookies_db = os.path.join(profile_path, "cookies.sqlite")
        cleanup_dir = None
        try:
            conn = sqlite3.connect(f"file:{cookies_db}?mode=ro", uri=True)
        except sqlite3.OperationalError:
            copied_path, cleanup_dir = self._templates_copy_sqlite_readonly(cookies_db)
            conn = sqlite3.connect(copied_path)

        jar = requests.cookies.RequestsCookieJar()
        try:
            cur = conn.cursor()
            cur.execute(
                "SELECT name, value, host, path, isSecure FROM moz_cookies WHERE host LIKE ?",
                (f"%{domain_suffix}",)
            )
            for name, value, host, path, isSecure in cur.fetchall():
                jar.set(name, value, domain=host, path=path, secure=bool(isSecure))
        finally:
            conn.close()
            if cleanup_dir and os.path.isdir(cleanup_dir):
                shutil.rmtree(cleanup_dir, ignore_errors=True)
        return jar

    def _templates_build_sonar_session(self, profile_path):
        """
        Erstellt eine requests.Session mit Sonar-Cookies und sinnvollen Headern.
        """
        jar = self._templates_load_firefox_cookies_for_domain(profile_path, "sonar-eu.amazon.com")
        s = requests.Session()
        s.cookies = jar
        s.headers.update({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:140.0) Gecko/20100101 Firefox/140.0",
            "Accept": "application/json, text/plain, */*",
            "Origin": "https://sonar-eu.amazon.com",
            "Referer": "https://sonar-eu.amazon.com/",
            "Connection": "keep-alive",
        })
        return s

    def _templates_cookiejar_from_selenium(self, cookies_list):
        jar = requests.cookies.RequestsCookieJar()
        for c in cookies_list:
            jar.set(
                c.get("name"),
                c.get("value"),
                domain=c.get("domain"),
                path=c.get("path", "/"),
                secure=bool(c.get("secure", False))
            )
        return jar

    def _templates_refresh_cookies_via_selenium(self, profile_path, headless=True):
        """
        Startet (headless) Firefox mit dem Profil, öffnet Sonar und liefert frische Cookies.
        """
        try:
            from selenium import webdriver
            from selenium.webdriver.firefox.service import Service as FxService
            from selenium.webdriver.firefox.options import Options as FxOptions
        except Exception:
            return None  # Selenium nicht verfügbar

        options = FxOptions()
        options.add_argument("-profile")
        options.add_argument(profile_path)
        if headless:
            options.add_argument("--headless")
            options.add_argument("--width=1920")
            options.add_argument("--height=1080")

        service = FxService("geckodriver.exe")  # ggf. anpassen
        driver = None
        try:
            driver = webdriver.Firefox(service=service, options=options)
            driver.get("https://sonar-eu.amazon.com/")
            time.sleep(3)
            return self._templates_cookiejar_from_selenium(driver.get_cookies())
        finally:
            if driver is not None:
                try:
                    driver.quit()
                except Exception:
                    pass

    def _templates_http_get_json(self, session, url, timeout=(5, 30)):
        r = session.get(url, timeout=timeout)
        r.raise_for_status()
        try:
            return r.json()
        except Exception:
            return json.loads(r.text)


    def _templates_import_from_campaign(self, link_or_id, default_marketplace_id: int = 4):
        """
        Importiert Template-Werte aus einer bestehenden Sonar-Campaign (bevorzugt) oder,
        falls nur eine Program-ID/-URL gegeben ist, direkt aus dem Program.
        Nutzt _fetch_json_from_sonar() → robust gegen Login/HTML, inkl. Selenium-Fallback.
        """
        # 1) ID & MP aus Link ziehen
        s = (link_or_id or "").strip()
        cid, mp = self._templates_parse_ids_from_link(s, default_mp=default_marketplace_id)
        if not cid:
            raise ValueError("Konnte keine ID im Link erkennen (Campaign oder Program).")
        mp = int(mp or default_marketplace_id)

        # 2) Firefox-Profil besorgen (für Cookies)
        try:
            from utils import get_firefox_profile
            profile_path = get_firefox_profile()
        except Exception:
            profile_path = None
        if not profile_path:
            raise RuntimeError("Kein Firefox-Profil gefunden. Bitte Firefox einmal öffnen/einloggen.")

        # Hilfsfunktion: Campaign laden (bevorzugt)
        def load_campaign(campaign_id: str):
            requester = (self.profile or {}).get("alias") or "me"
            camp_url = f"https://{SONAR_WEB_DOMAIN}/ajax/campaign/{campaign_id}?marketplaceId={mp}&requester={requester}"
            camp_raw = _fetch_json_from_sonar(camp_url, profile_path)
            return camp_raw.get("campaign") if isinstance(camp_raw, dict) and "campaign" in camp_raw else camp_raw

        # Hilfsfunktion: Program laden
        def load_program(program_id: str):
            prog_url = (
                f"https://{SONAR_WEB_DOMAIN}/ajax/program/{program_id}"
                f"?includeBindleInfo=true&marketplaceId={mp}"
            )
            return _fetch_json_from_sonar(prog_url, profile_path)

        # 3) Versuch: erst Campaign laden
        campaign = None
        program_id = None
        try:
            campaign = load_campaign(cid)
            program_id = str(
                campaign.get("programId")
                or (campaign.get("program") or {}).get("id")
                or ""
            ).strip()
        except Exception:
            campaign = None

        # 4a) Erfolgsweg über Campaign → Program + Zeitfenster berechnen
        if campaign and program_id:
            prog_json = load_program(program_id)

            management_type = campaign.get("managementType", "")
            lob_expr = campaign.get("lobExpression", "")
            business_group_id = (campaign.get("businessGroup") or {}).get("id")
            family_id = (campaign.get("family") or {}).get("id")

            start_midnight = campaign.get("campaignStartDate")
            end_midnight = campaign.get("campaignEndDate")
            st_ms = (campaign.get("sendingTime") or {}).get("start")
            en_ms = (campaign.get("sendingTime") or {}).get("end")

            def offset_minutes(point_ms, midnight_ms):
                if point_ms is None or midnight_ms is None:
                    return None
                return int(round((int(point_ms) - int(midnight_ms)) / 60000.0))

            start_off = offset_minutes(st_ms, start_midnight) or 540
            end_off = offset_minutes(en_ms, end_midnight) or 1260

            # Channel sauber normalisieren
            channel_raw = (
                prog_json.get("channelType")
                or prog_json.get("channel")
                or campaign.get("type")
                or ""
            )
            ch = str(channel_raw or "").upper().replace("-", "_").replace(" ", "_")
            if ch == "MOBILEPUSH":
                ch = "MOBILE_PUSH"
            elif ch not in ("MOBILE_PUSH", "EMAIL"):
                ch = ""  # unbekannt -> leer lassen
            channel = ch

            team_bindle = prog_json.get("teamBindle") or ""

            return {
                "channel": channel,
                "managementType": management_type,
                "teamBindle": team_bindle,
                "lobExpression": lob_expr,
                "businessGroupId": str(business_group_id or ""),
                "familyId": str(family_id or ""),
                "optOuts": [],
                "startTimeMinutesOffset": int(start_off),
                "endTimeMinutesOffset": int(end_off),
                "templateId": (campaign.get("template") or {}).get("id"),
                "marketplaceId": mp,
                "programId": program_id,
                "campaignId": str(campaign.get("id") or ""),
            }

        # 4b) Fallback: behandle die ID als Program-ID
        prog_json = load_program(cid)

        channel = prog_json.get("channelType") or prog_json.get("channel") or ""
        team_bindle = prog_json.get("teamBindle") or ""
        management_type = prog_json.get("managementType") or ""
        lob_expr = prog_json.get("lobExpression") or ""
        business_group_id = prog_json.get("businessGroupId")
        family_id = prog_json.get("familyId")

        # Ohne Campaign kennen wir kein Tages-Zeitfenster → Defaults
        return {
            "channel": channel,
            "managementType": str(management_type or ""),
            "teamBindle": str(team_bindle or ""),
            "lobExpression": str(lob_expr or ""),
            "businessGroupId": str(business_group_id or ""),
            "familyId": str(family_id or ""),
            "optOuts": [],
            "startTimeMinutesOffset": 540,
            "endTimeMinutesOffset": 1260,
            "marketplaceId": mp,
            "programId": str(prog_json.get("id") or ""),
            "campaignId": "",
        }





   


def main():
    root = tk.Tk()
    root.withdraw()  # hide window until preflight passes

    if not run_startup_preflight(root):
        try:
            root.destroy()
        except Exception:
            pass
        sys.exit(0)

    app = BullseyeApp(root)
    root.deiconify()  # show window after preflight
    root.mainloop()



if __name__ == "__main__":
    main()
